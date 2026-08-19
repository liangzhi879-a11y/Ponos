"""Excel 应用内编辑：结构读取 + 单元格写回（保留格式，公式格只读）。

用法:
  sheet_edit.py read  <path>          # 输出 {"ok":true,"sheets":[{name,rows,formulas}]}
  sheet_edit.py write <jsonPath>      # 入参 {"path","sheet"?,"updates":[{row,col,value}]}

- read: openpyxl 双加载取 active sheet 值 + 公式标记（'=' 开头格）
- write: 非只读加载 -> 非公式格赋值 -> save（字体/边框/填充/列宽保留）
- .xls 旧格式: xlrd 读 + xlutils.copy 写回（需 xlrd<2.0 才支持格式化保留，否则明确报错）
"""
import sys, json, os


def read_xlsx(path):
    from openpyxl import load_workbook
    wb_val = load_workbook(path, data_only=True)
    wb_f = load_workbook(path, data_only=False)
    ws_val = wb_val.active
    ws_f = wb_f.active
    rows, formulas = [], []
    for r in ws_val.iter_rows():
        row = []
        frow = []
        for cell in r:
            row.append(cell.value)
            v = ws_f.cell(row=cell.row, column=cell.column).value
            frow.append(isinstance(v, str) and v.startswith('='))
        rows.append(row)
        formulas.append(frow)
    return [{"name": ws_val.title, "rows": rows, "formulas": formulas}]


def read_xls(path):
    import xlrd
    wb = xlrd.open_workbook(path)
    ws = wb.sheet_by_index(0)
    rows = [[ws.cell_value(r, c) for c in range(ws.ncols)] for r in range(ws.nrows)]
    return [{"name": ws.name, "rows": rows, "formulas": [[False] * len(rows[0]) for _ in rows]}]


def write_xlsx(body):
    from openpyxl import load_workbook
    path = body["path"]
    wb = load_workbook(path)
    sheet_name = body.get("sheet") or ""
    ws = wb[sheet_name] if sheet_name else wb.active
    for u in body.get("updates", []):
        row, col = int(u["row"]), int(u["col"])
        cur = ws.cell(row=row, column=col)
        if isinstance(cur.value, str) and cur.value.startswith('='):
            continue  # 公式格只读，保留原公式
        cur.value = u["value"]
    wb.save(path)
    return {"ok": True}


def write_xls(body):
    path = body["path"]
    try:
        import xlrd, xlutils.copy  # noqa: E402
    except Exception as e:
        return {"ok": False, "error": f".xls 写回需要 xlrd(<2.0) 与 xlutils：{e}；请另存为 .xlsx 后编辑"}
    try:
        rb = xlrd.open_workbook(path, formatting_info=True)
    except Exception:
        return {"ok": False, "error": ".xls 写回需支持 formatting_info 的 xlrd(<2.0)；请另存为 .xlsx 后编辑"}
    wb = xlutils.copy.copy(rb)
    ws = wb.get_sheet(0)
    for u in body.get("updates", []):
        ws.write(int(u["row"]) - 1, int(u["col"]) - 1, u["value"])
    wb.save(path)
    return {"ok": True}


def main():
    mode, arg = sys.argv[1], sys.argv[2]
    if mode == "read":
        ext = os.path.splitext(arg)[1].lower()
        sheets = read_xlsx(arg) if ext == ".xlsx" else read_xls(arg)
        print(json.dumps({"ok": True, "sheets": sheets}))
    else:
        with open(arg, encoding="utf-8") as f:
            body = json.load(f)
        ext = os.path.splitext(body["path"])[1].lower()
        result = write_xlsx(body) if ext == ".xlsx" else write_xls(body)
        print(json.dumps(result))


try:
    main()
except Exception as e:
    print(json.dumps({"ok": False, "error": str(e)}))
