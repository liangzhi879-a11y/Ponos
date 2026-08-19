"""Convert an Excel file (.xls / .xlsx) to an HTML table, printed as JSON.

Single script for both formats: picks the reader library by file extension
(xlrd for legacy .xls, openpyxl for .xlsx). Output contract:
  {"ok": true, "html": "<table>..."}  or  {"ok": false, "error": "..."}
"""
import sys, json, os

def rows_for(path):
    ext = os.path.splitext(path)[1].lower()
    if ext == '.xlsx':
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True, data_only=True)
        try:
            return list(wb.active.iter_rows(values_only=True))
        finally:
            wb.close()
    import xlrd
    wb = xlrd.open_workbook(path)
    ws = wb.sheet_by_index(0)
    return [[ws.cell_value(r, c) for c in range(ws.ncols)] for r in range(ws.nrows)]

try:
    rows = rows_for(sys.argv[1])
    html = '<div style="font-family:system-ui,sans-serif;font-size:13px;padding:12px;overflow:auto;height:100%">'
    html += '<table style="border-collapse:collapse;width:max-content;min-width:100%">'
    def escape(v):
        return str(v if v is not None else '').replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
    for i, row in enumerate(rows):
        html += '<tr>'
        tag = 'th' if i == 0 else 'td'
        for cell in row:
            bg = '#1f2937' if i == 0 else 'transparent'
            fw = '600' if i == 0 else '400'
            fc = '#f3f4f6' if i == 0 else '#d1d5db'
            html += f'<{tag} style="border:1px solid #374151;padding:4px 12px;text-align:left;white-space:nowrap;background:{bg};font-weight:{fw};color:{fc}">{escape(cell)}</{tag}>'
        html += '</tr>'
    html += '</table></div>'
    print(json.dumps({"ok": True, "html": html}))
except Exception as e:
    print(json.dumps({"ok": False, "error": str(e)}))
