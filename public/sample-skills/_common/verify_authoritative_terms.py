#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""权威术语强制核验与自动矫正模块

核心功能：
    1. load_authoritative_terms() - 加载权威术语注册表
    2. scan_and_correct(text) - 扫描文本，自动矫正所有偏离项
    3. verify_tech_field(text) - 核验单个技术领域字段
    4. verify_output_file(filepath) - 核验输出文件（xlsx/docx/txt）
    5. generate_correction_report() - 生成矫正报告

设计原则：
    - 政府申报系统中所有官方术语必须精确匹配，不得有任何词序/连词差异
    - LLM 生成的内容中"与"/"及"/"和"等连词替换虽语义相同，但会导致系统核验失败
    - 本模块在输出前自动扫描并矫正，杜绝此类问题

用法：
    from verify_authoritative_terms import scan_and_correct, verify_tech_field

    # 扫描并矫正文本
    text = "该项目属于新能源及节能领域"
    corrected, corrections = scan_and_correct(text)
    # corrected = "该项目属于新能源与节能领域"
    # corrections = [{"original": "新能源及节能", "corrected": "新能源与节能", "category": "tech_fields_level1"}]

    # 核验技术领域
    result = verify_tech_field("新能源及节能")
    # result = {"valid": False, "corrected": "新能源与节能", "correction_made": True}
"""

import os
import re
import json
import sys
from datetime import datetime
from collections import OrderedDict


# ============================================================
# 常量
# ============================================================

_TERMS_REGISTRY = None
_TERMS_PATH = None
_CORRECTION_LOG = []  # 本次会话矫正记录


def _get_terms_path():
    """获取权威术语注册表路径"""
    global _TERMS_PATH
    if _TERMS_PATH:
        return _TERMS_PATH

    # 优先级：环境变量 > 相对于本文件的路径 > 绝对路径
    env_path = os.environ.get('AUTHORITATIVE_TERMS_PATH', '')
    if env_path and os.path.exists(env_path):
        _TERMS_PATH = env_path
        return _TERMS_PATH

    # 相对于本文件
    script_dir = os.path.dirname(os.path.abspath(__file__))
    local_path = os.path.join(script_dir, 'authoritative_terms.json')
    if os.path.exists(local_path):
        _TERMS_PATH = local_path
        return _TERMS_PATH

    # 绝对路径
    abs_path = os.path.join(
        os.path.dirname(script_dir),
        '_common', 'authoritative_terms.json'
    )
    if os.path.exists(abs_path):
        _TERMS_PATH = abs_path
        return _TERMS_PATH

    return None


# ============================================================
# 加载
# ============================================================

def load_authoritative_terms(force_reload=False):
    """加载权威术语注册表

    Returns:
        dict: 权威术语注册表，加载失败返回 None
    """
    global _TERMS_REGISTRY
    if _TERMS_REGISTRY is not None and not force_reload:
        return _TERMS_REGISTRY

    path = _get_terms_path()
    if not path:
        print('[WARN] authoritative_terms.json 未找到，权威术语核验不可用', file=sys.stderr)
        return None

    try:
        with open(path, 'r', encoding='utf-8') as f:
            _TERMS_REGISTRY = json.load(f)
        return _TERMS_REGISTRY
    except (json.JSONDecodeError, OSError) as e:
        print(f'[ERROR] 加载权威术语注册表失败: {e}', file=sys.stderr)
        return None


def get_mutations_map():
    """获取所有已知的术语变异 → 正确术语的映射表

    合并 common_mutations 和 connective_pattern 自动生成的变异。

    Returns:
        dict: {变异术语: 正确术语}
    """
    registry = load_authoritative_terms()
    if not registry:
        return {}

    mutations = {}

    # 1. 从 common_mutations 加载
    for cat_key in ('tech_fields_level1',):
        cat = registry.get('categories', {}).get(cat_key, {})
        for wrong, correct in cat.get('common_mutations', {}).items():
            # 移除空格后的归一化版本也加入映射
            normalized_wrong = re.sub(r'\s+', '', wrong)
            normalized_correct = re.sub(r'\s+', '', correct)
            if normalized_wrong != normalized_correct:
                mutations[wrong] = correct
                if normalized_wrong not in mutations:
                    mutations[normalized_wrong] = normalized_correct

    # 2. 词级变异：（原词→替换后能否匹配权威列表）
    _build_fuzzy_mutations(mutations)

    return mutations


def _build_fuzzy_mutations(mutations):
    """基于权威术语列表构建模糊匹配映射

    对8大领域名称，自动生成常见的 '与'→'及'/'和' 变异映射。
    """
    registry = load_authoritative_terms()
    if not registry:
        return

    cat = registry.get('categories', {}).get('tech_fields_level1', {})
    terms = cat.get('terms', [])
    connective_pattern = cat.get('connective_pattern', {})

    if not connective_pattern:
        return

    official_conn = connective_pattern.get('official_connective', '与')
    non_official = connective_pattern.get('non_official_connectives', ['及', '和'])

    for term in terms:
        if official_conn in term:
            for bad_conn in non_official:
                mutated = term.replace(official_conn, bad_conn)
                if mutated != term and mutated not in mutations:
                    mutations[mutated] = term


# ============================================================
# 核验
# ============================================================

def verify_tech_field(text, level=1):
    """核验技术领域字段是否与权威术语一致

    Args:
        text: 待核验的领域名称
        level: 领域级别（1=一级8大领域, 2=二级子领域, 3=三级细分）

    Returns:
        dict: {
            "valid": bool,
            "original": str,
            "corrected": str or None,
            "correction_made": bool,
            "matched_term": str or None,
            "severity": str  # CRITICAL/ERROR/WARNING/OK
        }
    """
    if not text or not str(text).strip():
        return {
            'valid': False, 'original': str(text),
            'corrected': None, 'correction_made': False,
            'matched_term': None, 'severity': 'ERROR'
        }

    original = str(text).strip()

    # 先尝试直接匹配
    registry = load_authoritative_terms()
    if not registry:
        return {
            'valid': True, 'original': original,
            'corrected': None, 'correction_made': False,
            'matched_term': original, 'severity': 'WARNING'
        }

    cat_key = f'tech_fields_level{level}'
    cat = registry.get('categories', {}).get(cat_key, {})

    if level == 1:
        terms = cat.get('terms', [])
        if original in terms:
            return {
                'valid': True, 'original': original,
                'corrected': None, 'correction_made': False,
                'matched_term': original, 'severity': 'OK'
            }

        # 尝试 common_mutations 矫正
        mutations = cat.get('common_mutations', {})
        if original in mutations:
            corrected = mutations[original]
            return {
                'valid': False, 'original': original,
                'corrected': corrected, 'correction_made': True,
                'matched_term': corrected,
                'severity': 'CRITICAL'  # 一级领域错误是一票否决
            }

        # 尝试 connective_pattern 自动矫正
        pattern = cat.get('connective_pattern', {})
        if pattern:
            auto_regex = pattern.get('auto_correct_regex', '')
            replacement = pattern.get('auto_correct_replacement', '')
            if auto_regex and replacement:
                corrected = re.sub(auto_regex, replacement, original)
                if corrected != original and corrected in terms:
                    return {
                        'valid': False, 'original': original,
                        'corrected': corrected, 'correction_made': True,
                        'matched_term': corrected, 'severity': 'CRITICAL'
                    }

        # 未匹配任何权威术语
        return {
            'valid': False, 'original': original,
            'corrected': None, 'correction_made': False,
            'matched_term': None,
            'severity': 'CRITICAL',
            'error': f'"{original}" 不在8大高新技术领域内，且无法自动矫正'
        }

    elif level == 2:
        # 二级领域：需知道其上级一级领域才能校验
        terms_by_l1 = cat.get('terms', {})
        all_terms = []
        for l1_terms in terms_by_l1.values():
            all_terms.extend(l1_terms)

        if original in all_terms:
            return {
                'valid': True, 'original': original,
                'corrected': None, 'correction_made': False,
                'matched_term': original, 'severity': 'OK'
            }

        return {
            'valid': False, 'original': original,
            'corrected': None, 'correction_made': False,
            'matched_term': None,
            'severity': 'ERROR',
            'error': f'"{original}" 不在已知的二级技术领域内'
        }

    return {
        'valid': True, 'original': original,
        'corrected': None, 'correction_made': False,
        'matched_term': original, 'severity': 'WARNING'
    }


def scan_and_correct(text, auto_correct=True, project_root=None):
    """双层核验：先溯源（provenance），再静态术语库兜底

    Layer 1: provenance — 以官方源文件提取值为准（优先级最高）
    Layer 2: authoritative_terms.json — 静态术语库兜底

    Args:
        text: 待扫描的文本
        auto_correct: 是否自动矫正（默认 True）
        project_root: 项目根目录（用于读取provenance）

    Returns:
        tuple: (corrected_text, corrections_list)
    """
    global _CORRECTION_LOG

    if not text or not isinstance(text, str):
        return text, []

    corrected = text
    all_corrections = []

    # Layer 1: provenance 溯源核验（v2.7新增）
    try:
        from provenance_manager import verify_against_provenance
        corrected, prov_corrections = verify_against_provenance(corrected, project_root)
        for c in prov_corrections:
            c['layer'] = 'provenance'
        all_corrections.extend(prov_corrections)
    except ImportError:
        pass

    # Layer 2: 静态权威术语库兜底
    mutations = get_mutations_map()
    if mutations:
        # 按长度降序排列变异（优先匹配长词）
        sorted_mutations = sorted(mutations.items(), key=lambda x: len(x[0]), reverse=True)
        for wrong, correct in sorted_mutations:
            if wrong in corrected:
                pos = corrected.find(wrong)
                corrected = corrected.replace(wrong, correct)
                correction = {
                    'original': wrong,
                    'corrected': correct,
                    'category': 'tech_fields_level1',
                    'position': pos,
                    'timestamp': datetime.now().isoformat(timespec='seconds'),
                    'layer': 'static_terms',
                }
                all_corrections.append(correction)
                _CORRECTION_LOG.append(correction)

    return corrected, all_corrections


def verify_output_file(filepath):
    """核验输出文件中的权威术语

    支持 xlsx、docx、txt 格式。
    扫描所有文本内容，检测术语偏离并自动矫正。

    Args:
        filepath: 输出文件路径

    Returns:
        dict: {
            "file": str,
            "verified": bool,
            "corrections_made": int,
            "corrections": [...],
            "errors": [...]
        }
    """
    if not os.path.exists(filepath):
        return {
            'file': filepath, 'verified': False,
            'corrections_made': 0, 'corrections': [],
            'errors': [f'文件不存在: {filepath}']
        }

    ext = os.path.splitext(filepath)[1].lower()
    all_corrections = []
    errors = []

    try:
        if ext == '.txt' or ext == '.json' or ext == '.md':
            with open(filepath, 'r', encoding='utf-8') as f:
                original = f.read()
            corrected, corrections = scan_and_correct(original)
            if corrections and corrected != original:
                with open(filepath, 'w', encoding='utf-8') as f:
                    f.write(corrected)
            all_corrections.extend(corrections)

        elif ext == '.xlsx':
            try:
                import openpyxl
            except ImportError:
                errors.append('openpyxl 未安装，跳过 xlsx 核验')
                return {
                    'file': filepath, 'verified': False,
                    'corrections_made': 0, 'corrections': [],
                    'errors': errors
                }

            wb = openpyxl.load_workbook(filepath)
            modified = False
            for ws_name in wb.sheetnames:
                ws = wb[ws_name]
                for row in range(1, ws.max_row + 1):
                    for col in range(1, ws.max_column + 1):
                        cell = ws.cell(row, col)
                        if cell.value and isinstance(cell.value, str):
                            corrected_text, cell_corrections = scan_and_correct(cell.value)
                            if cell_corrections:
                                cell.value = corrected_text
                                for c in cell_corrections:
                                    c['location'] = f'{ws_name}!{openpyxl.utils.get_column_letter(col)}{row}'
                                all_corrections.extend(cell_corrections)
                                modified = True
            if modified:
                wb.save(filepath)
            wb.close()

        elif ext == '.docx':
            errors.append('docx 核验暂不支持（TODO）')

    except Exception as e:
        errors.append(f'核验异常: {type(e).__name__}: {e}')

    # 写入矫正日志
    if all_corrections:
        _write_correction_log(filepath, all_corrections)

    return {
        'file': filepath,
        'verified': len(errors) == 0,
        'corrections_made': len(all_corrections),
        'corrections': all_corrections,
        'errors': errors,
    }


def _write_correction_log(source_file, corrections):
    """写入矫正日志"""
    registry = load_authoritative_terms()
    if not registry:
        return

    rules = registry.get('verification_rules', {})
    log_config = rules.get('correction_log', {})
    if not log_config.get('enabled', True):
        return

    log_path = log_config.get('output_path', '.trae/logs/term_corrections.log')
    log_dir = os.path.dirname(log_path)
    if log_dir and not os.path.exists(log_dir):
        os.makedirs(log_dir, exist_ok=True)

    try:
        with open(log_path, 'a', encoding='utf-8') as f:
            for c in corrections:
                ts = c.get('timestamp', datetime.now().isoformat(timespec='seconds'))
                cat = c.get('category', 'unknown')
                orig = c.get('original', '')
                corrected = c.get('corrected', '')
                line = f'{ts} | {cat} | {orig} → {corrected} | {source_file}\n'
                f.write(line)
    except OSError:
        pass  # 日志写入失败不阻塞主流程


def generate_correction_report():
    """生成本次会话的矫正汇总报告

    Returns:
        dict: {
            "total_corrections": int,
            "by_category": {category: count},
            "details": [...],
            "session_time": str
        }
    """
    global _CORRECTION_LOG
    by_cat = {}
    for c in _CORRECTION_LOG:
        cat = c.get('category', 'unknown')
        by_cat[cat] = by_cat.get(cat, 0) + 1

    return {
        'total_corrections': len(_CORRECTION_LOG),
        'by_category': by_cat,
        'details': _CORRECTION_LOG.copy(),
        'session_time': datetime.now().isoformat(timespec='seconds'),
    }


def clear_correction_log():
    """清空本次会话矫正记录"""
    global _CORRECTION_LOG
    _CORRECTION_LOG = []


# ============================================================
# 便捷函数（供技能直接调用）
# ============================================================

def validate_tech_field_policy_compliance(tech_field_l1, tech_field_l2=None, tech_field_l3=None):
    """全面核验技术领域三级的政策合规性

    Args:
        tech_field_l1: 一级领域名称
        tech_field_l2: 二级领域名称（可选）
        tech_field_l3: 三级领域名称（可选）

    Returns:
        dict: {
            "passed": bool,
            "level1": verify_tech_field 结果,
            "level2": verify_tech_field 结果 (如有),
            "level3": ...,
            "corrections": [...],
            "fatal": bool  # True=一级领域错误，一票否决
        }
    """
    results = {
        'passed': True,
        'corrections': [],
        'fatal': False,
    }

    if tech_field_l1:
        l1_result = verify_tech_field(tech_field_l1, level=1)
        results['level1'] = l1_result
        if l1_result['correction_made']:
            results['corrections'].append({
                'level': 1, 'original': l1_result['original'],
                'corrected': l1_result['corrected'],
            })
        if l1_result['severity'] == 'CRITICAL':
            results['passed'] = False
            results['fatal'] = True

    if tech_field_l2:
        l2_result = verify_tech_field(tech_field_l2, level=2)
        results['level2'] = l2_result
        if l2_result['correction_made']:
            results['corrections'].append({
                'level': 2, 'original': l2_result['original'],
                'corrected': l2_result['corrected'],
            })
        if not l2_result['valid']:
            results['passed'] = False

    if tech_field_l3:
        l3_result = verify_tech_field(tech_field_l3, level=3)
        results['level3'] = l3_result

    return results


# ============================================================
# CLI
# ============================================================

def main():
    """CLI 入口"""
    import argparse
    parser = argparse.ArgumentParser(
        description='权威术语强制核验与自动矫正',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog='''示例：
  python verify_authoritative_terms.py --text "新能源及节能"
  python verify_authoritative_terms.py --file "output.xlsx"
  python verify_authoritative_terms.py --scan-text "该项目属于新能源及节能领域，采用先进制造及自动化技术"
''',
    )
    parser.add_argument('--text', help='核验单个技术领域文本')
    parser.add_argument('--level', type=int, default=1, help='领域级别（1/2/3，默认1）')
    parser.add_argument('--file', help='核验输出文件（xlsx/txt）')
    parser.add_argument('--scan-text', help='扫描并矫正长文本')
    args = parser.parse_args()

    if args.text:
        result = verify_tech_field(args.text, level=args.level)
        print(json.dumps(result, ensure_ascii=False, indent=2))
        sys.exit(0 if result['valid'] else 1)

    elif args.file:
        result = verify_output_file(args.file)
        print(json.dumps(result, ensure_ascii=False, indent=2))
        sys.exit(0 if result['verified'] else 1)

    elif args.scan_text:
        corrected, corrections = scan_and_correct(args.scan_text)
        print(f'原始文本：{args.scan_text}')
        print(f'矫正文本：{corrected}')
        print(f'矫正项：{corrections}')
        sys.exit(0 if not corrections else 1)

    else:
        parser.print_help()
        sys.exit(1)


if __name__ == '__main__':
    main()
