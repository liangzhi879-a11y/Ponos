"""高新技术产品（PS）材料审核脚本

用途：审核高新技术企业认定申报中的高新技术产品（服务）证明材料是否合规，包括：
  1. 高新产品材料完整性：每个 PS 是否有对应的证明材料
  2. 产品收入/合同/发票校验：是否有关键财务证明
  3. PS 与 IP 关联校验：检查 PS 是否有关联的 IP
  4. 技术领域完整性校验 + 权威术语强制核验（v2.6新增）
  5. 是否主要产品标注校验
  6. PS 编号连续性校验

用法：
  python validate_ps.py --dir "PS材料目录"
  python validate_ps.py --dir "PS材料目录" --project-root "项目根目录"
  python validate_ps.py --file "高新技术产品明细表.xlsx" --project-root "项目根目录"

输出：JSON 格式审核报告，结构如下：
  {
    "passed": bool,
    "errors": [{"file": "...", "row": 0, "field": "...", "reason": "..."}],
    "warnings": [...],
    "stats": {"total_ps": N, "valid_ps": N, "invalid_ps": N, ...},
    "term_corrections": [...]   # v2.6新增：权威术语矫正记录
  }

退出码：审核通过 0，存在错误 1

依赖：openpyxl（用于读取 xlsx 汇总表）；verify_authoritative_terms（权威术语核验）
"""

import os
import re
import sys
import json
import argparse

# v2.6: 权威术语核验模块
try:
    from verify_authoritative_terms import scan_and_correct, verify_tech_field, validate_tech_field_policy_compliance
    _TERM_VERIFY_AVAILABLE = True
except ImportError:
    _TERM_VERIFY_AVAILABLE = False


# ============================================================
# 常量定义
# ============================================================

# 证明材料文件支持的扩展名
PROOF_EXTENSIONS = ('.pdf', '.jpg', '.jpeg', '.png', '.doc', '.docx', '.xls', '.xlsx')

# PS 表模板表头（与 validate_tables.py 中 ps 规格一致）
PS_HEADERS = [
    '产品（服务）编号', '产品（服务）名称',
    '技术领域（一级）', '技术领域（二级）', '技术领域（三级）',
    '技术来源', '上年度销售收入 （万元）', '是否主要产品 （服务）',
    '知识产权编号',
    '关键技术及主要技术指标（限400字）',
    '与同类产品（服务）的竞争优势（限400字）',
    '知识产权获得情况及其对产品（服务）在技术上发挥的支持作用（限400字）',
]

# 是否主要产品有效值
VALID_MAIN_PRODUCT_VALUES = {'是', '否', 'Y', 'N', '√', '×'}

# 技术来源有效值（参考 validate_tables.py 中的校验）
VALID_TECH_SOURCES = {'企业自有技术', '受让', '受赠', '并购', '其他'}

# 8 大高新技术领域名称（一级）- 保留硬编码作为快速回退，但优先使用权威术语核验
TECH_FIELDS_LEVEL1 = {
    '电子信息', '生物与新医药', '航空航天', '新材料',
    '高技术服务', '新能源与节能', '资源与环境', '先进制造与自动化',
}

# PS 编号正则
PS_NUMBER_PATTERN = re.compile(r'PS(\d{1,3})', re.IGNORECASE)
IP_NUMBER_PATTERN = re.compile(r'IP(\d{1,3})', re.IGNORECASE)


def _load_project_index(project_root):
    """加载 project_index.json，用于 PS 与 IP 关联校验"""
    if not project_root:
        return None
    candidates = [
        os.path.join(project_root, '.trae', 'project_knowledge', 'project_index.json'),
        os.path.join(project_root, '.trae', 'project_index.json'),
    ]
    for path in candidates:
        if os.path.exists(path):
            try:
                with open(path, 'r', encoding='utf-8') as f:
                    return json.load(f)
            except (json.JSONDecodeError, OSError):
                return None
    return None


def _extract_ps_number(value):
    """从字符串中提取 PS 编号（如 PS01 → 1）"""
    if value is None:
        return None
    m = PS_NUMBER_PATTERN.search(str(value))
    return int(m.group(1)) if m else None


def _extract_ip_numbers(value):
    """从单元格值中提取所有 IP 编号"""
    if value is None:
        return []
    return [int(m.group(1)) for m in IP_NUMBER_PATTERN.finditer(str(value))]


def _scan_ps_proof_files(proof_dir):
    """扫描 PS 证明材料目录下的文件

    返回：{ps_no: [文件名列表], ...}（按 PS 编号分组）
    """
    proofs = {}
    if not os.path.isdir(proof_dir):
        return proofs
    for entry in sorted(os.listdir(proof_dir)):
        ext = os.path.splitext(entry)[1].lower()
        if ext not in PROOF_EXTENSIONS:
            continue
        # 从文件名提取 PS 编号
        ps_no = _extract_ps_number(entry)
        if ps_no is not None:
            proofs.setdefault(ps_no, []).append(entry)
    return proofs


def _read_ps_table(file_path):
    """读取高新技术产品（服务）明细表

    返回：[{'row': 行号, 'ps_no': 编号, 'name': 名称, 'tech_field_l1': ...,
           'tech_field_l2': ..., 'tech_field_l3': ..., 'tech_source': ...,
           'revenue': 收入, 'is_main': 是否主要产品, 'ip_raw': IP编号原始值,
           'ip': [IP编号列表], ...}, ...]
    """
    try:
        import openpyxl
    except ImportError:
        return None, 'openpyxl 未安装，跳过表格读取'

    if not os.path.exists(file_path):
        return None, f'文件不存在: {file_path}'

    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
    except Exception as e:
        return None, f'加载 Excel 失败: {e}'

    ws = wb.active
    header_row = 1
    # 自动检测表头行
    for r in range(1, min(5, ws.max_row + 1)):
        v = ws.cell(r, 1).value
        if v and '编号' in str(v) and ('产品' in str(v) or 'PS' in str(v).upper()):
            header_row = r
            break

    rows = []
    for r in range(header_row + 1, ws.max_row + 1):
        ps_no_raw = ws.cell(r, 1).value
        # 跳过空行和统计行
        if ps_no_raw is None or str(ps_no_raw).strip() == '':
            continue
        if isinstance(ps_no_raw, str) and ('总计' in ps_no_raw or '合计' in ps_no_raw or '说明' in ps_no_raw):
            continue

        ps_no = _extract_ps_number(ps_no_raw)
        row_data = {
            'row': r,
            'ps_no_raw': ps_no_raw,
            'ps_no': ps_no,
            'name': ws.cell(r, 2).value,
            'tech_field_l1': ws.cell(r, 3).value,
            'tech_field_l2': ws.cell(r, 4).value,
            'tech_field_l3': ws.cell(r, 5).value,
            'tech_source': ws.cell(r, 6).value,
            'revenue': ws.cell(r, 7).value,
            'is_main': ws.cell(r, 8).value,
            'ip_raw': ws.cell(r, 9).value,
            'ip': _extract_ip_numbers(ws.cell(r, 9).value),
        }
        rows.append(row_data)

    return rows, None


def validate_ps_directory(ps_dir, project_root=None):
    """审核高新产品材料目录

    参数：
      ps_dir: PS 证明材料目录路径（含汇总表 + 证明材料）
      project_root: 项目根目录（用于读取 project_index.json）

    返回：审核报告 dict
    """
    errors = []
    warnings = []
    term_corrections = []  # v2.6: 权威术语矫正记录
    stats = {
        'total_ps': 0,
        'valid_ps': 0,
        'invalid_ps': 0,
        'main_product_count': 0,
        'with_ip_relation': 0,
        'without_ip_relation': 0,
        'with_proof_material': 0,
        'without_proof_material': 0,
        'ps_numbers': [],
        'term_corrections_count': 0,  # v2.6
    }

    # ============================================================
    # 校验 1：目录结构校验
    # ============================================================
    if not os.path.exists(ps_dir):
        errors.append({
            'file': ps_dir, 'row': 0, 'field': 'directory',
            'reason': f'PS 证明材料目录不存在: {ps_dir}',
        })
        return {'passed': False, 'errors': errors, 'warnings': warnings, 'stats': stats}

    # 查找 PS 汇总表
    table_file = None
    table_patterns = [r'高新技术产品.*明细.*\.xlsx$', r'高新产品.*明细.*\.xlsx$']
    for entry in os.listdir(ps_dir):
        if not entry.endswith('.xlsx'):
            continue
        for p in table_patterns:
            if re.search(p, entry):
                table_file = os.path.join(ps_dir, entry)
                break
        if table_file:
            break

    # 加载 project_index.json
    project_index = _load_project_index(project_root)

    # 从 project_index 获取 PS-IP 关联
    ps_ip_relations = {}  # ps_no -> [ip_nos]
    if project_index and isinstance(project_index.get('knowledge_graph'), dict):
        edges = project_index['knowledge_graph'].get('edges', []) or []
        for edge in edges:
            if not isinstance(edge, dict):
                continue
            src = str(edge.get('source', ''))
            tgt = str(edge.get('target', ''))
            rel = str(edge.get('relation', ''))
            if rel in ('related_ip', 'ps_ip', 'supports'):
                for ps_str, ip_str in [(src, tgt), (tgt, src)]:
                    ps_m = PS_NUMBER_PATTERN.search(ps_str)
                    ip_m = IP_NUMBER_PATTERN.search(ip_str)
                    if ps_m and ip_m:
                        ps_no = int(ps_m.group(1))
                        ip_no = int(ip_m.group(1))
                        ps_ip_relations.setdefault(ps_no, []).append(ip_no)

    # 扫描证明材料文件
    proofs = _scan_ps_proof_files(ps_dir)

    # ============================================================
    # 校验汇总表
    # ============================================================
    table_rows = []
    if table_file is None:
        warnings.append({
            'file': '', 'row': 0, 'field': 'table',
            'reason': '未找到高新技术产品（服务）明细表（xlsx），跳过表格校验',
        })
    else:
        rows, err = _read_ps_table(table_file)
        if err:
            warnings.append({
                'file': os.path.basename(table_file), 'row': 0, 'field': 'table',
                'reason': err,
            })
        else:
            table_rows = rows or []
            stats['total_ps'] = len(table_rows)

    # ============================================================
    # 逐行校验 PS 记录
    # ============================================================
    ps_nos_found = []
    prev_ps_no = None

    for row_data in table_rows:
        r = row_data['row']
        fname = os.path.basename(table_file) if table_file else ''
        has_error = False

        ps_no = row_data['ps_no']
        if ps_no is not None:
            ps_nos_found.append(ps_no)
            stats['ps_numbers'].append(ps_no)

        # 校验 6：PS 编号连续性校验
        if ps_no is not None:
            if prev_ps_no is not None and ps_no != prev_ps_no + 1:
                if ps_no == prev_ps_no:
                    errors.append({
                        'file': fname, 'row': r, 'field': 'ps_no',
                        'reason': f'第{r}行 PS 编号重复: PS{ps_no:02d}',
                    })
                    has_error = True
                elif ps_no > prev_ps_no + 1:
                    missing = list(range(prev_ps_no + 1, ps_no))
                    warnings.append({
                        'file': fname, 'row': r, 'field': 'ps_no',
                        'reason': f'PS 编号不连续，缺少: {",".join(f"PS{n:02d}" for n in missing)}',
                    })
                elif ps_no < prev_ps_no:
                    warnings.append({
                        'file': fname, 'row': r, 'field': 'ps_no',
                        'reason': f'PS 编号乱序: PS{ps_no:02d} 出现在 PS{prev_ps_no:02d} 之后',
                    })
            prev_ps_no = ps_no

        # 校验 1：证明材料完整性
        if ps_no is not None:
            if ps_no in proofs:
                stats['with_proof_material'] += 1
            else:
                stats['without_proof_material'] += 1
                warnings.append({
                    'file': '', 'row': r, 'field': 'proof_material',
                    'reason': f'第{r}行 PS{ps_no:02d} 未找到对应证明材料文件',
                })

        # 校验 3：PS 与 IP 关联校验
        has_ip_in_table = len(row_data['ip']) > 0
        has_ip_in_index = ps_no is not None and ps_no in ps_ip_relations

        if has_ip_in_table or has_ip_in_index:
            stats['with_ip_relation'] += 1
        else:
            stats['without_ip_relation'] += 1
            errors.append({
                'file': fname, 'row': r, 'field': 'ip_relation',
                'reason': f'第{r}行 PS{ps_no:02d}（{row_data.get("name", "")}）未关联任何 IP',
            })
            has_error = True

        # 校验 4：技术领域完整性校验 + 权威术语强制核验（v2.6）
        tech_l1 = row_data.get('tech_field_l1')
        tech_l2 = row_data.get('tech_field_l2')
        tech_l3 = row_data.get('tech_field_l3')
        if not tech_l1 or str(tech_l1).strip() == '':
            errors.append({
                'file': fname, 'row': r, 'field': 'tech_field_l1',
                'reason': f'第{r}行 技术领域（一级）为空',
            })
            has_error = True
        else:
            l1_str = str(tech_l1).strip()
            # v2.7: 优先使用 provenance 溯源核验 + authoritative_terms 兜底
            if _TERM_VERIFY_AVAILABLE:
                corrected_text, corrections = scan_and_correct(l1_str, project_root=project_root)
                if corrections:
                    for c in corrections:
                        term_corrections.append({
                            'file': fname, 'row': r, 'field': 'tech_field_l1',
                            'original': c.get('original', c.get('found', l1_str)),
                            'corrected': c.get('corrected', c.get('source_value', l1_str)),
                            'severity': c.get('severity', 'ERROR'),
                            'layer': c.get('layer', 'unknown'),
                        })
                    stats['term_corrections_count'] += 1
                # 矫正后仍不在权威列表中则为真错误
                verify_result = verify_tech_field(corrected_text, level=1)
                if not verify_result['valid'] and not verify_result.get('correction_made'):
                    errors.append({
                        'file': fname, 'row': r, 'field': 'tech_field_l1',
                        'reason': f'第{r}行 技术领域（一级）"{l1_str}"不在 8 大高新技术领域内，且无法自动矫正',
                    })
                    has_error = True
            else:
                # 回退：硬编码集合校验
                if l1_str not in TECH_FIELDS_LEVEL1 and not any(
                    f in l1_str for f in TECH_FIELDS_LEVEL1
                ):
                    warnings.append({
                        'file': fname, 'row': r, 'field': 'tech_field_l1',
                        'reason': f'第{r}行 技术领域（一级）"{l1_str}"不在 8 大高新技术领域内',
                    })

        if not tech_l2 or str(tech_l2).strip() == '':
            warnings.append({
                'file': fname, 'row': r, 'field': 'tech_field_l2',
                'reason': f'第{r}行 技术领域（二级）为空',
            })
        else:
            # v2.6: 二级领域也进行权威术语核验
            if _TERM_VERIFY_AVAILABLE:
                l2_str = str(tech_l2).strip()
                verify_result = verify_tech_field(l2_str, level=2)
                if not verify_result['valid']:
                    warnings.append({
                        'file': fname, 'row': r, 'field': 'tech_field_l2',
                        'reason': f'第{r}行 技术领域（二级）"{l2_str}"不在已知的二级技术领域内',
                    })
        if not tech_l3 or str(tech_l3).strip() == '':
            warnings.append({
                'file': fname, 'row': r, 'field': 'tech_field_l3',
                'reason': f'第{r}行 技术领域（三级）为空',
            })

        # 校验 5：是否主要产品标注校验
        is_main = row_data.get('is_main')
        if not is_main or str(is_main).strip() == '':
            errors.append({
                'file': fname, 'row': r, 'field': 'is_main',
                'reason': f'第{r}行 是否主要产品（服务）为空',
            })
            has_error = True
        else:
            is_main_str = str(is_main).strip()
            if is_main_str in ('是', 'Y', '√', 'y'):
                stats['main_product_count'] += 1
            elif is_main_str not in VALID_MAIN_PRODUCT_VALUES:
                warnings.append({
                    'file': fname, 'row': r, 'field': 'is_main',
                    'reason': f'第{r}行 是否主要产品值"{is_main_str}"不规范（应为 是/否）',
                })

        # 校验 2：产品收入校验（关键财务证明）
        revenue = row_data.get('revenue')
        if revenue is None or str(revenue).strip() == '':
            warnings.append({
                'file': fname, 'row': r, 'field': 'revenue',
                'reason': f'第{r}行 上年度销售收入为空',
            })
        else:
            try:
                rev_val = float(revenue)
                if rev_val <= 0:
                    warnings.append({
                        'file': fname, 'row': r, 'field': 'revenue',
                        'reason': f'第{r}行 上年度销售收入 ≤ 0',
                    })
            except (ValueError, TypeError):
                warnings.append({
                    'file': fname, 'row': r, 'field': 'revenue',
                    'reason': f'第{r}行 上年度销售收入非数值: {revenue}',
                })

        if has_error:
            stats['invalid_ps'] += 1
        else:
            stats['valid_ps'] += 1

    # ============================================================
    # 校验 2 补充：产品收入/合同/发票校验（检查证明材料目录）
    # ============================================================
    all_files = []
    if os.path.isdir(ps_dir):
        all_files = os.listdir(ps_dir)

    # 检查是否有关键财务证明文件（合同/发票）
    has_contract = any(
        re.search(r'合同|contract', f, re.IGNORECASE) for f in all_files
    )
    has_invoice = any(
        re.search(r'发票|invoice', f, re.IGNORECASE) for f in all_files
    )
    stats['has_contract'] = has_contract
    stats['has_invoice'] = has_invoice

    if table_rows:  # 有汇总表时才校验财务证明
        if not has_contract:
            warnings.append({
                'file': '', 'row': 0, 'field': 'contract',
                'reason': 'PS 材料目录中未发现销售合同文件',
            })
        if not has_invoice:
            warnings.append({
                'file': '', 'row': 0, 'field': 'invoice',
                'reason': 'PS 材料目录中未发现销售发票文件',
            })

    # 整体校验：至少应有 1 项主要产品
    if table_rows and stats['main_product_count'] == 0:
        errors.append({
            'file': '', 'row': 0, 'field': 'main_product',
            'reason': '所有 PS 均未标注为"主要产品"，至少应有 1 项主要产品（服务）',
        })

    passed = len(errors) == 0
    return {
        'passed': passed,
        'errors': errors,
        'warnings': warnings,
        'stats': stats,
        'term_corrections': term_corrections,  # v2.6新增
    }


def validate_single_ps_file(file_path, project_root=None):
    """审核单个 PS 汇总表文件"""
    errors = []
    warnings = []
    stats = {
        'total_ps': 0,
        'valid_ps': 0,
        'invalid_ps': 0,
        'main_product_count': 0,
        'with_ip_relation': 0,
        'without_ip_relation': 0,
    }

    if not os.path.exists(file_path):
        errors.append({
            'file': file_path, 'row': 0, 'field': 'file',
            'reason': f'文件不存在: {file_path}',
        })
        return {'passed': False, 'errors': errors, 'warnings': warnings, 'stats': stats}

    rows, err = _read_ps_table(file_path)
    if err:
        errors.append({
            'file': os.path.basename(file_path), 'row': 0, 'field': 'table',
            'reason': err,
        })
        return {'passed': False, 'errors': errors, 'warnings': warnings, 'stats': stats}

    rows = rows or []
    stats['total_ps'] = len(rows)
    fname = os.path.basename(file_path)

    project_index = _load_project_index(project_root)
    ps_ip_relations = {}
    if project_index and isinstance(project_index.get('knowledge_graph'), dict):
        edges = project_index['knowledge_graph'].get('edges', []) or []
        for edge in edges:
            if not isinstance(edge, dict):
                continue
            src = str(edge.get('source', ''))
            tgt = str(edge.get('target', ''))
            rel = str(edge.get('relation', ''))
            if rel in ('related_ip', 'ps_ip', 'supports'):
                for ps_str, ip_str in [(src, tgt), (tgt, src)]:
                    ps_m = PS_NUMBER_PATTERN.search(ps_str)
                    ip_m = IP_NUMBER_PATTERN.search(ip_str)
                    if ps_m and ip_m:
                        ps_no = int(ps_m.group(1))
                        ps_ip_relations.setdefault(ps_no, []).append(int(ip_m.group(1)))

    for row_data in rows:
        r = row_data['row']
        has_error = False
        ps_no = row_data['ps_no']

        # PS-IP 关联
        has_ip = len(row_data['ip']) > 0 or (ps_no is not None and ps_no in ps_ip_relations)
        if has_ip:
            stats['with_ip_relation'] += 1
        else:
            stats['without_ip_relation'] += 1
            errors.append({
                'file': fname, 'row': r, 'field': 'ip_relation',
                'reason': f'第{r}行 PS{ps_no:02d if ps_no else "??"} 未关联任何 IP',
            })
            has_error = True

        # 技术领域
        if not row_data.get('tech_field_l1') or str(row_data.get('tech_field_l1', '')).strip() == '':
            errors.append({
                'file': fname, 'row': r, 'field': 'tech_field_l1',
                'reason': f'第{r}行 技术领域（一级）为空',
            })
            has_error = True

        # 是否主要产品
        is_main = row_data.get('is_main')
        if not is_main or str(is_main).strip() == '':
            errors.append({
                'file': fname, 'row': r, 'field': 'is_main',
                'reason': f'第{r}行 是否主要产品为空',
            })
            has_error = True
        elif str(is_main).strip() in ('是', 'Y', '√', 'y'):
            stats['main_product_count'] += 1

        if has_error:
            stats['invalid_ps'] += 1
        else:
            stats['valid_ps'] += 1

    if rows and stats['main_product_count'] == 0:
        errors.append({
            'file': fname, 'row': 0, 'field': 'main_product',
            'reason': '所有 PS 均未标注为"主要产品"',
        })

    passed = len(errors) == 0
    return {
        'passed': passed,
        'errors': errors,
        'warnings': warnings,
        'stats': stats,
    }


def main():
    """CLI 入口：解析参数并执行高新产品材料审核"""
    parser = argparse.ArgumentParser(
        description='高新技术产品（PS）材料审核脚本 - 校验高企认定 PS 证明材料合规性',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog='''示例：
  python validate_ps.py --dir "高新产品证明材料目录"
  python validate_ps.py --dir "高新产品证明材料目录" --project-root "项目根目录"
  python validate_ps.py --file "高新技术产品明细表.xlsx" --project-root "项目根目录"
''',
    )
    parser.add_argument('--dir', help='PS 证明材料所在目录')
    parser.add_argument('--file', help='单个 PS 汇总表 xlsx 文件路径')
    parser.add_argument('--project-root', dest='project_root',
                        help='项目根目录（用于读取 project_index.json 进行关联校验）')
    args = parser.parse_args()

    if not args.dir and not args.file:
        parser.print_help()
        sys.exit(1)

    if args.dir:
        report = validate_ps_directory(args.dir, args.project_root)
    else:
        report = validate_single_ps_file(args.file, args.project_root)

    # 输出 JSON 格式审核报告
    print(json.dumps(report, ensure_ascii=False, indent=2, default=str))
    sys.exit(0 if report['passed'] else 1)


if __name__ == '__main__':
    main()
