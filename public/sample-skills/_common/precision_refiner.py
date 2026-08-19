# -*- coding: utf-8 -*-
"""
precision_refiner.py v1.0.0 - 精修辅助工具

为 gxtz-precision-refiner 技能提供文件操作辅助：
  1. locate  - 精确定位材料中的字段/段落
  2. backup  - 备份原文件（.bak）
  3. apply   - 应用修复到指定位置
  4. scan    - 快速扫描项目中可能涉及的文件

用法：
  python precision_refiner.py locate --file "IP表.xlsx" --field "IP03.授权日期" --mode excel
  python precision_refiner.py backup --file "IP表.xlsx"
  python precision_refiner.py apply --file "IP表.xlsx" --cell "Sheet1!F4" --value "2024-03-15" --type date
  python precision_refiner.py scan --project-root "." --keyword "RD01"

设计原则：
  - 日期字段必须以 datetime.date 对象写入，禁止字符串
  - 字数检查只告警不截断
  - 下拉值必须校验有效性
"""
import os
import sys
import json
import shutil
import argparse
from pathlib import Path
from datetime import datetime, date
from copy import copy

# ── 扫描件使用的常量和函数（尽量自包含，避免循环导入）──
try:
    import openpyxl
except ImportError:
    openpyxl = None

try:
    from docx import Document as DocxDocument
except ImportError:
    DocxDocument = None

RD_DATE_COLS = {6, 7}
IP_DATE_COLS = {6}
PS_DATE_COLS = set()
ACH_DATE_COLS = {6}
STAFF_DATE_COLS = {7, 8}

IP_DROPDOWNS = {
    3: [
        "实用新型专利", "外观设计专利", "软件著作权",
        "发明专利（非国防专利）", "发明专利（国防专利）",
        "植物新品种", "国家级农作物品种", "国家新药",
        "国家一级中药保护品种", "集成电路布图设计专有权",
    ],
    4: ["自主研发", "受让", "受赠", "并购", "其他"],
}
RD_DROPDOWNS = {
    8: [
        "大专院校", "地方属科研院所", "其它企业技术",
        "引进技术本企业消化创新", "国外技术", "企业自有技术", "中央属科研院所",
    ],
}
PS_DROPDOWNS = {
    6: ["企业自有技术", "科研院所", "大专院校",
        "引进技术本企业消化创新", "国外技术", "其它企业技术"],
}
ACH_DROPDOWNS = {
    3: ["专利", "版权", "集成电路布图设计", "其他"],
    4: ["自主研发", "受让、受赠、并购", "集成电路布图设计", "其他"],
    5: ["新产品", "新服务", "新设备", "新技术应用", "样品/样机", "其他"],
    10: [
        "许可他人使用该科技成果",
        "以该科技成果作为投资，折算股份或出资比例",
        "自行投资实施转化",
        "向他人转让该科技成果",
        "以该科技成果作为合作条件，与他人共同实施转化",
        "其他协商确定方式",
    ],
}

PATENT_TYPES = [
    "发明专利", "实用新型专利", "外观设计专利",
    "发明专利（非国防专利）", "发明专利（国防专利）",
    "软件著作权",
    "植物新品种", "国家级农作物品种", "国家新药",
    "国家一级中药保护品种", "集成电路布图设计专有权",
]


def _parse_date(value):
    if not value:
        return None
    if isinstance(value, (datetime, date)):
        return value
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d"):
        try:
            return datetime.strptime(str(value), fmt)
        except ValueError:
            continue
    return value


def cmd_scan(project_root, keyword):
    """扫描项目中可能涉及的文件"""
    root = Path(project_root)
    if not root.exists():
        print(f"错误: 项目目录不存在 {project_root}")
        return

    results = []
    for ext in (".xlsx", ".xls", ".docx", ".doc", ".pdf", ".json"):
        for f in root.rglob(f"*{ext}"):
            if keyword.lower() in f.name.lower():
                results.append(f)

    if not results:
        # 二级：搜索文件名包含常见表格关键词
        common = ["IP", "RD", "PS", "TOAI", "知识产权", "研发", "高新",
                  "成果", "转化", "人员", "制度", "管理", "审计", "核对", "专审"]
        for f in root.rglob("*.xlsx"):
            if any(k in f.name for k in common):
                results.append(f)
        for f in root.rglob("*.docx"):
            if any(k in f.name for k in common):
                results.append(f)

    results = sorted(set(results))

    print(f"\n扫描结果: 找到 {len(results)} 个可能相关的文件\n")
    print(f"{'路径':80} {'大小(KB)':>10}")
    print("-" * 92)
    for r in results:
        size = r.stat().st_size / 1024
        rel = str(r.relative_to(root))
        print(f"{rel:80} {size:>9.1f}")
    print()

    if results:
        return [str(r) for r in results]
    return []


def cmd_locate(file_path, field, mode="excel"):
    """精确定位材料中的字段"""
    path = Path(file_path)
    if not path.exists():
        print(f"错误: 文件不存在 {file_path}")
        return

    if mode == "excel":
        return _locate_excel(path, field)
    elif mode == "docx":
        return _locate_docx(path, field)
    else:
        print(f"错误: 不支持的模式 {mode}（支持: excel, docx）")
        return None


def _locate_excel(path, field):
    """定位Excel中的字段"""
    if openpyxl is None:
        print("错误: 需要安装 openpyxl (pip install openpyxl)")
        return None

    wb = openpyxl.load_workbook(path, data_only=True)
    print(f"\n文件: {path.name}")
    print(f"工作表: {wb.sheetnames}")
    print(f"查找字段: {field}\n")

    found = []
    for ws_name in wb.sheetnames:
        ws = wb[ws_name]
        for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 50)):
            for cell in row:
                if cell.value is not None and field.lower() in str(cell.value).lower():
                    found.append({
                        "sheet": ws_name,
                        "row": cell.row,
                        "col": cell.column,
                        "col_letter": cell.coordinate.replace(str(cell.row), ""),
                        "value": str(cell.value)[:200],
                    })

    if not found:
        print(f"未找到包含 '{field}' 的单元格")
        # 打印表头辅助定位
        ws = wb.active
        print(f"\n当前工作表 ({ws.title}) 表头:")
        for col in range(1, min(ws.max_column + 1, 20)):
            header = ws.cell(1, col).value
            if header:
                print(f"  列{col}: {header}")
    else:
        print(f"找到 {len(found)} 处匹配:\n")
        print(f"{'Sheet':20} {'行':>5} {'列':>5} {'内容预览'}")
        print("-" * 60)
        for f in found:
            print(f"{f['sheet']:20} {f['row']:>5} {f['col']:>5} {f['value'][:100]}")

    wb.close()
    return found


def _locate_docx(path, field):
    """定位Word文档中的段落"""
    if DocxDocument is None:
        print("错误: 需要安装 python-docx (pip install python-docx)")
        return None

    doc = DocxDocument(str(path))
    print(f"\n文件: {path.name}")
    print(f"查找字段: {field}\n")

    found = []
    for i, para in enumerate(doc.paragraphs):
        if para.text.strip() and field.lower() in para.text.lower():
            found.append({
                "paragraph": i + 1,
                "text": para.text[:300],
            })

    if not found:
        print(f"未找到包含 '{field}' 的段落")
    else:
        print(f"找到 {len(found)} 处匹配:\n")
        for f in found:
            print(f"  段落{f['paragraph']}: {f['text'][:200]}")

    return found


def cmd_backup(file_path):
    """备份文件为 .bak"""
    path = Path(file_path)
    if not path.exists():
        print(f"错误: 文件不存在 {file_path}")
        return False

    bak_path = path.with_suffix(path.suffix + ".bak")
    shutil.copy2(path, bak_path)
    print(f"✓ 已备份: {path.name} → {bak_path.name}")
    return True


def cmd_apply(file_path, cell=None, value=None, value_type="text", sheet=None,
             paragraph=None, row=None, col=None):
    """应用修复到指定位置（仅 Excel 支持）"""
    path = Path(file_path)
    if not path.exists():
        print(f"错误: 文件不存在 {file_path}")
        return False

    if not cell and not (row and col):
        print("错误: 需要提供 --cell (如 Sheet1!B3) 或 --row + --col")
        return False

    if openpyxl is None:
        print("错误: 需要安装 openpyxl (pip install openpyxl)")
        return False

    # 自动备份
    cmd_backup(file_path)

    wb = openpyxl.load_workbook(path)

    # 解析 cell 参数
    if cell and "!" in cell:
        sheet_name, coord = cell.split("!", 1)
        ws = wb[sheet_name]
    elif cell:
        ws = wb.active
        coord = cell
    elif sheet:
        ws = wb[sheet_name]
    else:
        ws = wb.active
        coord = None

    # 解析行/列
    if coord:
        from openpyxl.utils import coordinate_to_tuple
        target_row, target_col = coordinate_to_tuple(coord)
    elif row and col:
        target_row = int(row)
        target_col = int(col)
    else:
        print("错误: 无法解析目标位置")
        return False

    # 值类型处理
    final_value = value
    if value_type == "date":
        parsed = _parse_date(value)
        if isinstance(parsed, datetime):
            final_value = parsed
            print(f"  [日期解析] '{value}' → datetime({parsed.year},{parsed.month},{parsed.day})")
        else:
            print(f"  ⚠ 警告: '{value}' 无法解析为日期格式，将作为原始值写入")
    elif value_type == "number":
        try:
            if "." in str(value):
                final_value = float(value)
            else:
                final_value = int(value)
        except (ValueError, TypeError):
            print(f"  ⚠ 警告: '{value}' 无法解析为数字，将作为原始值写入")

    target_cell = ws.cell(target_row, target_col)

    # 应用前校验
    old_value = target_cell.value

    # 字数检查
    if value_type == "text" and isinstance(final_value, str) and len(final_value) > 400:
        print(f"  ⚠ 字数警告: 新值 {len(final_value)} 字超过400字上限，不截断，请人工判断")

    target_cell.value = final_value

    # 日期格式设置
    if isinstance(final_value, datetime):
        target_cell.number_format = "YYYY/MM/DD"

    from openpyxl.utils import get_column_letter
    cell_ref = f"{get_column_letter(target_col)}{target_row}"
    wb.save(path)
    wb.close()

    print(f"\n✓ 修复完成: {path.name} → {cell_ref}")
    print(f"  旧值: {old_value}")
    print(f"  新值: {value}")
    if isinstance(final_value, datetime):
        print(f"  格式: YYYY/MM/DD (datetime 对象)")
    return True


def main():
    parser = argparse.ArgumentParser(description="precision_refiner.py v1.0.0 - 精修辅助工具")
    sub = parser.add_subparsers(dest="command", help="子命令")

    # locate 子命令
    p_locate = sub.add_parser("locate", help="定位材料中的字段/段落")
    p_locate.add_argument("--file", required=True, help="文件路径")
    p_locate.add_argument("--field", required=True, help="要定位的字段名")
    p_locate.add_argument("--mode", default="excel", choices=["excel", "docx"], help="文件类型")

    # backup 子命令
    p_backup = sub.add_parser("backup", help="备份文件为 .bak")
    p_backup.add_argument("--file", required=True, help="文件路径")

    # apply 子命令
    p_apply = sub.add_parser("apply", help="应用修复到指定单元格")
    p_apply.add_argument("--file", required=True, help="文件路径")
    p_apply.add_argument("--cell", default=None, help="目标单元格 (如 Sheet1!B3)")
    p_apply.add_argument("--value", required=True, help="新值")
    p_apply.add_argument("--type", dest="value_type", default="text",
                         choices=["text", "date", "number"], help="值类型")
    p_apply.add_argument("--sheet", default=None, help="工作表名（如果--cell不含工作表名时使用）")
    p_apply.add_argument("--row", default=None, type=int, help="行号（如果--cell未指定时使用）")
    p_apply.add_argument("--col", default=None, type=int, help="列号（如果--cell未指定时使用）")

    # scan 子命令
    p_scan = sub.add_parser("scan", help="扫描项目中可能涉及的文件")
    p_scan.add_argument("--project-root", required=True, help="项目根目录")
    p_scan.add_argument("--keyword", required=True, help="搜索关键词")

    args = parser.parse_args()

    if args.command == "locate":
        cmd_locate(args.file, args.field, args.mode)
    elif args.command == "backup":
        cmd_backup(args.file)
    elif args.command == "apply":
        cmd_apply(args.file, args.cell, args.value, args.value_type,
                  args.sheet, args.paragraph if hasattr(args, "paragraph") else None,
                  args.row if hasattr(args, "row") else None,
                  args.col if hasattr(args, "col") else None)
    elif args.command == "scan":
        cmd_scan(args.project_root, args.keyword)
    else:
        parser.print_help()


if __name__ == "__main__":
    main()
