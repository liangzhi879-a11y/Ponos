#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
filing_mapper.py - 高新认定材料归档映射器

将企业项目材料按模板结构映射归档，生成缺失清单。

用法：
  python filing_mapper.py map --source "源目录" --template "模板目录" --target "目标目录" --enterprise "企业名称"

设计逻辑：
  1. 扫描模板目录，构建结构模型（含 root_level 标识）
  2. 扫描源目录所有文件，按内容+文件名分类
  3. 将源文件匹配到模板槽位
  4. 创建目标目录结构（镜像模板）
  5. 复制匹配文件到目标，保持命名规范
  6. 生成 _缺失清单.md 报告
"""

import argparse
import fnmatch
import os
import re
import shutil
import sys
from collections import defaultdict, OrderedDict
from datetime import datetime
from pathlib import Path


# ============================================================
# 模板结构定义（基准模板）
# ============================================================
TEMPLATE_STRUCTURE = OrderedDict([
    ("知识产权证书扫描件", {
        "path": "知识产权证书扫描件/",
        "pattern": "IP*.pdf",
        "category": "IP证书",
        "root_level": False,
    }),
    ("立项报告原件", {
        "path": "21-23年研发项目/",
        "pattern": "*.doc",
        "category": "RD立项报告",
        "root_level": False,
    }),
    ("立项报告扫描件", {
        "path": "云充科技立项报告扫描件/",
        "pattern": "RD*.pdf",
        "category": "RD扫描件",
        "root_level": False,
    }),
    ("科技成果转化证明", {
        "path": "科技成果转化证明资料/",
        "pattern": "*.pdf",
        "category": "成果转化",
        "root_level": False,
    }),
    ("PS产品资料", {
        "path": "",
        "pattern": "PS*.pdf",
        "category": "PS产品",
        "root_level": True,
    }),
    ("营业执照", {
        "path": "",
        "pattern": "*营业执照*.pdf",
        "category": "企业资质",
        "root_level": True,
    }),
    ("承诺书", {
        "path": "",
        "pattern": "*承诺书*.pdf",
        "category": "企业资质",
        "root_level": True,
    }),
    ("申请书", {
        "path": "",
        "pattern": "*申请书*.pdf",
        "category": "申请书",
        "root_level": True,
    }),
    ("封皮", {
        "path": "",
        "pattern": "*封皮*.pdf",
        "category": "申请书",
        "root_level": True,
    }),
    ("财务审计报告", {
        "path": "",
        "pattern": "*财务审计报告*.pdf",
        "category": "财务审计",
        "root_level": True,
    }),
    ("纳税申报表", {
        "path": "",
        "pattern": "*企业所得税纳税申报表*.pdf",
        "category": "纳税申报",
        "root_level": True,
    }),
    ("研发费用台账", {
        "path": "",
        "pattern": "*研发费用台账*.xlsx",
        "category": "研发台账",
        "root_level": True,
    }),
    ("专项审计报告", {
        "path": "",
        "pattern": "*专审报告*.pdf",
        "category": "专项审计",
        "root_level": True,
    }),
    ("管理制度证明材料", {
        "path": "",
        "pattern": "*.pdf",
        "category": "管理制度",
        "root_level": True,
    }),
    ("科技人员材料", {
        "path": "",
        "pattern": "*科技人员*.pdf",
        "category": "科技人员",
        "root_level": True,
    }),
    ("销售合同与发票", {
        "path": "",
        "pattern": "*销售合同*.pdf",
        "category": "销售合同发票",
        "root_level": True,
    }),
    ("核心表格", {
        "path": "",
        "pattern": "*.xlsx",
        "category": "核心表格",
        "root_level": True,
    }),
    ("研发设备清单", {
        "path": "",
        "pattern": "*设备清单*.xlsx",
        "category": "设备清单",
        "root_level": True,
    }),
])


# ============================================================
# 分类关键词映射（用于内容/文件名匹配）
# ============================================================
CLASSIFICATION_KEYWORDS = {
    "IP证书": {
        "filename_patterns": [r"^IP\d+"],
        "content_keywords": [
            "发明专利", "实用新型", "外观设计", "专利证书", "知识产权",
            "证书号", "专利号", "专利权", "中华人民共和国国家知识产权局",
        ],
    },
    "RD立项报告": {
        "filename_patterns": [r"^RD\d+", r"研发立项", r"立项报告", r"立项书"],
        "content_keywords": [
            "立项报告", "研发项目", "立项书", "研发立项",
            "研究开发", "项目编号", "项目负责人", "研发目标",
        ],
    },
    "RD扫描件": {
        "filename_patterns": [r"^RD\d+.*\.pdf$"],
        "content_keywords": [
            "立项报告", "研发项目", "立项书",
        ],
    },
    "成果转化": {
        "filename_patterns": [r"成果转化", r"转化证明"],
        "content_keywords": [
            "成果转化", "转化证明", "科技成果",
        ],
    },
    "PS产品": {
        "filename_patterns": [r"^PS\d+", r"高新产品"],
        "content_keywords": [
            "高新技术产品", "高新产品", "PS",
        ],
    },
    "企业资质": {
        "filename_patterns": [],
        "content_keywords": [],
    },
    "申请书": {
        "filename_patterns": [r"申请书", r"封皮"],
        "content_keywords": [
            "高新技术企业认定", "申请书", "企业注册号",
        ],
    },
    "财务审计": {
        "filename_patterns": [r"财务审计", r"审计报告"],
        "content_keywords": [
            "审计报告", "财务报表", "资产负债表", "利润表", "现金流量表",
        ],
    },
    "纳税申报": {
        "filename_patterns": [r"纳税申报", r"企业所得税"],
        "content_keywords": [
            "企业所得税", "纳税申报", "应纳税所得额",
        ],
    },
    "研发台账": {
        "filename_patterns": [r"研发费用台账", r"研发台账", r"辅助账"],
        "content_keywords": [
            "研发费用", "台账", "辅助账",
        ],
    },
    "专项审计": {
        "filename_patterns": [r"专审", r"专项审计", r"高新收入审计"],
        "content_keywords": [
            "专项审计", "研发费用", "高新收入", "专审",
        ],
    },
    "管理制度": {
        "filename_patterns": [r"管理制度", r"研发制度", r"产学研", r"激励奖励",
                              r"^1[2-5].*制度", r"^1[2-5].*管理", r"组织管理",
                              r"研发投入核算", r"科研条件"],
        "content_keywords": [
            "管理制度", "研发制度", "组织管理", "产学研",
            "科技成果转化", "激励奖励", "人才绩效",
        ],
    },
    "科技人员": {
        "filename_patterns": [r"科技人员", r"人员比例", r"社保"],
        "content_keywords": [
            "科技人员", "社保", "学历证书", "花名册",
            "人员名单", "劳动合同",
        ],
    },
    "销售合同发票": {
        "filename_patterns": [r"销售合同", r"发票"],
        "content_keywords": [
            "销售合同", "发票", "购销合同",
        ],
    },
    "核心表格": {
        "filename_patterns": [r"RDPS", r"知识产权表", r"成果转化.*表", r"TO.?AI", r"汇总表", r"明细表"],
        "content_keywords": [
            "研发活动汇总", "知识产权表", "成果转化", "高新产品明细",
        ],
    },
    "设备清单": {
        "filename_patterns": [r"设备清单"],
        "content_keywords": [
            "设备清单", "研发设备",
        ],
    },
}

SPECIFIC_FILENAME_RULES = {
    "营业执照": [r"营业执照"],
    "承诺书": [r"承诺书"],
}


# ============================================================
# 文本提取工具
# ============================================================

def _extract_pdf_text(filepath):
    """尝试从PDF提取文本内容"""
    text = ""
    try:
        import pdfplumber
        with pdfplumber.open(filepath) as pdf:
            for page in pdf.pages[:3]:
                page_text = page.extract_text()
                if page_text:
                    text += page_text + "\n"
    except ImportError:
        pass
    except Exception:
        pass

    if not text:
        try:
            import fitz
            _stderr_fd = 2
            _saved_stderr = os.dup(_stderr_fd)
            _devnull = os.open(os.devnull, os.O_WRONLY)
            os.dup2(_devnull, _stderr_fd)
            os.close(_devnull)
            try:
                doc = fitz.open(filepath)
                for page in doc[:3]:
                    text += page.get_text()
                doc.close()
            finally:
                os.dup2(_saved_stderr, _stderr_fd)
                os.close(_saved_stderr)
        except (ImportError, OSError):
            try:
                import fitz
                doc = fitz.open(filepath)
                for page in doc[:3]:
                    text += page.get_text()
                doc.close()
            except ImportError:
                pass
            except Exception:
                pass
        except Exception:
            pass

    if not text:
        try:
            from PyPDF2 import PdfReader
            reader = PdfReader(filepath)
            for page in reader.pages[:3]:
                page_text = page.extract_text()
                if page_text:
                    text += page_text + "\n"
        except ImportError:
            pass
        except Exception:
            pass

    return text


def _extract_doc_text(filepath):
    """尝试从DOCX提取文本内容"""
    text = ""
    ext = os.path.splitext(filepath)[1].lower()
    if ext in ('.docx',):
        try:
            import docx
            doc = docx.Document(filepath)
            for para in doc.paragraphs[:50]:
                if para.text.strip():
                    text += para.text + "\n"
        except ImportError:
            pass
        except Exception:
            pass
    elif ext in ('.doc',):
        try:
            import subprocess
            import tempfile
            result = subprocess.run(
                ['antiword', filepath],
                capture_output=True, text=True, timeout=10,
            )
            if result.returncode == 0:
                text = result.stdout
        except Exception:
            pass
    return text


def _extract_xlsx_text(filepath):
    """尝试从XLSX提取工作表名和内容"""
    text = ""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        text += " ".join(wb.sheetnames) + "\n"
        for sheet_name in wb.sheetnames[:3]:
            ws = wb[sheet_name]
            row_count = 0
            for row in ws.iter_rows(values_only=True):
                row_text = " ".join(str(c) for c in row if c is not None)
                if row_text.strip():
                    text += row_text + "\n"
                row_count += 1
                if row_count > 20:
                    break
        wb.close()
    except ImportError:
        pass
    except Exception:
        pass
    return text


EXTRACTORS = {
    '.pdf': _extract_pdf_text,
    '.doc': _extract_doc_text,
    '.docx': _extract_doc_text,
    '.xlsx': _extract_xlsx_text,
    '.xls': _extract_xlsx_text,
}


# ============================================================
# 模板扫描
# ============================================================

def scan_template(template_dir):
    """扫描模板目录，构建结构模型

    Returns:
        dict: {
            category_name: {
                "path": str,           # 模板中的相对路径
                "pattern": str,        # glob 匹配模式
                "category": str,       # 分类名
                "root_level": bool,    # 是否在模板根目录
                "existing_files": [],  # 模板中已存在的文件列表
                "expected_count": int, # 模板中该槽位的文件数量
            }
        }
    """
    template_path = Path(template_dir)
    if not template_path.exists():
        print(f"[ERROR] 模板目录不存在: {template_dir}")
        sys.exit(1)

    structure = OrderedDict()

    for slot_name, slot_config in TEMPLATE_STRUCTURE.items():
        entry = dict(slot_config)
        entry["existing_files"] = []
        entry["expected_count"] = 0
        entry["existing_names"] = set()

        if slot_config.get("root_level", False):
            search_dir = template_path
        else:
            search_dir = template_path / slot_config["path"]

        if search_dir.exists():
            pattern = slot_config["pattern"]
            matched = sorted(search_dir.glob(pattern))
            if slot_config.get("root_level", False):
                matched = [f for f in matched if f.is_file()]
            else:
                matched = [f for f in matched if f.is_file()]
            entry["existing_files"] = [str(m) for m in matched]
            entry["existing_names"] = {m.name for m in matched}
            entry["expected_count"] = len(matched)

        structure[slot_name] = entry

    return structure


# ============================================================
# 源文件扫描与分类
# ============================================================

def _is_text_file_content_match(text, keywords):
    """检查文本内容是否匹配关键词"""
    if not text or not keywords:
        return 0
    text_lower = text.lower()
    score = 0
    for kw in keywords:
        if kw.lower() in text_lower:
            score += 1
    return score


def _is_filename_match(filename, patterns):
    """检查文件名是否匹配正则模式（同时检查含扩展名和不含扩展名）"""
    basename = os.path.basename(filename)
    basename_noext = os.path.splitext(basename)[0]
    for pattern in patterns:
        if re.search(pattern, basename, re.IGNORECASE):
            return True
        if re.search(pattern, basename_noext, re.IGNORECASE):
            return True
    return False


def _classify_by_specific_rules(filename):
    """通过精确文件名规则分类"""
    basename = os.path.basename(filename)
    for category, patterns in SPECIFIC_FILENAME_RULES.items():
        for pattern in patterns:
            if re.search(pattern, basename, re.IGNORECASE):
                return category
    return None


def classify_source_file(filepath):
    """分类单个源文件

    分类策略（优先级从高到低）：
    1. 精确文件名规则匹配（营业执照、承诺书等）
    2. 文件名模式匹配 + 扩展名约束（IP01.pdf、RD01.doc、RD01.pdf等）
    3. 内容关键词匹配（读取文件内容）
    4. 文件扩展名+目录名辅助判断
    5. 无法分类 → None

    Returns:
        str or None: 分类名
    """
    filename = os.path.basename(filepath)
    ext = os.path.splitext(filename)[1].lower()

    specific_category = _classify_by_specific_rules(filename)
    if specific_category:
        return specific_category

    file_text = ""
    classifications = []

    for category, config in CLASSIFICATION_KEYWORDS.items():
        score = 0

        if _is_filename_match(filename, config["filename_patterns"]):
            score += 3

        if config["content_keywords"] and ext in EXTRACTORS:
            if not file_text:
                file_text = EXTRACTORS[ext](filepath)
            content_score = _is_text_file_content_match(file_text, config["content_keywords"])
            if content_score > 0:
                score += content_score * 2

        if score > 0:
            classifications.append((score, category))

    classifications.sort(key=lambda x: x[0], reverse=True)

    if classifications:
        classifications = _filter_implausible(classifications, ext)
        if not classifications:
            classifications = [(0, None)]
        top_results = [(s, cat) for s, cat in classifications if s == classifications[0][0]]
        if len(top_results) > 1:
            best_score, best_category = _resolve_tie(top_results, filename, ext, filepath)
        else:
            best_score, best_category = classifications[0]
        if best_score >= 2 and best_category:
            return best_category

    ext_category_map = {
        '.doc': 'RD立项报告',
        '.docx': 'RD立项报告',
    }
    if ext in ext_category_map:
        return ext_category_map[ext]

    category = _classify_by_directory(filepath)
    if category:
        return category

    return None


def _classify_by_directory(filepath):
    """通过文件所在目录名进行兜底分类"""
    dirpath = os.path.dirname(filepath)
    dirname = os.path.basename(dirpath)

    dir_category_map = [
        (r'科技成果转化', '成果转化'),
        (r'知识产权', 'IP证书'),
        (r'立项报告扫描', 'RD扫描件'),
        (r'研发项目', 'RD立项报告'),
        (r'立项报告原件', 'RD立项报告'),
    ]

    for pattern, category in dir_category_map:
        if re.search(pattern, dirname, re.IGNORECASE):
            return category

    return None


def _resolve_tie(top_results, filename, ext, filepath):
    """当多个分类得分相同时，通过扩展名和更精确的模式决出胜者"""
    rd_doc_cat = 'RD立项报告'
    rd_scan_cat = 'RD扫描件'
    ip_cat = 'IP证书'

    candidates = {cat for _, cat in top_results}

    if rd_doc_cat in candidates and rd_scan_cat in candidates:
        if ext in ('.doc', '.docx'):
            return (100, rd_doc_cat)
        elif ext == '.pdf':
            return (100, rd_scan_cat)

    if ip_cat in candidates and rd_scan_cat in candidates:
        basename = os.path.basename(filename)
        if re.match(r'^RD\d+', basename, re.IGNORECASE):
            return (100, rd_scan_cat)
        if re.match(r'^IP\d+', basename, re.IGNORECASE):
            return (100, ip_cat)

    if ip_cat in candidates and '核心表格' in candidates:
        if ext in ('.xlsx', '.xls'):
            return (100, '核心表格')

    return top_results[0]


def _filter_implausible(classifications, ext):
    """过滤不合理的分类组合（如 xlsx 不应分为 IP证书）"""
    spreadsheet_exts = ('.xlsx', '.xls')
    if ext in spreadsheet_exts:
        return [(s, cat) for s, cat in classifications if cat != 'IP证书']
    return classifications


def scan_source(source_dir):
    """扫描源目录，对所有文件分类

    Returns:
        dict: {category: [filepath, ...]}
        list: [(filepath, reason)] 无法分类的文件
    """
    source_path = Path(source_dir)
    if not source_path.exists():
        print(f"[ERROR] 源目录不存在: {source_dir}")
        sys.exit(1)

    categorized = defaultdict(list)
    unclassified = []

    all_files = []
    for root, dirs, files in os.walk(source_dir):
        for f in files:
            if f.startswith('~$') or f.startswith('._'):
                continue
            filepath = os.path.join(root, f)
            all_files.append(filepath)

    total = len(all_files)
    print(f"[INFO] 扫描源目录: {source_dir}")
    print(f"[INFO] 发现 {total} 个文件，正在分类...")

    for i, filepath in enumerate(all_files, 1):
        filename = os.path.basename(filepath)
        ext = os.path.splitext(filename)[1].lower()

        if ext not in ('.pdf', '.doc', '.docx', '.xlsx', '.xls', '.jpg', '.jpeg', '.png'):
            continue

        category = classify_source_file(filepath)

        if category:
            categorized[category].append(filepath)
        else:
            unclassified.append((filepath, "无法确定文件类别"))

        if i % 20 == 0:
            print(f"  进度: {i}/{total}")

    print(f"[INFO] 分类完成: {len(categorized)} 个类别, {sum(len(v) for v in categorized.values())} 个文件已分类, {len(unclassified)} 个未分类")

    return dict(categorized), unclassified


# ============================================================
# 匹配与模板分类细化
# ============================================================

def _match_category_to_slot(category):
    """将通用分类映射到模板槽位"""
    mapping = {
        "IP证书": ["知识产权证书扫描件"],
        "RD立项报告": ["立项报告原件"],
        "RD扫描件": ["立项报告扫描件"],
        "成果转化": ["科技成果转化证明"],
        "PS产品": ["PS产品资料"],
        "企业资质": ["营业执照", "承诺书"],
        "营业执照": ["营业执照"],
        "承诺书": ["承诺书"],
        "申请书": ["申请书", "封皮"],
        "财务审计": ["财务审计报告"],
        "纳税申报": ["纳税申报表"],
        "研发台账": ["研发费用台账"],
        "专项审计": ["专项审计报告"],
        "管理制度": ["管理制度证明材料"],
        "科技人员": ["科技人员材料"],
        "销售合同发票": ["销售合同与发票"],
        "核心表格": ["核心表格"],
        "设备清单": ["研发设备清单"],
    }
    return mapping.get(category, [])


def _refine_match_to_slot(category, filepath, slot_names, template_structure):
    """对于映射到多个槽位的分类，进一步精确匹配"""
    if len(slot_names) == 1:
        return slot_names[0]

    filename = os.path.basename(filepath)
    basename = os.path.splitext(filename)[0]

    for slot_name in slot_names:
        slot = template_structure[slot_name]
        pattern = slot["pattern"]
        if fnmatch.fnmatch(filename, pattern):
            return slot_name

    return slot_names[0]


def match_source_to_template(categorized, template_structure, enterprise_name):
    """将分类后的源文件匹配到模板槽位

    Returns:
        dict: {slot_name: [(src_path, dest_filename), ...]}
    """
    matches = defaultdict(list)
    enterprise_safe = re.sub(r'[\\/:*?"<>|]', '_', enterprise_name)

    for category, file_list in categorized.items():
        slot_names = _match_category_to_slot(category)
        if not slot_names:
            continue

        for filepath in file_list:
            slot_name = _refine_match_to_slot(category, filepath, slot_names, template_structure)
            if not slot_name:
                continue

            src_filename = os.path.basename(filepath)
            slot = template_structure[slot_name]

            dest_filename = _generate_dest_filename(
                src_filename, slot_name, slot, enterprise_safe
            )

            matches[slot_name].append((filepath, dest_filename))

    return dict(matches)


def _generate_dest_filename(src_filename, slot_name, slot_config, enterprise_safe):
    """为目标文件生成规范命名"""
    basename, ext = os.path.splitext(src_filename)
    existing = slot_config.get("existing_names", set())

    ip_match = re.match(r'(IP\d+)', basename, re.IGNORECASE)
    rd_match = re.match(r'(RD\d+)', basename, re.IGNORECASE)
    ps_match = re.match(r'(PS\d+)', basename, re.IGNORECASE)

    if ip_match:
        return f"{ip_match.group(1).upper()}_{enterprise_safe}_知识产权证书{ext}"
    elif rd_match:
        return f"{rd_match.group(1).upper()}_{enterprise_safe}_立项报告{ext}"
    elif ps_match:
        return f"{ps_match.group(1).upper()}_{enterprise_safe}_高新产品{ext}"

    clean_name = re.sub(r'[-_\s]+', '_', basename)
    clean_name = re.sub(r'[\\/:*?"<>|]', '_', clean_name)
    for old_name in existing:
        old_base = os.path.splitext(old_name)[0]
        if old_base in clean_name or clean_name in old_base:
            return f"{old_name}"
    return f"{enterprise_safe}_{clean_name}{ext}"


# ============================================================
# 目录创建与文件复制
# ============================================================

def create_target_structure(template_structure, target_dir):
    """在目标目录创建镜像模板的目录结构"""
    target_path = Path(target_dir)
    target_path.mkdir(parents=True, exist_ok=True)

    created_dirs = []
    for slot_name, slot_config in template_structure.items():
        if slot_config.get("root_level", False):
            dest_dir = target_path / slot_name
        else:
            path_part = slot_config["path"].rstrip("/")
            dest_dir = target_path / path_part

        dest_dir.mkdir(parents=True, exist_ok=True)
        created_dirs.append(str(dest_dir))

    return created_dirs


def copy_matched_files(matches, template_structure, target_dir):
    """将匹配的文件复制到目标目录"""
    target_path = Path(target_dir)
    copied_count = 0
    copy_log = []

    for slot_name, file_list in matches.items():
        slot_config = template_structure[slot_name]
        if slot_config.get("root_level", False):
            dest_dir = target_path / slot_name
        else:
            path_part = slot_config["path"].rstrip("/")
            dest_dir = target_path / path_part

        dest_dir.mkdir(parents=True, exist_ok=True)

        for src_path, dest_filename in file_list:
            dest_path = dest_dir / dest_filename
            try:
                shutil.copy2(src_path, str(dest_path))
                copy_log.append((slot_name, os.path.basename(src_path), str(dest_path), True))
                copied_count += 1
            except Exception as e:
                copy_log.append((slot_name, os.path.basename(src_path), str(dest_path), False, str(e)))

    print(f"[INFO] 已复制 {copied_count} 个文件到目标目录")
    return copy_log


# ============================================================
# 缺失清单报告
# ============================================================

def generate_missing_report(template_structure, matches, unclassified, target_dir, enterprise_name):
    """生成 _缺失清单.md 报告"""
    report_lines = []
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    report_lines.append(f"# 材料归档缺失清单")
    report_lines.append(f"")
    report_lines.append(f"**企业名称**: {enterprise_name}")
    report_lines.append(f"**生成时间**: {now}")
    report_lines.append(f"**目标目录**: {target_dir}")
    report_lines.append(f"")

    # 汇总统计
    total_slots = len(template_structure)
    filled_slots = 0
    total_matched = 0

    for slot_name in template_structure:
        if slot_name in matches and matches[slot_name]:
            filled_slots += 1
            total_matched += len(matches[slot_name])

    report_lines.append(f"## 汇总统计")
    report_lines.append(f"")
    report_lines.append(f"| 指标 | 数值 |")
    report_lines.append(f"|------|------|")
    report_lines.append(f"| 模板槽位总数 | {total_slots} |")
    report_lines.append(f"| 已填充槽位 | {filled_slots} |")
    report_lines.append(f"| 缺失槽位 | {total_slots - filled_slots} |")
    report_lines.append(f"| 已匹配文件数 | {total_matched} |")
    report_lines.append(f"| 填充率 | {filled_slots / total_slots * 100:.1f}% |")
    report_lines.append(f"")

    # 缺失清单
    report_lines.append(f"## 缺失项清单")
    report_lines.append(f"")
    report_lines.append(f"| 状态 | 槽位名称 | 类别 | 模板数量 | 已匹配数量 |")
    report_lines.append(f"|------|----------|------|----------|------------|")

    missing_count = 0
    for slot_name, slot_config in template_structure.items():
        is_filled = slot_name in matches and matches[slot_name]
        existing_count = slot_config["expected_count"]

        if is_filled:
            status = "✅"
            count = len(matches[slot_name])
            report_lines.append(f"| {status} | {slot_name} | {slot_config['category']} | {existing_count} | {count} |")
        else:
            status = "❌ 缺失"
            missing_count += 1
            report_lines.append(f"| {status} | {slot_name} | {slot_config['category']} | {existing_count} | 0 |")

    if missing_count == 0:
        report_lines.append(f"")
        report_lines.append(f"> ✅ 所有模板槽位均已填充，无缺失项。")

    report_lines.append(f"")

    # 已匹配详情
    report_lines.append(f"## 已匹配文件详情")
    report_lines.append(f"")

    for slot_name in template_structure:
        if slot_name in matches and matches[slot_name]:
            report_lines.append(f"### {slot_name}（{template_structure[slot_name]['category']}）")
            report_lines.append(f"")
            file_list = matches[slot_name]
            report_lines.append(f"共 {len(file_list)} 个文件：")
            report_lines.append(f"")
            for src_path, dest_filename in file_list:
                src_name = os.path.basename(src_path)
                report_lines.append(f"- `{src_name}` → `{dest_filename}`")
            report_lines.append(f"")

    # 未分类文件
    if unclassified:
        report_lines.append(f"## ⚠ 未分类文件")
        report_lines.append(f"")
        report_lines.append(f"以下文件无法自动分类，请手动处理：")
        report_lines.append(f"")
        for filepath, reason in unclassified:
            report_lines.append(f"- `{os.path.basename(filepath)}` — {reason}")
            report_lines.append(f"  路径: `{filepath}`")
        report_lines.append(f"")

    report_lines.append(f"---")
    report_lines.append(f"")
    report_lines.append(f"*报告由 filing_mapper.py 自动生成*")

    report_content = "\n".join(report_lines)
    return report_content


def write_report(content, target_dir):
    """写入报告文件"""
    report_path = os.path.join(target_dir, "_缺失清单.md")
    with open(report_path, "w", encoding="utf-8") as f:
        f.write(content)
    print(f"[INFO] 缺失清单已生成: {report_path}")
    return report_path


# ============================================================
# 主流程
# ============================================================

def cmd_map(args):
    """执行 map 子命令"""
    source_dir = args.source
    template_dir = args.template
    target_dir = args.target
    enterprise_name = args.enterprise

    print("=" * 60)
    print(f"  高新认定材料归档映射器")
    print(f"  企业: {enterprise_name}")
    print(f"  源目录: {source_dir}")
    print(f"  模板目录: {template_dir}")
    print(f"  目标目录: {target_dir}")
    print("=" * 60)
    print()

    print("[步骤 1/5] 扫描模板目录...")
    template_structure = scan_template(template_dir)
    print(f"  模板共 {len(template_structure)} 个槽位")
    for name, slot in template_structure.items():
        loc = "根目录" if slot.get("root_level") else slot["path"]
        print(f"    [{slot['category']}] {name}: {slot['expected_count']} 个文件 ({loc})")
    print()

    print("[步骤 2/5] 扫描源目录并分类...")
    categorized, unclassified = scan_source(source_dir)
    print()

    print("[步骤 3/5] 匹配源文件到模板槽位...")
    matches = match_source_to_template(categorized, template_structure, enterprise_name)
    total_matched = sum(len(v) for v in matches.values())
    print(f"  匹配结果: {len(matches)} 个槽位有匹配, 共 {total_matched} 个文件")
    for slot_name, file_list in matches.items():
        print(f"    {slot_name}: {len(file_list)} 个文件")
    print()

    print("[步骤 4/5] 创建目标目录结构并复制文件...")
    create_target_structure(template_structure, target_dir)
    copy_log = copy_matched_files(matches, template_structure, target_dir)
    print()

    print("[步骤 5/5] 生成缺失清单报告...")
    report_content = generate_missing_report(
        template_structure, matches, unclassified, target_dir, enterprise_name
    )
    report_path = write_report(report_content, target_dir)
    print()

    # 终末统计
    total_slots = len(template_structure)
    filled_slots = sum(1 for s in template_structure if s in matches and matches[s])
    missing_slots = total_slots - filled_slots

    print("=" * 60)
    print(f"  映射完成!")
    print(f"  总槽位: {total_slots}  |  已填充: {filled_slots}  |  缺失: {missing_slots}")
    print(f"  已复制文件: {total_matched} 个")
    print(f"  缺失清单: {report_path}")
    print("=" * 60)

    return {
        "total_slots": total_slots,
        "filled_slots": filled_slots,
        "missing_slots": missing_slots,
        "matched_files": total_matched,
        "unclassified_count": len(unclassified),
        "report_path": report_path,
    }


def main():
    parser = argparse.ArgumentParser(
        description="高新认定材料归档映射器 - 将企业材料按模板结构映射归档",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
示例:
  python filing_mapper.py map \\
    --source "D:\\\\企业材料" \\
    --template "D:\\\\00 高新模板（全）" \\
    --target "D:\\\\归档输出" \\
    --enterprise "云充科技"
        """,
    )
    subparsers = parser.add_subparsers(dest="command", help="子命令")

    map_parser = subparsers.add_parser("map", help="执行材料归档映射")
    map_parser.add_argument("--source", required=True, help="源目录（企业材料所在目录）")
    map_parser.add_argument("--template", required=True, help="模板目录（基准模板所在目录）")
    map_parser.add_argument("--target", required=True, help="目标目录（归档输出目录）")
    map_parser.add_argument("--enterprise", required=True, help="企业名称")

    args = parser.parse_args()

    if args.command == "map":
        result = cmd_map(args)
        sys.exit(0 if result["missing_slots"] == 0 else 0)
    else:
        parser.print_help()
        sys.exit(1)


if __name__ == "__main__":
    main()
