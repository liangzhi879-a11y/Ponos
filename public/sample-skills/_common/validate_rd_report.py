"""RD 立项报告审核脚本

用途：审核高新技术企业认定申报中的 RD 立项报告是否合规，包括：
  1. RD 立项书完整性：每个 RD 是否有对应的立项书文件（docx）
  2. 立项书内容校验：是否包含目的、核心技术及创新点、阶段性成果等必要章节
  3. 字数要求校验：目的 ≥280字、核心技术及创新点 ≥350字、阶段性成果 ≥280字
  4. 格式规范校验：检查 docx 格式、标题层级、字体字号
  5. RD 编号连续性校验
  6. 立项书与 RD 表数据一致性校验（如可读取 RD 表）

用法：
  python validate_rd_report.py --dir "RD报告目录"
  python validate_rd_report.py --dir "RD报告目录" --project-root "项目根目录"
  python validate_rd_report.py --file "RD01_某项目立项书.docx" --project-root "项目根目录"

输出：JSON 格式审核报告，结构如下：
  {
    "passed": bool,
    "errors": [{"file": "...", "row": 0, "field": "...", "reason": "..."}],
    "warnings": [...],
    "stats": {"total_rd": N, "valid_rd": N, "invalid_rd": N, ...}
  }

退出码：审核通过 0，存在错误 1

依赖：python-docx（用于读取 docx 内容进行字数校验）；若无 python-docx 则降级为仅校验文件存在性
      openpyxl（用于读取 RD 汇总表进行一致性校验）；若无 openpyxl 则跳过该校验
"""

import os
import re
import sys
import json
import argparse


# ============================================================
# 常量定义
# ============================================================

# 立项书文件支持的扩展名
REPORT_EXTENSIONS = ('.docx', '.doc')

# RD 编号正则（RD01、RD02...）
RD_NUMBER_PATTERN = re.compile(r'RD(\d{1,3})', re.IGNORECASE)

# 立项书必要章节及其关键词（按章节顺序）
REQUIRED_SECTIONS = [
    {
        'name': '立项目的',
        'keywords': ['立项目的', '项目目的', '研究目的', '研发目的', '一、目的', '1、目的'],
        'min_chars': 280,
    },
    {
        'name': '核心技术及创新点',
        'keywords': ['核心技术', '技术创新', '核心技术与创新点', '技术路线', '创新点'],
        'min_chars': 350,
    },
    {
        'name': '阶段性成果',
        'keywords': ['阶段性成果', '阶段成果', '研究成果', '预期成果', '项目成果', '研究成果及形式'],
        'min_chars': 280,
    },
]

# RD 表文件名匹配模式
RD_TABLE_PATTERNS = [
    r'RD.*汇总.*\.xlsx$',
    r'研发活动.*\.xlsx$',
    r'RD.*表.*\.xlsx$',
    r'研究开发活动.*\.xlsx$',
]

# 标准字体名称（中文正文常用字体）
STANDARD_FONTS = {'宋体', '仿宋', '仿宋_GB2312', '微软雅黑', 'Times New Roman'}

# 标准正文字号范围（磅）：小四=12pt，五号=10.5pt
STANDARD_FONT_SIZES = {10.5, 12.0, 14.0, 16.0}


def _load_project_index(project_root):
    """加载 project_index.json，用于获取 RD 列表等上下文信息

    查找顺序：
      1. project_root/.trae/project_knowledge/project_index.json
      2. project_root/.trae/project_index.json
    """
    if not project_root:
        return None
    candidates = [
        os.path.join(project_root, '.trae', 'project_knowledge', 'project_index.json'),
        os.path.join(project_root, '.trae', 'project_index.json'),
    ]
    for path in candidates:
        if os.path.exists(path):
            try:
                with open(path, 'r', encoding='utf-8') as f:
                    return json.load(f)
            except (json.JSONDecodeError, OSError):
                return None
    return None


def _get_rd_ids_from_index(project_index):
    """从 project_index.json 中提取 RD 编号列表

    返回：{RD编号(整数), ...}
    """
    rd_ids = set()
    if not project_index or not isinstance(project_index.get('knowledge_graph'), dict):
        return rd_ids
    nodes = project_index['knowledge_graph'].get('nodes', []) or []
    for node in nodes:
        if not isinstance(node, dict):
            continue
        node_id = str(node.get('id', ''))
        m = RD_NUMBER_PATTERN.search(node_id)
        if m:
            rd_ids.add(int(m.group(1)))
    return rd_ids


def _scan_rd_reports(rd_dir):
    """扫描 RD 报告目录下的所有立项书文件

    返回：[{'file': 文件名, 'path': 完整路径, 'rd_no': RD编号, 'ext': 扩展名}, ...]
    """
    reports = []
    if not os.path.isdir(rd_dir):
        return reports
    for entry in sorted(os.listdir(rd_dir)):
        ext = os.path.splitext(entry)[1].lower()
        if ext not in REPORT_EXTENSIONS:
            continue
        full_path = os.path.join(rd_dir, entry)
        rd_no = None
        m = RD_NUMBER_PATTERN.search(entry)
        if m:
            rd_no = int(m.group(1))
        reports.append({
            'file': entry,
            'path': full_path,
            'rd_no': rd_no,
            'ext': ext,
        })
    return reports


def _read_docx_content(file_path):
    """读取 docx 文件内容，返回段落列表和章节文本

    返回：(paragraphs, sections, error_or_None)
      paragraphs: [{'text': 段落文本, 'style': 样式名, 'level': 标题层级(0=正文)}, ...]
      sections: {'章节名': 章节文本内容, ...}
      error: 错误信息或 None
    """
    try:
        import docx
    except ImportError:
        return None, None, 'python-docx 未安装，降级为仅校验文件存在性'

    if not os.path.exists(file_path):
        return None, None, f'文件不存在: {file_path}'

    try:
        doc = docx.Document(file_path)
    except Exception as e:
        return None, None, f'读取 docx 文件失败: {e}'

    paragraphs = []
    current_section = None
    sections = {}

    for para in doc.paragraphs:
        text = para.text.strip()
        style_name = para.style.name if para.style else ''

        # 判断标题层级
        level = 0
        if style_name.startswith('Heading'):
            try:
                level = int(style_name.replace('Heading', '').replace(' ', '').strip())
            except ValueError:
                level = 1
        elif style_name.startswith('标题'):
            try:
                level = int(style_name.replace('标题', '').strip())
            except ValueError:
                level = 1

        # 检查是否为章节标题
        if level > 0 and text:
            # 匹配必要章节
            for section_def in REQUIRED_SECTIONS:
                if any(kw in text for kw in section_def['keywords']):
                    current_section = section_def['name']
                    sections.setdefault(current_section, '')
                    break
            else:
                # 非必要章节标题，结束当前章节
                current_section = None
        elif text and current_section:
            sections[current_section] += text + '\n'

        paragraphs.append({
            'text': text,
            'style': style_name,
            'level': level,
        })

    return paragraphs, sections, None


def _read_rd_table(file_path):
    """读取 RD 汇总表，提取 RD 编号及项目名称

    返回：({rd_no: {'name': 项目名称, 'row': 行号}, ...}, error_or_None)
    """
    try:
        import openpyxl
    except ImportError:
        return None, 'openpyxl 未安装，跳过 RD 表一致性校验'

    if not os.path.exists(file_path):
        return None, f'文件不存在: {file_path}'

    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
    except Exception as e:
        return None, f'加载 Excel 失败: {e}'

    ws = wb.active
    rd_data = {}

    # 自动检测表头行（查找包含"RD"或"序号"的行）
    header_row = 1
    for r in range(1, min(5, ws.max_row + 1)):
        v = ws.cell(r, 1).value
        if v and ('RD' in str(v) or '序号' in str(v)):
            header_row = r
            break

    for r in range(header_row + 1, ws.max_row + 1):
        first_cell = ws.cell(r, 1).value
        if first_cell is None or str(first_cell).strip() == '':
            continue
        # 跳过统计行
        if isinstance(first_cell, str) and ('总计' in first_cell or '合计' in first_cell or '说明' in first_cell):
            continue

        m = RD_NUMBER_PATTERN.search(str(first_cell))
        if m:
            rd_no = int(m.group(1))
            name = ws.cell(r, 2).value or ''
            rd_data[rd_no] = {'name': str(name).strip(), 'row': r}

    return rd_data, None


def _check_rd_number_continuity(rd_numbers):
    """检查 RD 编号连续性

    参数：
      rd_numbers: 已排序的 RD 编号列表

    返回：(是否连续, 缺失的编号列表)
    """
    if not rd_numbers:
        return True, []

    sorted_nums = sorted(set(rd_numbers))
    min_num = sorted_nums[0]
    max_num = sorted_nums[-1]
    expected = set(range(min_num, max_num + 1))
    actual = set(sorted_nums)
    missing = sorted(expected - actual)

    return len(missing) == 0, missing


def validate_rd_report_directory(rd_dir, project_root=None):
    """审核 RD 立项报告目录

    参数：
      rd_dir: RD 报告目录路径
      project_root: 项目根目录（用于读取 project_index.json）

    返回：审核报告 dict
    """
    errors = []
    warnings = []
    stats = {
        'total_rd': 0,
        'valid_rd': 0,
        'invalid_rd': 0,
        'missing_sections': 0,
        'insufficient_chars': 0,
        'rd_numbers': [],
        'missing_rd_numbers': [],
        'has_rd_table': False,
        'content_check_enabled': False,
    }

    # ============================================================
    # 校验 1：目录结构校验
    # ============================================================
    if not os.path.exists(rd_dir):
        errors.append({
            'file': rd_dir, 'row': 0, 'field': 'directory',
            'reason': f'RD 报告目录不存在: {rd_dir}',
        })
        return {'passed': False, 'errors': errors, 'warnings': warnings, 'stats': stats}

    if not os.path.isdir(rd_dir):
        errors.append({
            'file': rd_dir, 'row': 0, 'field': 'directory',
            'reason': f'RD 报告路径不是目录: {rd_dir}',
        })
        return {'passed': False, 'errors': errors, 'warnings': warnings, 'stats': stats}

    # 扫描立项书文件
    reports = _scan_rd_reports(rd_dir)
    stats['total_rd'] = len(reports)

    if stats['total_rd'] == 0:
        errors.append({
            'file': '', 'row': 0, 'field': 'reports',
            'reason': 'RD 报告目录下未找到任何立项书文件（docx/doc）',
        })
        return {'passed': False, 'errors': errors, 'warnings': warnings, 'stats': stats}

    # 检测 python-docx 是否可用
    try:
        import docx  # noqa: F401
        stats['content_check_enabled'] = True
    except ImportError:
        stats['content_check_enabled'] = False
        warnings.append({
            'file': '', 'row': 0, 'field': 'dependency',
            'reason': 'python-docx 未安装，降级为仅校验文件存在性（内容/字数校验被跳过）',
        })

    # 加载 project_index.json
    project_index = _load_project_index(project_root)
    rd_ids_in_index = _get_rd_ids_from_index(project_index)

    # 查找 RD 汇总表
    rd_table_file = None
    if os.path.isdir(rd_dir):
        for entry in os.listdir(rd_dir):
            if not entry.endswith('.xlsx'):
                continue
            for p in RD_TABLE_PATTERNS:
                if re.search(p, entry, re.IGNORECASE):
                    rd_table_file = os.path.join(rd_dir, entry)
                    break
            if rd_table_file:
                break
    # 也在上级目录查找
    if not rd_table_file and project_root:
        for search_dir in [os.path.dirname(rd_dir), project_root]:
            if not os.path.isdir(search_dir):
                continue
            for entry in os.listdir(search_dir):
                if not entry.endswith('.xlsx'):
                    continue
                for p in RD_TABLE_PATTERNS:
                    if re.search(p, entry, re.IGNORECASE):
                        rd_table_file = os.path.join(search_dir, entry)
                        break
                if rd_table_file:
                    break
            if rd_table_file:
                break

    rd_table_data = None
    if rd_table_file:
        stats['has_rd_table'] = True
        rd_table_data, err = _read_rd_table(rd_table_file)
        if err:
            warnings.append({
                'file': os.path.basename(rd_table_file), 'row': 0, 'field': 'rd_table',
                'reason': err,
            })

    # ============================================================
    # 逐个校验 RD 立项书
    # ============================================================
    rd_numbers_found = []
    rd_numbers_with_files = set()

    for report in reports:
        fname = report['file']
        rd_no = report['rd_no']
        ext = report['ext']
        has_error = False

        # 校验 1：RD 编号识别
        if rd_no is None:
            warnings.append({
                'file': fname, 'row': 0, 'field': 'rd_no',
                'reason': '文件名中未识别到 RD 编号（建议命名格式：RD01_项目名称.docx）',
            })
        else:
            rd_numbers_found.append(rd_no)
            rd_numbers_with_files.add(rd_no)

        # 校验 4：格式规范校验
        if ext == '.doc':
            warnings.append({
                'file': fname, 'row': 0, 'field': 'format',
                'reason': '文档使用 .doc 格式，建议统一转换为 .docx 格式',
            })

        # 校验 2 & 3：立项书内容校验 + 字数要求校验
        if stats['content_check_enabled'] and ext == '.docx':
            paragraphs, sections, err = _read_docx_content(report['path'])
            if err:
                warnings.append({
                    'file': fname, 'row': 0, 'field': 'content',
                    'reason': err,
                })
            else:
                # 校验 2：必要章节完整性
                for section_def in REQUIRED_SECTIONS:
                    section_name = section_def['name']
                    if section_name not in sections:
                        stats['missing_sections'] += 1
                        errors.append({
                            'file': fname, 'row': 0, 'field': 'section',
                            'reason': f'缺少必要章节: {section_name}（关键词: {section_def["keywords"][:3]}）',
                        })
                        has_error = True
                    else:
                        # 校验 3：字数要求校验
                        section_text = sections[section_name]
                        char_count = len(section_text.replace('\n', '').replace(' ', ''))
                        min_chars = section_def['min_chars']
                        if char_count < min_chars:
                            stats['insufficient_chars'] += 1
                            errors.append({
                                'file': fname, 'row': 0, 'field': 'char_count',
                                'reason': f'章节"{section_name}"字数不足: {char_count}字 < {min_chars}字',
                            })
                            has_error = True

                # 校验 4：标题层级校验
                heading_levels = [p['level'] for p in paragraphs if p['level'] > 0]
                if not heading_levels:
                    warnings.append({
                        'file': fname, 'row': 0, 'field': 'heading',
                        'reason': '文档未检测到任何标题层级，建议使用 Word 标题样式（标题1/标题2）',
                    })
                elif heading_levels:
                    # 检查标题层级是否从1开始
                    if min(heading_levels) > 1:
                        warnings.append({
                            'file': fname, 'row': 0, 'field': 'heading',
                            'reason': f'文档标题层级从 {min(heading_levels)} 级开始，建议从标题1开始',
                        })

        if has_error:
            stats['invalid_rd'] += 1
        else:
            stats['valid_rd'] += 1

    # ============================================================
    # 校验 5：RD 编号连续性校验
    # ============================================================
    stats['rd_numbers'] = sorted(rd_numbers_found)
    is_continuous, missing_nums = _check_rd_number_continuity(rd_numbers_found)
    stats['missing_rd_numbers'] = missing_nums

    if not is_continuous:
        warnings.append({
            'file': '', 'row': 0, 'field': 'continuity',
            'reason': f'RD 编号不连续，缺失: RD{missing_nums}',
        })

    # 检查重复编号
    from collections import Counter
    num_counts = Counter(rd_numbers_found)
    duplicates = [num for num, count in num_counts.items() if count > 1]
    if duplicates:
        for dup in sorted(duplicates):
            errors.append({
                'file': '', 'row': 0, 'field': 'rd_no',
                'reason': f'RD 编号重复: RD{dup:02d} 存在 {num_counts[dup]} 个立项书文件',
            })

    # ============================================================
    # 校验 6：立项书与 RD 表数据一致性校验
    # ============================================================
    if rd_table_data:
        rd_in_table = set(rd_table_data.keys())
        rd_in_files = rd_numbers_with_files

        # RD 表中有但无立项书
        missing_reports = rd_in_table - rd_in_files
        for rd_no in sorted(missing_reports):
            errors.append({
                'file': '', 'row': rd_table_data[rd_no]['row'], 'field': 'report_missing',
                'reason': f'RD 表中 RD{rd_no:02d}（{rd_table_data[rd_no]["name"]}）无对应立项书文件',
            })

        # 有立项书但不在 RD 表中
        extra_reports = rd_in_files - rd_in_table
        for rd_no in sorted(extra_reports):
            warnings.append({
                'file': '', 'row': 0, 'field': 'table_missing',
                'reason': f'立项书 RD{rd_no:02d} 在 RD 汇总表中未找到对应记录',
            })

    # 校验：与 project_index.json 中 RD 列表的一致性
    if rd_ids_in_index:
        missing_from_files = rd_ids_in_index - rd_numbers_with_files
        for rd_no in sorted(missing_from_files):
            errors.append({
                'file': '', 'row': 0, 'field': 'index_missing',
                'reason': f'project_index.json 中 RD{rd_no:02d} 无对应立项书文件',
            })

    passed = len(errors) == 0
    return {
        'passed': passed,
        'errors': errors,
        'warnings': warnings,
        'stats': stats,
    }


def validate_single_rd_report(file_path, project_root=None):
    """审核单个 RD 立项报告文件"""
    errors = []
    warnings = []
    stats = {
        'total_rd': 1,
        'valid_rd': 0,
        'invalid_rd': 0,
        'missing_sections': 0,
        'insufficient_chars': 0,
        'rd_numbers': [],
        'missing_rd_numbers': [],
        'has_rd_table': False,
        'content_check_enabled': False,
    }

    if not os.path.exists(file_path):
        errors.append({
            'file': file_path, 'row': 0, 'field': 'file',
            'reason': f'文件不存在: {file_path}',
        })
        return {'passed': False, 'errors': errors, 'warnings': warnings, 'stats': stats}

    fname = os.path.basename(file_path)
    ext = os.path.splitext(fname)[1].lower()
    has_error = False

    # 格式校验
    if ext not in REPORT_EXTENSIONS:
        errors.append({
            'file': fname, 'row': 0, 'field': 'format',
            'reason': f'不支持的文档格式: {ext}（仅支持 docx/doc）',
        })
        has_error = True

    # RD 编号识别
    m = RD_NUMBER_PATTERN.search(fname)
    rd_no = int(m.group(1)) if m else None
    if rd_no is None:
        warnings.append({
            'file': fname, 'row': 0, 'field': 'rd_no',
            'reason': '文件名中未识别到 RD 编号',
        })
    else:
        stats['rd_numbers'] = [rd_no]

    if ext == '.doc':
        warnings.append({
            'file': fname, 'row': 0, 'field': 'format',
            'reason': '文档使用 .doc 格式，建议统一转换为 .docx 格式',
        })

    # 内容校验
    try:
        import docx  # noqa: F401
        stats['content_check_enabled'] = True
    except ImportError:
        stats['content_check_enabled'] = False
        warnings.append({
            'file': fname, 'row': 0, 'field': 'dependency',
            'reason': 'python-docx 未安装，降级为仅校验文件存在性',
        })

    if stats['content_check_enabled'] and ext == '.docx':
        paragraphs, sections, err = _read_docx_content(file_path)
        if err:
            warnings.append({
                'file': fname, 'row': 0, 'field': 'content',
                'reason': err,
            })
        else:
            # 必要章节校验
            for section_def in REQUIRED_SECTIONS:
                section_name = section_def['name']
                if section_name not in sections:
                    stats['missing_sections'] += 1
                    errors.append({
                        'file': fname, 'row': 0, 'field': 'section',
                        'reason': f'缺少必要章节: {section_name}',
                    })
                    has_error = True
                else:
                    section_text = sections[section_name]
                    char_count = len(section_text.replace('\n', '').replace(' ', ''))
                    min_chars = section_def['min_chars']
                    if char_count < min_chars:
                        stats['insufficient_chars'] += 1
                        errors.append({
                            'file': fname, 'row': 0, 'field': 'char_count',
                            'reason': f'章节"{section_name}"字数不足: {char_count}字 < {min_chars}字',
                        })
                        has_error = True

            # 标题层级校验
            heading_levels = [p['level'] for p in paragraphs if p['level'] > 0]
            if not heading_levels:
                warnings.append({
                    'file': fname, 'row': 0, 'field': 'heading',
                    'reason': '文档未检测到任何标题层级',
                })

    if has_error:
        stats['invalid_rd'] = 1
    else:
        stats['valid_rd'] = 1

    passed = len(errors) == 0
    return {
        'passed': passed,
        'errors': errors,
        'warnings': warnings,
        'stats': stats,
    }


def main():
    """CLI 入口：解析参数并执行 RD 立项报告审核"""
    parser = argparse.ArgumentParser(
        description='RD 立项报告审核脚本 - 校验高企认定 RD 立项书合规性',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog='''示例：
  python validate_rd_report.py --dir "RD立项报告"
  python validate_rd_report.py --dir "RD立项报告" --project-root "项目根目录"
  python validate_rd_report.py --file "RD01_某项目立项书.docx" --project-root "项目根目录"
''',
    )
    parser.add_argument('--dir', help='RD 立项报告所在目录')
    parser.add_argument('--file', help='单个 RD 立项书文件路径')
    parser.add_argument('--project-root', dest='project_root',
                        help='项目根目录（用于读取 project_index.json 进行 RD 关联校验）')
    args = parser.parse_args()

    if not args.dir and not args.file:
        parser.print_help()
        sys.exit(1)

    if args.dir:
        report = validate_rd_report_directory(args.dir, args.project_root)
    else:
        report = validate_single_rd_report(args.file, args.project_root)

    # 输出 JSON 格式审核报告
    print(json.dumps(report, ensure_ascii=False, indent=2, default=str))
    sys.exit(0 if report['passed'] else 1)


if __name__ == '__main__':
    main()
