"""科技人员材料审核脚本

用途：审核高新技术企业认定申报中的科技人员证明材料是否合规，包括：
  1. 科技人员清单完整性（三方对比：花名册×社保×台账）
  2. 183 天在职天数校验（符合条件人员是否全部 ≥183 天）
  3. 社保占比校验（科技人员占比是否 ≥10%）
  4. 社保缴费记录完整性（上年12月带公章的社保缴费记录）
  5. 花名册完整性校验
  6. 台账人员与花名册一致性校验

用法：
  python validate_staff.py --dir "人员材料目录"
  python validate_staff.py --dir "人员材料目录" --project-root "项目根目录"
  python validate_staff.py --file "科技人员清单.xlsx" --project-root "项目根目录"

输出：JSON 格式审核报告，结构如下：
  {
    "passed": bool,
    "errors": [{"file": "...", "row": 0, "field": "...", "reason": "..."}],
    "warnings": [...],
    "stats": {"total_staff": N, "qualified_staff": N, "staff_ratio": ..., ...}
  }

退出码：审核通过 0，存在错误 1

依赖：openpyxl（用于读取 xlsx 清单）；若无 openpyxl 则跳过表格校验
"""

import os
import re
import sys
import json
import argparse


# ============================================================
# 常量定义
# ============================================================

# 证明材料文件支持的扩展名
PROOF_EXTENSIONS = ('.pdf', '.jpg', '.jpeg', '.png', '.xls', '.xlsx', '.doc', '.docx')

# 科技人员清单表头（与 validate_tables.py 中 staff 规格一致）
STAFF_HEADERS = [
    '序号', '姓名', '身份证号码', '性别', '部门', '岗位',
    '入职日期', '离职日期', '上年在职天数',
    '花名册', '社保清单', '台账',
    '是否符合科技人员条件', '不符合原因', '备注',
]

# 三方对比来源列有效标记值
THREE_WAY_MARKS = {'是', '√', 'Y', 'y', '1', True, 1}

# 统计行标签
SUMMARY_LABELS = [
    '花名册总人数', '社保清单总人数', '台账纳入总人数',
    '符合科技人员条件人数', '科技人员占比（基于社保）',
]

# 政策阈值
MIN_WORK_DAYS = 183        # 科技人员累计工作时长下限
MIN_STAFF_RATIO = 0.10     # 科技人员占比下限
SOCIAL_SECURITY_KEYWORDS = ['社保', '社会保险', '缴费记录']
ROSTER_KEYWORDS = ['花名册', '人员名册', '员工名册']
LEDGER_KEYWORDS = ['台账', '研发费用台账']

# 身份证号正则（18 位，最后一位可为 X）
ID_CARD_PATTERN = re.compile(r'\d{17}[\dXx]')


def _load_project_index(project_root):
    """加载 project_index.json，用于获取企业总人数等基础数据"""
    if not project_root:
        return None
    candidates = [
        os.path.join(project_root, '.trae', 'project_knowledge', 'project_index.json'),
        os.path.join(project_root, '.trae', 'project_index.json'),
    ]
    for path in candidates:
        if os.path.exists(path):
            try:
                with open(path, 'r', encoding='utf-8') as f:
                    return json.load(f)
            except (json.JSONDecodeError, OSError):
                return None
    return None


def _is_marked_yes(value):
    """判断三方对比列的值是否表示"存在/是" """
    if value is None:
        return False
    if isinstance(value, bool):
        return value
    s = str(value).strip().lower()
    return s in ('是', '√', 'y', '1', 'true', 'yes')


def _is_qualified_staff(value):
    """判断"是否符合科技人员条件"列的值是否为"是" """
    if value is None:
        return False
    if isinstance(value, bool):
        return value
    s = str(value).strip().lower()
    return s in ('是', '√', 'y', '1', 'true', 'yes')


def _parse_work_days(value):
    """解析在职天数（支持整数和浮点数字符串）"""
    if value is None:
        return None
    try:
        return int(float(str(value).strip()))
    except (ValueError, TypeError):
        # 尝试提取数字
        m = re.search(r'\d+', str(value))
        return int(m.group(0)) if m else None


def _parse_ratio(value):
    """解析占比值（支持 "10.5%" / 0.105 / "10.5" 等格式）"""
    if value is None:
        return None
    s = str(value).strip()
    if '%' in s:
        try:
            return float(s.replace('%', '').strip()) / 100.0
        except ValueError:
            return None
    try:
        v = float(s)
        # 如果值大于 1，认为是百分数形式（如 10.5 表示 10.5%）
        if v > 1:
            return v / 100.0
        return v
    except ValueError:
        return None


def _read_staff_table(file_path):
    """读取科技人员清单（三方对比表）

    返回：({'data_rows': [...], 'summary_rows': {...}, 'headers': [...]}, error_or_None)
    """
    try:
        import openpyxl
    except ImportError:
        return None, 'openpyxl 未安装，跳过表格读取'

    if not os.path.exists(file_path):
        return None, f'文件不存在: {file_path}'

    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
    except Exception as e:
        return None, f'加载 Excel 失败: {e}'

    ws = wb.active
    header_row = 1
    # 自动检测表头行（查找包含"姓名"或"身份证"的行）
    for r in range(1, min(5, ws.max_row + 1)):
        v = ws.cell(r, 2).value  # 第2列通常是"姓名"
        if v and ('姓名' in str(v) or '身份证' in str(v)):
            header_row = r
            break

    # 读取表头
    headers = []
    for c in range(1, ws.max_column + 1):
        v = ws.cell(header_row, c).value
        headers.append(str(v) if v is not None else '')

    data_rows = []
    summary_rows = {}

    for r in range(header_row + 1, ws.max_row + 1):
        first_cell = ws.cell(r, 1).value
        # 跳过空行
        if first_cell is None or str(first_cell).strip() == '':
            continue

        # 检查是否为统计行
        first_str = str(first_cell).strip()
        if any(label in first_str for label in SUMMARY_LABELS):
            summary_rows[first_str] = {
                'row': r,
                'label': first_str,
                'value': ws.cell(r, 3).value,  # 统计值通常在第3列
            }
            continue

        # 跳过非数据行（如"说明"、"备注"等）
        if isinstance(first_cell, str) and ('说明' in first_str or '备注' in first_str):
            continue

        # 数据行
        row_data = {
            'row': r,
            'seq': first_cell,
            'name': ws.cell(r, 2).value,
            'id_card': ws.cell(r, 3).value,
            'gender': ws.cell(r, 4).value,
            'department': ws.cell(r, 5).value,
            'position': ws.cell(r, 6).value,
            'hire_date': ws.cell(r, 7).value,
            'resign_date': ws.cell(r, 8).value,
            'work_days': ws.cell(r, 9).value,
            'in_roster': ws.cell(r, 10).value,
            'in_social': ws.cell(r, 11).value,
            'in_ledger': ws.cell(r, 12).value,
            'is_qualified': ws.cell(r, 13).value,
            'unqualified_reason': ws.cell(r, 14).value,
            'remark': ws.cell(r, 15).value,
        }
        data_rows.append(row_data)

    return {'data_rows': data_rows, 'summary_rows': summary_rows, 'headers': headers}, None


def _scan_staff_proof_files(staff_dir):
    """扫描人员材料目录，识别花名册/社保/台账等关键文件

    返回：{'roster': [...], 'social_security': [...], 'ledger': [...], 'others': [...]}
    """
    result = {'roster': [], 'social_security': [], 'ledger': [], 'others': []}
    if not os.path.isdir(staff_dir):
        return result
    for entry in os.listdir(staff_dir):
        ext = os.path.splitext(entry)[1].lower()
        if ext not in PROOF_EXTENSIONS:
            continue
        if any(kw in entry for kw in SOCIAL_SECURITY_KEYWORDS):
            result['social_security'].append(entry)
        elif any(kw in entry for kw in ROSTER_KEYWORDS):
            result['roster'].append(entry)
        elif any(kw in entry for kw in LEDGER_KEYWORDS):
            result['ledger'].append(entry)
        else:
            result['others'].append(entry)
    return result


def validate_staff_directory(staff_dir, project_root=None):
    """审核科技人员材料目录

    参数：
      staff_dir: 人员材料目录路径
      project_root: 项目根目录（用于读取 project_index.json）

    返回：审核报告 dict
    """
    errors = []
    warnings = []
    stats = {
        'total_staff': 0,
        'qualified_staff': 0,
        'unqualified_staff': 0,
        'staff_ratio': None,
        'total_employees': None,
        'roster_count': 0,
        'social_security_count': 0,
        'ledger_count': 0,
        'insufficient_days_count': 0,
        'three_way_consistent': 0,
        'three_way_inconsistent': 0,
    }

    # ============================================================
    # 校验 1：目录结构校验
    # ============================================================
    if not os.path.exists(staff_dir):
        errors.append({
            'file': staff_dir, 'row': 0, 'field': 'directory',
            'reason': f'人员材料目录不存在: {staff_dir}',
        })
        return {'passed': False, 'errors': errors, 'warnings': warnings, 'stats': stats}

    # 扫描证明材料文件
    proof_files = _scan_staff_proof_files(staff_dir)

    # 查找科技人员清单表
    table_file = None
    table_patterns = [
        r'科技人员清单.*三方对比.*\.xlsx$',
        r'科技人员清单.*\.xlsx$',
        r'科技人员.*信息表.*\.xlsx$',
        r'科技人员.*\.xlsx$',
    ]
    if os.path.isdir(staff_dir):
        for entry in os.listdir(staff_dir):
            if not entry.endswith('.xlsx'):
                continue
            for p in table_patterns:
                if re.search(p, entry):
                    table_file = os.path.join(staff_dir, entry)
                    break
            if table_file:
                break

    # 加载 project_index.json（获取企业总人数）
    project_index = _load_project_index(project_root)

    # ============================================================
    # 校验 4：社保缴费记录完整性（上年12月带公章的社保缴费记录）
    # ============================================================
    stats['social_security_count'] = len(proof_files['social_security'])
    if stats['social_security_count'] == 0:
        errors.append({
            'file': '', 'row': 0, 'field': 'social_security',
            'reason': '未找到社保缴费记录文件（需上年12月带公章的社保缴费记录）',
        })
    else:
        # 检查是否包含上年12月关键词
        last_year = None
        if project_index and project_index.get('last_year'):
            last_year = project_index['last_year']
        has_december = False
        for fname in proof_files['social_security']:
            if '12月' in fname or '12' in fname:
                has_december = True
                break
            # 如果文件名含上年年份，也算
            if last_year and str(last_year) in fname:
                has_december = True
                break
        if not has_december and last_year:
            warnings.append({
                'file': '', 'row': 0, 'field': 'social_security',
                'reason': f'社保缴费记录文件名未体现"{last_year}年12月"，需确认是否为上年12月记录',
            })

    # ============================================================
    # 校验 5：花名册完整性校验
    # ============================================================
    stats['roster_count'] = len(proof_files['roster'])
    if stats['roster_count'] == 0:
        errors.append({
            'file': '', 'row': 0, 'field': 'roster',
            'reason': '未找到花名册文件',
        })

    # ============================================================
    # 校验汇总表
    # ============================================================
    if table_file is None:
        warnings.append({
            'file': '', 'row': 0, 'field': 'table',
            'reason': '未找到科技人员清单（xlsx），跳过表格校验',
        })
    else:
        table_data, err = _read_staff_table(table_file)
        if err:
            warnings.append({
                'file': os.path.basename(table_file), 'row': 0, 'field': 'table',
                'reason': err,
            })
        else:
            data_rows = table_data['data_rows']
            summary_rows = table_data['summary_rows']
            stats['total_staff'] = len(data_rows)
            fname = os.path.basename(table_file)

            # ============================================================
            # 逐行校验
            # ============================================================
            for row_data in data_rows:
                r = row_data['row']
                has_error = False

                # 校验 2：183 天在职天数校验
                is_qualified = _is_qualified_staff(row_data['is_qualified'])
                work_days = _parse_work_days(row_data['work_days'])

                if is_qualified:
                    stats['qualified_staff'] += 1
                    if work_days is None:
                        errors.append({
                            'file': fname, 'row': r, 'field': 'work_days',
                            'reason': f'第{r}行 {row_data["name"]} 符合科技人员条件但"上年在职天数"为空',
                        })
                        has_error = True
                    elif work_days < MIN_WORK_DAYS:
                        stats['insufficient_days_count'] += 1
                        errors.append({
                            'file': fname, 'row': r, 'field': 'work_days',
                            'reason': f'第{r}行 {row_data["name"]} 符合科技人员条件但在职天数 {work_days} 天 < {MIN_WORK_DAYS} 天',
                        })
                        has_error = True
                else:
                    stats['unqualified_staff'] += 1

                # 校验 1：三方对比完整性（花名册×社保×台账）
                in_roster = _is_marked_yes(row_data['in_roster'])
                in_social = _is_marked_yes(row_data['in_social'])
                in_ledger = _is_marked_yes(row_data['in_ledger'])

                # 三方标记一致性
                marks = [in_roster, in_social, in_ledger]
                if all(marks):
                    stats['three_way_consistent'] += 1
                elif not any(marks):
                    stats['three_way_inconsistent'] += 1
                    warnings.append({
                        'file': fname, 'row': r, 'field': 'three_way',
                        'reason': f'第{r}行 {row_data["name"]} 三方（花名册/社保/台账）均未标记',
                    })
                else:
                    # 部分标记，检查不一致情况
                    missing = []
                    if not in_roster:
                        missing.append('花名册')
                    if not in_social:
                        missing.append('社保')
                    if not in_ledger:
                        missing.append('台账')
                    stats['three_way_inconsistent'] += 1
                    # 仅当符合科技人员条件时，三方缺失才视为错误
                    if is_qualified:
                        if not in_roster:
                            errors.append({
                                'file': fname, 'row': r, 'field': 'roster',
                                'reason': f'第{r}行 {row_data["name"]} 符合科技人员条件但未在花名册中',
                            })
                            has_error = True
                        if not in_social:
                            errors.append({
                                'file': fname, 'row': r, 'field': 'social',
                                'reason': f'第{r}行 {row_data["name"]} 符合科技人员条件但未在社保清单中',
                            })
                            has_error = True

                # 校验 6：台账人员与花名册一致性
                # 台账中有但花名册中没有，或反之，视为不一致
                if in_ledger and not in_roster:
                    warnings.append({
                        'file': fname, 'row': r, 'field': 'ledger_roster',
                        'reason': f'第{r}行 {row_data["name"]} 在台账中但不在花名册中',
                    })

                # 身份证号完整性校验
                id_card = row_data['id_card']
                if not id_card or str(id_card).strip() == '':
                    warnings.append({
                        'file': fname, 'row': r, 'field': 'id_card',
                        'reason': f'第{r}行 {row_data["name"]} 身份证号为空',
                    })
                elif not ID_CARD_PATTERN.match(str(id_card).strip()):
                    warnings.append({
                        'file': fname, 'row': r, 'field': 'id_card',
                        'reason': f'第{r}行 {row_data["name"]} 身份证号格式不规范',
                    })

            # ============================================================
            # 校验 3：社保占比校验（科技人员占比是否 ≥10%）
            # ============================================================
            # 优先从统计行获取占比
            ratio_value = None
            for label in SUMMARY_LABELS:
                if '占比' in label and label in summary_rows:
                    ratio_value = _parse_ratio(summary_rows[label]['value'])
                    break

            # 从 project_index 获取企业总人数
            total_employees = None
            if project_index and project_index.get('data_summary'):
                ds = project_index['data_summary']
                if isinstance(ds, dict):
                    total_employees = ds.get('total_employees') or ds.get('staff_count')

            # 如果统计行有占比值，直接使用
            if ratio_value is not None:
                stats['staff_ratio'] = ratio_value
                if ratio_value < MIN_STAFF_RATIO:
                    errors.append({
                        'file': fname, 'row': 0, 'field': 'staff_ratio',
                        'reason': f'科技人员占比 {ratio_value*100:.2f}% < {MIN_STAFF_RATIO*100:.0f}%（认定要求下限）',
                    })
            elif stats['qualified_staff'] > 0 and total_employees and total_employees > 0:
                # 自行计算占比
                ratio = stats['qualified_staff'] / total_employees
                stats['staff_ratio'] = ratio
                stats['total_employees'] = total_employees
                if ratio < MIN_STAFF_RATIO:
                    errors.append({
                        'file': fname, 'row': 0, 'field': 'staff_ratio',
                        'reason': f'科技人员占比 {ratio*100:.2f}% < {MIN_STAFF_RATIO*100:.0f}%（{stats["qualified_staff"]}/{total_employees}）',
                    })
            elif stats['qualified_staff'] > 0:
                # 无法计算占比
                stats['staff_ratio'] = None
                warnings.append({
                    'file': fname, 'row': 0, 'field': 'staff_ratio',
                    'reason': '无法计算科技人员占比（缺少企业总人数数据）',
                })

            # 校验统计行完整性
            missing_summaries = [s for s in SUMMARY_LABELS if s not in summary_rows]
            if missing_summaries:
                warnings.append({
                    'file': fname, 'row': 0, 'field': 'summary',
                    'reason': f'科技人员清单缺少统计行: {missing_summaries}',
                })

    passed = len(errors) == 0
    return {
        'passed': passed,
        'errors': errors,
        'warnings': warnings,
        'stats': stats,
    }


def validate_single_staff_file(file_path, project_root=None):
    """审核单个科技人员清单文件"""
    errors = []
    warnings = []
    stats = {
        'total_staff': 0,
        'qualified_staff': 0,
        'unqualified_staff': 0,
        'staff_ratio': None,
        'insufficient_days_count': 0,
        'three_way_consistent': 0,
        'three_way_inconsistent': 0,
    }

    if not os.path.exists(file_path):
        errors.append({
            'file': file_path, 'row': 0, 'field': 'file',
            'reason': f'文件不存在: {file_path}',
        })
        return {'passed': False, 'errors': errors, 'warnings': warnings, 'stats': stats}

    table_data, err = _read_staff_table(file_path)
    if err:
        errors.append({
            'file': os.path.basename(file_path), 'row': 0, 'field': 'table',
            'reason': err,
        })
        return {'passed': False, 'errors': errors, 'warnings': warnings, 'stats': stats}

    data_rows = table_data['data_rows']
    summary_rows = table_data['summary_rows']
    stats['total_staff'] = len(data_rows)
    fname = os.path.basename(file_path)

    project_index = _load_project_index(project_root)

    for row_data in data_rows:
        r = row_data['row']
        is_qualified = _is_qualified_staff(row_data['is_qualified'])
        work_days = _parse_work_days(row_data['work_days'])

        if is_qualified:
            stats['qualified_staff'] += 1
            if work_days is not None and work_days < MIN_WORK_DAYS:
                stats['insufficient_days_count'] += 1
                errors.append({
                    'file': fname, 'row': r, 'field': 'work_days',
                    'reason': f'第{r}行 {row_data["name"]} 在职天数 {work_days} 天 < {MIN_WORK_DAYS} 天',
                })
            elif work_days is None:
                errors.append({
                    'file': fname, 'row': r, 'field': 'work_days',
                    'reason': f'第{r}行 {row_data["name"]} 符合条件但在职天数为空',
                })
        else:
            stats['unqualified_staff'] += 1

        # 三方对比
        in_roster = _is_marked_yes(row_data['in_roster'])
        in_social = _is_marked_yes(row_data['in_social'])
        in_ledger = _is_marked_yes(row_data['in_ledger'])
        if all([in_roster, in_social, in_ledger]):
            stats['three_way_consistent'] += 1
        else:
            stats['three_way_inconsistent'] += 1
            if is_qualified:
                if not in_roster:
                    errors.append({
                        'file': fname, 'row': r, 'field': 'roster',
                        'reason': f'第{r}行 {row_data["name"]} 符合条件但不在花名册中',
                    })
                if not in_social:
                    errors.append({
                        'file': fname, 'row': r, 'field': 'social',
                        'reason': f'第{r}行 {row_data["name"]} 符合条件但不在社保清单中',
                    })

    # 占比校验
    ratio_value = None
    for label in SUMMARY_LABELS:
        if '占比' in label and label in summary_rows:
            ratio_value = _parse_ratio(summary_rows[label]['value'])
            break

    if ratio_value is not None:
        stats['staff_ratio'] = ratio_value
        if ratio_value < MIN_STAFF_RATIO:
            errors.append({
                'file': fname, 'row': 0, 'field': 'staff_ratio',
                'reason': f'科技人员占比 {ratio_value*100:.2f}% < {MIN_STAFF_RATIO*100:.0f}%',
            })
    else:
        total_employees = None
        if project_index and project_index.get('data_summary'):
            ds = project_index['data_summary']
            if isinstance(ds, dict):
                total_employees = ds.get('total_employees') or ds.get('staff_count')
        if total_employees and total_employees > 0 and stats['qualified_staff'] > 0:
            ratio = stats['qualified_staff'] / total_employees
            stats['staff_ratio'] = ratio
            stats['total_employees'] = total_employees
            if ratio < MIN_STAFF_RATIO:
                errors.append({
                    'file': fname, 'row': 0, 'field': 'staff_ratio',
                    'reason': f'科技人员占比 {ratio*100:.2f}% < {MIN_STAFF_RATIO*100:.0f}%',
                })

    passed = len(errors) == 0
    return {
        'passed': passed,
        'errors': errors,
        'warnings': warnings,
        'stats': stats,
    }


def main():
    """CLI 入口：解析参数并执行科技人员材料审核"""
    parser = argparse.ArgumentParser(
        description='科技人员材料审核脚本 - 校验高企认定科技人员证明材料合规性',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog='''示例：
  python validate_staff.py --dir "人员材料目录"
  python validate_staff.py --dir "人员材料目录" --project-root "项目根目录"
  python validate_staff.py --file "科技人员清单.xlsx" --project-root "项目根目录"
''',
    )
    parser.add_argument('--dir', help='科技人员材料所在目录')
    parser.add_argument('--file', help='单个科技人员清单 xlsx 文件路径')
    parser.add_argument('--project-root', dest='project_root',
                        help='项目根目录（用于读取 project_index.json 获取企业总人数等数据）')
    args = parser.parse_args()

    if not args.dir and not args.file:
        parser.print_help()
        sys.exit(1)

    if args.dir:
        report = validate_staff_directory(args.dir, args.project_root)
    else:
        report = validate_single_staff_file(args.file, args.project_root)

    # 输出 JSON 格式审核报告
    print(json.dumps(report, ensure_ascii=False, indent=2, default=str))
    sys.exit(0 if report['passed'] else 1)


if __name__ == '__main__':
    main()
