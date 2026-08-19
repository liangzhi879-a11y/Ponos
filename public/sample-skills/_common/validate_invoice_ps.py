#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""发票 PS 筛选结果审核脚本

功能说明：
    审核 invoice_ps_matcher.py 生成的发票 PS 筛选结果的完整性和合规性。

核心审核规则：
    1. PS 基线来源校验：检查 PS 基线是否来自申请书（非发票反推）
    2. 匹配覆盖率校验：已匹配发票金额占比是否合理（≥60%为高新收入达标）
    3. 未匹配 PS 校验：申请书 PS 中是否有完全无发票匹配的 PS（需确认是否取消）
    4. 未匹配发票校验：是否有大额发票未匹配到任何 PS（需确认是否扩展 PS）
    5. PS 发票标注表完整性：每个 PS 的发票清单是否完整
    6. 统计表数据一致性：统计表金额与标注表明细是否一致
    7. 占比分析报告：高新收入占比是否≥60%

CLI 用法：
    python validate_invoice_ps.py --dir "输出目录"
    python validate_invoice_ps.py --dir "输出目录" --project-root "项目根目录"
    python validate_invoice_ps.py --dir "输出目录" --project-root "项目根目录" --threshold 0.6

输出目录需包含：
    - PS发票标注表.xlsx   （来自 generate-report）
    - PS统计表.xlsx       （来自 generate-report）
    - 占比分析报告.json    （来自 generate-report）
    - 匹配结果.json        （可选，匹配阶段输出，用于深度校验）
    - PS基线.json          （可选，extract-ps-baseline 阶段输出，用于来源校验）

审核报告格式：
    {
      "passed": bool,
      "errors": [{"rule": "...", "file": "...", "reason": "..."}],
      "warnings": [{"rule": "...", "file": "...", "reason": "..."}],
      "stats": {...}
    }

退出码：审核通过 0，存在错误 1
依赖：openpyxl（用于读取 xlsx）；若未安装则跳过表格校验
"""

import os
import sys
import json
import argparse
from datetime import datetime

# ============================================================
# 依赖加载
# ============================================================

try:
    import openpyxl
    _OPENPYXL_AVAILABLE = True
except ImportError:
    _OPENPYXL_AVAILABLE = False


# ============================================================
# 常量定义
# ============================================================

# 高新收入占比达标线（可被 CLI --threshold 覆盖）
DEFAULT_RATIO_THRESHOLD = 0.60

# 大额未匹配发票告警阈值（默认 10000 元）
LARGE_INVOICE_THRESHOLD = 10000.0

# 输出目录中需要校验的文件名
FILE_MARKED_TABLE = 'PS发票标注表.xlsx'
FILE_SUMMARY_TABLE = 'PS统计表.xlsx'
FILE_RATIO_REPORT = '占比分析报告.json'
FILE_MATCH_RESULT = '匹配结果.json'
FILE_PS_BASELINE = 'PS基线.json'

# PS 基线来源白名单（必须是"申请书"，不能是发票反推）
VALID_PS_SOURCES = {'申请书', '申请材料', '高新申报书'}

# 发票反推来源特征（出现在 source 字段中表示违规）
INVALID_PS_SOURCE_MARKERS = ['发票', '货物', '反推', '推断']


# ============================================================
# 工具函数
# ============================================================

def _to_float(value, default=0.0):
    """安全转浮点"""
    if value is None:
        return default
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip().replace(',', '').replace('，', '').replace('%', '')
    try:
        return float(s) if s else default
    except ValueError:
        return default


def _parse_ratio(value):
    """解析占比字符串或小数（"60.00%" → 0.6，0.6 → 0.6）"""
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip().replace('%', '').replace(',', '')
    try:
        v = float(s) if s else 0.0
        return v / 100.0 if v > 1.0 else v
    except ValueError:
        return 0.0


def _load_json_safe(path):
    """安全加载 JSON 文件，失败返回 None"""
    if not os.path.exists(path):
        return None
    try:
        with open(path, 'r', encoding='utf-8') as f:
            return json.load(f)
    except (json.JSONDecodeError, OSError):
        return None


def _read_xlsx_safe(path):
    """安全加载 xlsx，返回 workbook；失败返回 None"""
    if not _OPENPYXL_AVAILABLE:
        return None
    if not os.path.exists(path):
        return None
    try:
        return openpyxl.load_workbook(path, data_only=True)
    except Exception:
        return None


# ============================================================
# 各项校验规则
# ============================================================

def check_ps_baseline_source(dir_path, project_root=None):
    """规则1：PS 基线来源校验

    检查 PS 基线是否来自申请书（非发票反推）。
    校验对象：输出目录或项目根目录下的 PS基线.json。

    Args:
        dir_path: 输出目录
        project_root: 项目根目录（备用查找位置）

    Returns:
        dict: {"errors": [...], "warnings": [...], "stats": {...}}
    """
    errors = []
    warnings = []
    stats = {}

    # 查找 PS 基线 JSON
    candidates = [
        os.path.join(dir_path, FILE_PS_BASELINE),
    ]
    if project_root:
        candidates.append(os.path.join(project_root, '.trae', 'skills', '_common', FILE_PS_BASELINE))
        candidates.append(os.path.join(project_root, FILE_PS_BASELINE))

    baseline = None
    baseline_path = None
    for c in candidates:
        baseline = _load_json_safe(c)
        if baseline:
            baseline_path = c
            break

    if baseline is None:
        warnings.append({
            'rule': 'PS基线来源校验',
            'file': FILE_PS_BASELINE,
            'reason': '未找到 PS基线.json，无法校验 PS 来源（请确认是否已运行 extract-ps-baseline）',
        })
        return {'errors': errors, 'warnings': warnings,
                'stats': {'baseline_found': False}}

    source = baseline.get('source', '')
    stats['baseline_found'] = True
    stats['baseline_source'] = source
    stats['baseline_path'] = baseline_path
    stats['ps_count'] = len(baseline.get('ps_items', []))

    # 校验来源
    if not source:
        errors.append({
            'rule': 'PS基线来源校验',
            'file': baseline_path,
            'reason': 'PS 基线 source 字段为空，无法确认是否来自申请书',
        })
    elif source not in VALID_PS_SOURCES:
        # 检查是否含发票反推特征
        if any(marker in source for marker in INVALID_PS_SOURCE_MARKERS):
            errors.append({
                'rule': 'PS基线来源校验',
                'file': baseline_path,
                'reason': f'PS 基线来源 "{source}" 含发票反推特征，违反核心原则（PS 名称必须来源于申请书）',
            })
        else:
            warnings.append({
                'rule': 'PS基线来源校验',
                'file': baseline_path,
                'reason': f'PS 基线来源 "{source}" 不在白名单 {VALID_PS_SOURCES} 内，请人工确认',
            })

    # 校验 PS 项是否非空
    if not baseline.get('ps_items'):
        errors.append({
            'rule': 'PS基线来源校验',
            'file': baseline_path,
            'reason': 'PS 基线 ps_items 为空，无法进行后续匹配',
        })

    return {'errors': errors, 'warnings': warnings, 'stats': stats}


def check_match_coverage(dir_path, threshold=DEFAULT_RATIO_THRESHOLD):
    """规则2：匹配覆盖率校验

    检查已匹配发票金额占比是否合理（≥阈值视为达标）。
    优先从占比分析报告.json 读取，其次从匹配结果.json 读取。

    Args:
        dir_path: 输出目录
        threshold: 高新收入占比达标线（默认 0.60）

    Returns:
        dict: {"errors": [...], "warnings": [...], "stats": {...}}
    """
    errors = []
    warnings = []
    stats = {}

    ratio_report = _load_json_safe(os.path.join(dir_path, FILE_RATIO_REPORT))
    match_result = _load_json_safe(os.path.join(dir_path, FILE_MATCH_RESULT))

    coverage = None
    matched_amount = 0.0
    total_amount = 0.0

    if ratio_report:
        coverage = ratio_report.get('high_tech_ratio')
        matched_amount = ratio_report.get('high_tech_revenue', 0.0)
        total_amount = ratio_report.get('total_invoice_amount', 0.0)
        stats['source'] = FILE_RATIO_REPORT
    elif match_result:
        mstats = match_result.get('stats', {})
        matched_amount = mstats.get('matched_amount', 0.0)
        total_amount = mstats.get('total_invoice_amount', 0.0)
        coverage = match_result.get('coverage')
        stats['source'] = FILE_MATCH_RESULT

    if coverage is None:
        errors.append({
            'rule': '匹配覆盖率校验',
            'file': f'{FILE_RATIO_REPORT} / {FILE_MATCH_RESULT}',
            'reason': '未找到占比分析报告或匹配结果，无法计算覆盖率',
        })
        return {'errors': errors, 'warnings': warnings, 'stats': stats}

    coverage = _parse_ratio(coverage) if isinstance(coverage, str) else float(coverage)
    stats['coverage'] = round(coverage, 4)
    stats['matched_amount'] = round(matched_amount, 2)
    stats['total_amount'] = round(total_amount, 2)
    stats['threshold'] = threshold

    if coverage < threshold:
        errors.append({
            'rule': '匹配覆盖率校验',
            'file': FILE_RATIO_REPORT,
            'reason': f'高新收入占比 {coverage * 100:.2f}% 未达 {threshold * 100:.0f}% 达标线（差额 {(threshold - coverage) * 100:.2f}%）',
        })
    else:
        stats['qualified'] = True

    return {'errors': errors, 'warnings': warnings, 'stats': stats}


def check_unmatched_ps(dir_path):
    """规则3：未匹配 PS 校验

    检查申请书 PS 中是否有完全无发票匹配的 PS（需确认是否取消）。

    Args:
        dir_path: 输出目录

    Returns:
        dict: {"errors": [...], "warnings": [...], "stats": {...}}
    """
    errors = []
    warnings = []
    stats = {}

    match_result = _load_json_safe(os.path.join(dir_path, FILE_MATCH_RESULT))
    if match_result is None:
        warnings.append({
            'rule': '未匹配PS校验',
            'file': FILE_MATCH_RESULT,
            'reason': '未找到匹配结果.json，跳过未匹配 PS 校验',
        })
        return {'errors': errors, 'warnings': warnings, 'stats': stats}

    unmatched_ps = match_result.get('unmatched_ps', [])
    stats['unmatched_ps_count'] = len(unmatched_ps)
    stats['unmatched_ps_list'] = [
        {'ps_id': p.get('ps_id'), 'ps_name': p.get('ps_name')}
        for p in unmatched_ps
    ]

    if unmatched_ps:
        # 未匹配 PS 是 warning（不一定 error，因为可能是用户主动取消）
        for p in unmatched_ps:
            warnings.append({
                'rule': '未匹配PS校验',
                'file': FILE_MATCH_RESULT,
                'reason': f"PS {p.get('ps_id', '?')} {p.get('ps_name', '?')} 无任何发票匹配（请确认是否取消该 PS 或扩展匹配规则）",
            })

    return {'errors': errors, 'warnings': warnings, 'stats': stats}


def check_unmatched_invoices(dir_path, large_threshold=LARGE_INVOICE_THRESHOLD):
    """规则4：未匹配发票校验

    检查是否有大额发票未匹配到任何 PS（需确认是否扩展 PS）。

    Args:
        dir_path: 输出目录
        large_threshold: 大额发票告警阈值（默认 10000 元）

    Returns:
        dict: {"errors": [...], "warnings": [...], "stats": {...}}
    """
    errors = []
    warnings = []
    stats = {}

    match_result = _load_json_safe(os.path.join(dir_path, FILE_MATCH_RESULT))
    if match_result is None:
        warnings.append({
            'rule': '未匹配发票校验',
            'file': FILE_MATCH_RESULT,
            'reason': '未找到匹配结果.json，跳过未匹配发票校验',
        })
        return {'errors': errors, 'warnings': warnings, 'stats': stats}

    unmatched_invoices = match_result.get('unmatched_invoices', [])
    stats['unmatched_invoice_count'] = len(unmatched_invoices)

    # 大额未匹配发票
    large_unmatched = [
        inv for inv in unmatched_invoices
        if _to_float(inv.get('total_amount', 0.0)) >= large_threshold
    ]
    stats['large_unmatched_count'] = len(large_unmatched)
    stats['large_unmatched_amount'] = round(
        sum(_to_float(inv.get('total_amount', 0.0)) for inv in large_unmatched), 2
    )
    stats['large_threshold'] = large_threshold

    # 大额未匹配发票累计金额超过全量发票 5% → error；否则 warning
    total_amount = match_result.get('stats', {}).get('total_invoice_amount', 0.0)
    if total_amount > 0 and large_unmatched:
        ratio = stats['large_unmatched_amount'] / total_amount
        stats['large_unmatched_ratio'] = round(ratio, 4)

        for inv in large_unmatched:
            severity = 'error' if ratio >= 0.05 else 'warning'
            reason = (
                f"发票 {inv.get('invoice_no', '?')} 货物 {inv.get('goods_name', '?')} "
                f"金额 {_to_float(inv.get('total_amount', 0.0)):.2f} 元未匹配到任何 PS"
            )
            if severity == 'error':
                errors.append({
                    'rule': '未匹配发票校验',
                    'file': FILE_MATCH_RESULT,
                    'reason': reason + '（大额未匹配累计占比 ≥5%，建议扩展 PS 或调整匹配规则）',
                })
            else:
                warnings.append({
                    'rule': '未匹配发票校验',
                    'file': FILE_MATCH_RESULT,
                    'reason': reason,
                })

    return {'errors': errors, 'warnings': warnings, 'stats': stats}


def check_marked_table_completeness(dir_path):
    """规则5：PS 发票标注表完整性校验

    检查每个 PS 的发票清单是否完整（标注表 sheet 数与匹配结果 PS 数一致，
    且每个 sheet 的发票明细数与匹配结果一致）。

    Args:
        dir_path: 输出目录

    Returns:
        dict: {"errors": [...], "warnings": [...], "stats": {...}}
    """
    errors = []
    warnings = []
    stats = {}

    marked_path = os.path.join(dir_path, FILE_MARKED_TABLE)
    wb = _read_xlsx_safe(marked_path)
    if wb is None:
        if _OPENPYXL_AVAILABLE:
            errors.append({
                'rule': 'PS发票标注表完整性',
                'file': FILE_MARKED_TABLE,
                'reason': f'未找到或无法读取 {FILE_MARKED_TABLE}',
            })
        else:
            warnings.append({
                'rule': 'PS发票标注表完整性',
                'file': FILE_MARKED_TABLE,
                'reason': 'openpyxl 未安装，跳过标注表完整性校验',
            })
        return {'errors': errors, 'warnings': warnings, 'stats': stats}

    # 收集每个 sheet 的发票数量（标题行=1，表头行=2，数据从第3行起）
    sheet_counts = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # 跳过明显非数据 sheet
        if sheet_name in ('无匹配',):
            continue
        # 数据行数 = max_row - 2（标题+表头）
        data_rows = max(0, ws.max_row - 2)
        sheet_counts[sheet_name] = data_rows
    wb.close()

    stats['sheet_count'] = len(sheet_counts)
    stats['sheet_invoice_counts'] = sheet_counts

    # 与匹配结果对比（若可读）
    match_result = _load_json_safe(os.path.join(dir_path, FILE_MATCH_RESULT))
    if match_result:
        matched = match_result.get('matched', [])
        stats['matched_ps_count_in_json'] = len(matched)
        # 检查每个 PS 的发票数量一致性
        for bucket in matched:
            ps_id = bucket.get('ps_id', '')
            ps_name = bucket.get('ps_name', '')
            expected = len(bucket.get('invoices', []))
            # 在 sheet_counts 中查找匹配的 sheet（按 ps_id 前缀）
            actual = None
            for sname, cnt in sheet_counts.items():
                if sname.startswith(ps_id) or ps_id in sname:
                    actual = cnt
                    break
            if actual is None:
                errors.append({
                    'rule': 'PS发票标注表完整性',
                    'file': FILE_MARKED_TABLE,
                    'reason': f'PS {ps_id} {ps_name} 在标注表中找不到对应 sheet',
                })
            elif actual != expected:
                errors.append({
                    'rule': 'PS发票标注表完整性',
                    'file': FILE_MARKED_TABLE,
                    'reason': f'PS {ps_id} {ps_name} 标注表发票数 ({actual}) 与匹配结果 ({expected}) 不一致',
                })
    else:
        warnings.append({
            'rule': 'PS发票标注表完整性',
            'file': FILE_MATCH_RESULT,
            'reason': '未找到匹配结果.json，仅校验标注表存在性，未做数量一致性校验',
        })

    return {'errors': errors, 'warnings': warnings, 'stats': stats}


def check_summary_consistency(dir_path):
    """规则6：统计表数据一致性校验

    检查统计表金额与标注表明细是否一致：
      - 统计表合计行金额 vs 占比分析报告中的高新收入
      - 统计表合计行金额 vs 标注表各 sheet 发票合计

    Args:
        dir_path: 输出目录

    Returns:
        dict: {"errors": [...], "warnings": [...], "stats": {...}}
    """
    errors = []
    warnings = []
    stats = {}

    summary_path = os.path.join(dir_path, FILE_SUMMARY_TABLE)
    wb = _read_xlsx_safe(summary_path)
    if wb is None:
        if _OPENPYXL_AVAILABLE:
            errors.append({
                'rule': '统计表数据一致性',
                'file': FILE_SUMMARY_TABLE,
                'reason': f'未找到或无法读取 {FILE_SUMMARY_TABLE}',
            })
        else:
            warnings.append({
                'rule': '统计表数据一致性',
                'file': FILE_SUMMARY_TABLE,
                'reason': 'openpyxl 未安装，跳过统计表一致性校验',
            })
        return {'errors': errors, 'warnings': warnings, 'stats': stats}

    # 读取统计表（跳过第1行标题，第2行为表头）
    ws = wb.active
    headers = []
    for c in range(1, ws.max_column + 1):
        headers.append(str(ws.cell(2, c).value or '').strip())

    # 定位列
    col_idx = {}
    for i, h in enumerate(headers, 1):
        if '价税合计' in h and 'total_col' not in col_idx:
            col_idx['total_col'] = i
        elif '金额' in h and 'amount_col' not in col_idx and '合计' not in h:
            col_idx['amount_col'] = i
        elif '发票数量' in h or '数量' in h:
            col_idx['count_col'] = i
        elif 'PS编号' in h or h == 'PS编号':
            col_idx['ps_id_col'] = i

    # 累加各 PS 行（直到合计行）
    summary_total = 0.0
    summary_amount = 0.0
    summary_count = 0
    ps_rows = []
    for r in range(3, ws.max_row + 1):
        first_cell = str(ws.cell(r, 1).value or '').strip()
        if first_cell in ('合计', '总计'):
            # 合计行
            if 'total_col' in col_idx:
                summary_total = _to_float(ws.cell(r, col_idx['total_col']).value)
            if 'amount_col' in col_idx:
                summary_amount = _to_float(ws.cell(r, col_idx['amount_col']).value)
            if 'count_col' in col_idx:
                summary_count = _to_float(ws.cell(r, col_idx['count_col']).value)
            break
        ps_id = first_cell
        total = _to_float(ws.cell(r, col_idx.get('total_col', 5)).value) if 'total_col' in col_idx else 0.0
        amount = _to_float(ws.cell(r, col_idx.get('amount_col', 4)).value) if 'amount_col' in col_idx else 0.0
        count = _to_float(ws.cell(r, col_idx.get('count_col', 3)).value) if 'count_col' in col_idx else 0.0
        ps_rows.append({'ps_id': ps_id, 'total': total, 'amount': amount, 'count': count})
    wb.close()

    # 各 PS 行累计
    sum_of_rows_total = round(sum(p['total'] for p in ps_rows), 2)
    sum_of_rows_amount = round(sum(p['amount'] for p in ps_rows), 2)
    sum_of_rows_count = int(sum(p['count'] for p in ps_rows))

    stats['summary_total'] = summary_total
    stats['summary_amount'] = summary_amount
    stats['summary_count'] = summary_count
    stats['sum_of_rows_total'] = sum_of_rows_total
    stats['sum_of_rows_amount'] = sum_of_rows_amount
    stats['sum_of_rows_count'] = sum_of_rows_count

    # 校验：各 PS 行累计 = 合计行
    if abs(sum_of_rows_total - summary_total) > 0.5:
        errors.append({
            'rule': '统计表数据一致性',
            'file': FILE_SUMMARY_TABLE,
            'reason': f'统计表各 PS 价税合计累计 ({sum_of_rows_total}) 与合计行 ({summary_total}) 不一致',
        })
    if abs(sum_of_rows_amount - summary_amount) > 0.5:
        warnings.append({
            'rule': '统计表数据一致性',
            'file': FILE_SUMMARY_TABLE,
            'reason': f'统计表各 PS 金额累计 ({sum_of_rows_amount}) 与合计行 ({summary_amount}) 不一致',
        })
    if sum_of_rows_count != int(summary_count):
        errors.append({
            'rule': '统计表数据一致性',
            'file': FILE_SUMMARY_TABLE,
            'reason': f'统计表各 PS 发票数量累计 ({sum_of_rows_count}) 与合计行 ({int(summary_count)}) 不一致',
        })

    # 与占比分析报告对比
    ratio_report = _load_json_safe(os.path.join(dir_path, FILE_RATIO_REPORT))
    if ratio_report:
        ratio_high_tech = ratio_report.get('high_tech_revenue', 0.0)
        stats['ratio_high_tech_revenue'] = ratio_high_tech
        if abs(summary_total - ratio_high_tech) > 0.5:
            errors.append({
                'rule': '统计表数据一致性',
                'file': FILE_SUMMARY_TABLE,
                'reason': f'统计表合计行价税合计 ({summary_total}) 与占比分析报告高新收入 ({ratio_high_tech}) 不一致',
            })

    # 与标注表对比
    marked_path = os.path.join(dir_path, FILE_MARKED_TABLE)
    marked_wb = _read_xlsx_safe(marked_path)
    if marked_wb:
        marked_total = 0.0
        marked_count = 0
        for sheet_name in marked_wb.sheetnames:
            if sheet_name == '无匹配':
                continue
            ws_m = marked_wb[sheet_name]
            # 数据从第3行起，价税合计列固定为第6列
            for r in range(3, ws_m.max_row + 1):
                val = _to_float(ws_m.cell(r, 6).value)
                if val > 0:
                    marked_total += val
                    marked_count += 1
        marked_wb.close()
        marked_total = round(marked_total, 2)
        stats['marked_total'] = marked_total
        stats['marked_count'] = marked_count
        if abs(marked_total - summary_total) > 0.5:
            errors.append({
                'rule': '统计表数据一致性',
                'file': FILE_SUMMARY_TABLE,
                'reason': f'统计表合计 ({summary_total}) 与标注表明细合计 ({marked_total}) 不一致',
            })
        if marked_count != sum_of_rows_count:
            errors.append({
                'rule': '统计表数据一致性',
                'file': FILE_SUMMARY_TABLE,
                'reason': f'统计表发票数量合计 ({sum_of_rows_count}) 与标注表发票数 ({marked_count}) 不一致',
            })

    return {'errors': errors, 'warnings': warnings, 'stats': stats}


def check_ratio_report(dir_path, threshold=DEFAULT_RATIO_THRESHOLD):
    """规则7：占比分析报告校验

    检查高新收入占比是否≥阈值，以及报告字段完整性。

    Args:
        dir_path: 输出目录
        threshold: 高新收入占比达标线

    Returns:
        dict: {"errors": [...], "warnings": [...], "stats": {...}}
    """
    errors = []
    warnings = []
    stats = {}

    ratio_report = _load_json_safe(os.path.join(dir_path, FILE_RATIO_REPORT))
    if ratio_report is None:
        errors.append({
            'rule': '占比分析报告',
            'file': FILE_RATIO_REPORT,
            'reason': f'未找到或无法读取 {FILE_RATIO_REPORT}',
        })
        return {'errors': errors, 'warnings': warnings, 'stats': stats}

    # 字段完整性
    required_fields = ['high_tech_revenue', 'total_invoice_amount', 'high_tech_ratio',
                       'threshold', 'qualified', 'conclusion']
    missing = [f for f in required_fields if f not in ratio_report]
    if missing:
        errors.append({
            'rule': '占比分析报告',
            'file': FILE_RATIO_REPORT,
            'reason': f'占比分析报告缺少字段: {missing}',
        })

    ratio = _parse_ratio(ratio_report.get('high_tech_ratio'))
    stats['high_tech_ratio'] = ratio
    stats['high_tech_revenue'] = ratio_report.get('high_tech_revenue', 0.0)
    stats['total_invoice_amount'] = ratio_report.get('total_invoice_amount', 0.0)
    stats['qualified'] = ratio_report.get('qualified')
    stats['threshold'] = threshold

    # 占比是否≥阈值
    if ratio < threshold:
        errors.append({
            'rule': '占比分析报告',
            'file': FILE_RATIO_REPORT,
            'reason': f'高新收入占比 {ratio * 100:.2f}% 未达 {threshold * 100:.0f}% 达标线',
        })
    else:
        stats['qualified'] = True

    # qualified 字段与实际计算一致性
    expected_qualified = ratio >= threshold
    if 'qualified' in ratio_report and ratio_report['qualified'] is not expected_qualified:
        errors.append({
            'rule': '占比分析报告',
            'file': FILE_RATIO_REPORT,
            'reason': f'qualified 字段 ({ratio_report["qualified"]}) 与实际计算 ({expected_qualified}) 不一致',
        })

    return {'errors': errors, 'warnings': warnings, 'stats': stats}


# ============================================================
# 主审核函数
# ============================================================

def validate_invoice_ps(dir_path, project_root=None, threshold=DEFAULT_RATIO_THRESHOLD,
                        large_invoice_threshold=LARGE_INVOICE_THRESHOLD):
    """执行全部 7 项校验

    Args:
        dir_path: 输出目录（含 PS发票标注表.xlsx / PS统计表.xlsx / 占比分析报告.json）
        project_root: 项目根目录（用于查找 PS 基线）
        threshold: 高新收入占比达标线
        large_invoice_threshold: 大额未匹配发票告警阈值

    Returns:
        dict: {
            "passed": bool,
            "errors": [...],
            "warnings": [...],
            "stats": {...},
            "checked_at": "ISO时间",
            "dir": "..."
        }
    """
    if not os.path.isdir(dir_path):
        return {
            'passed': False,
            'errors': [{'rule': '目录校验', 'file': dir_path,
                        'reason': f'输出目录不存在: {dir_path}'}],
            'warnings': [],
            'stats': {},
            'checked_at': datetime.now().isoformat(timespec='seconds'),
            'dir': dir_path,
        }

    all_errors = []
    all_warnings = []
    all_stats = {}

    # 执行 7 项校验
    checks = [
        ('ps_baseline_source', lambda: check_ps_baseline_source(dir_path, project_root)),
        ('match_coverage', lambda: check_match_coverage(dir_path, threshold)),
        ('unmatched_ps', lambda: check_unmatched_ps(dir_path)),
        ('unmatched_invoices', lambda: check_unmatched_invoices(dir_path, large_invoice_threshold)),
        ('marked_table_completeness', lambda: check_marked_table_completeness(dir_path)),
        ('summary_consistency', lambda: check_summary_consistency(dir_path)),
        ('ratio_report', lambda: check_ratio_report(dir_path, threshold)),
    ]

    for name, fn in checks:
        try:
            result = fn()
            all_errors.extend(result.get('errors', []))
            all_warnings.extend(result.get('warnings', []))
            all_stats[name] = result.get('stats', {})
        except Exception as e:
            all_errors.append({
                'rule': name,
                'file': '',
                'reason': f'校验异常: {type(e).__name__}: {e}',
            })

    return {
        'passed': len(all_errors) == 0,
        'errors': all_errors,
        'warnings': all_warnings,
        'stats': all_stats,
        'checked_at': datetime.now().isoformat(timespec='seconds'),
        'dir': os.path.abspath(dir_path),
        'threshold': threshold,
        'error_count': len(all_errors),
        'warning_count': len(all_warnings),
    }


# ============================================================
# CLI
# ============================================================

def build_parser():
    """构建 argparse 解析器"""
    parser = argparse.ArgumentParser(
        prog='validate_invoice_ps',
        description='发票 PS 筛选结果审核脚本（校验 7 项规则）',
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument('--dir', required=True, help='输出目录（含 PS发票标注表/PS统计表/占比分析报告）')
    parser.add_argument('--project-root', default=None, help='项目根目录（用于查找 PS 基线）')
    parser.add_argument('--threshold', type=float, default=DEFAULT_RATIO_THRESHOLD,
                        help=f'高新收入占比达标线（默认 {DEFAULT_RATIO_THRESHOLD}）')
    parser.add_argument('--large-invoice-threshold', type=float,
                        default=LARGE_INVOICE_THRESHOLD,
                        help=f'大额未匹配发票告警阈值（默认 {LARGE_INVOICE_THRESHOLD} 元）')
    parser.add_argument('--output', default=None, help='审核报告 JSON 输出路径（省略则输出到 stdout）')
    return parser


def main():
    """CLI 主入口"""
    parser = build_parser()
    args = parser.parse_args()

    report = validate_invoice_ps(
        dir_path=args.dir,
        project_root=args.project_root,
        threshold=args.threshold,
        large_invoice_threshold=args.large_invoice_threshold,
    )

    # 输出
    if args.output:
        os.makedirs(os.path.dirname(os.path.abspath(args.output)) or '.', exist_ok=True)
        with open(args.output, 'w', encoding='utf-8') as f:
            json.dump(report, f, ensure_ascii=False, indent=2)
        print(f'[OK] 审核报告已保存：{args.output}')
    else:
        print(json.dumps(report, ensure_ascii=False, indent=2))

    # 摘要
    print(f'\n[摘要] 通过={report["passed"]}，错误={report["error_count"]}，'
          f'警告={report["warning_count"]}', file=sys.stderr)
    if report['errors']:
        print('\n[错误项]', file=sys.stderr)
        for e in report['errors']:
            print(f'  - [{e.get("rule", "?")}] {e.get("reason", "?")}', file=sys.stderr)
    if report['warnings']:
        print('\n[警告项]', file=sys.stderr)
        for w in report['warnings']:
            print(f'  - [{w.get("rule", "?")}] {w.get("reason", "?")}', file=sys.stderr)

    return 0 if report['passed'] else 1


if __name__ == '__main__':
    sys.exit(main())
