#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
contract_invoice_extractor.py - 合同发票字段专用提取器

从OCR文本中提取合同/发票的结构化字段，输出JSON+Excel双格式。

合同字段：合同编号、甲方、乙方、签订日期、合同金额、标的物、有效期
发票字段：发票号、开票日期、购方名称、销方名称、货物明细、金额、税额、价税合计

用法：
  python contract_invoice_extractor.py extract --file "合同.pdf" --output-dir "提取结果/"
  python contract_invoice_extractor.py batch --dir "发票目录/" --output-dir "提取结果/"
  python contract_invoice_extractor.py detect --file "未知文件.pdf"

设计逻辑（agent可据此自主调改）：
  extract: 调用ocr_engine识别 → 检测文档类型 → 提取字段 → 保存JSON+Excel
  batch: 遍历目录下所有PDF，逐个extract
  detect: 仅检测文档类型，不提取字段
"""

import argparse
import json
import os
import re
import sys
from datetime import datetime
from pathlib import Path

# 确保 _common 目录在 sys.path 中（支持直接运行和外部调用）
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

COMMON_DIR = Path(__file__).parent


# ============================================================
# 正则模式库
# ============================================================

# 日期正则：覆盖 OCR 输出的各种日期格式
_DATE_PATTERNS = [
    # 中文格式带"日"
    r'(20\d{2})\s*年\s*(\d{1,2})\s*月\s*(\d{1,2})\s*日',
    # 中文格式无"日"（合同/发票常见）
    r'(20\d{2})\s*年\s*(\d{1,2})\s*月\s*(\d{1,2})(?:日|号)?',
    # ISO格式
    r'(20\d{2})-(\d{1,2})-(\d{1,2})',
    # 斜杠分隔
    r'(20\d{2})/(\d{1,2})/(\d{1,2})',
    # 点分隔（OCR常见）
    r'(20\d{2})\.(\d{1,2})\.(\d{1,2})',
    # 空格分隔
    r'(20\d{2})\s+(\d{1,2})\s+(\d{1,2})(?!\d)',
    # 无分隔连续8位
    r'(20\d{2})(0[1-9]|1[0-2])(0[1-9]|[12]\d|3[01])',
]

def _find_date_in_text(text, prefix_patterns=None):
    """在文本中搜索日期，适配多种OCR输出格式

    Args:
        text: OCR文本
        prefix_patterns: 日期前缀关键词列表，如 ['签订日期', '开票日期']

    Returns:
        (year, month, day) 或 (None, None, None)
    """
    search_text = text

    # 如果指定了前缀，先截取前缀后的文本区域
    if prefix_patterns:
        for prefix in prefix_patterns:
            m = re.search(rf'{prefix}\s*[:：]?\s*', search_text)
            if m:
                # 取前缀后100字符进行日期匹配
                start = m.end()
                search_text = text[start:start + 100]
                break

    for pattern in _DATE_PATTERNS:
        m = re.search(pattern, search_text)
        if m:
            year = int(m.group(1))
            month = int(m.group(2))
            day = int(m.group(3))
            # 基本校验
            if 2000 <= year <= 2099 and 1 <= month <= 12 and 1 <= day <= 31:
                return year, month, day
    return None, None, None


CONTRACT_PATTERNS = {
    "contract_no": [
        r'合同编号[:：\s]*([A-Za-z0-9\-/]+)',
        r'合同号[:：\s]*([A-Za-z0-9\-/]+)',
        r'编号[:：\s]*([A-Za-z0-9\-/]+)',
    ],
    "party_a": [
        r'甲方[:：\s]*([\u4e00-\u9fa5（）()股份有限公司有限责任公司集团]+)',
        r'买方[:：\s]*([\u4e00-\u9fa5（）()股份有限公司有限责任公司集团]+)',
        r'发包方[:：\s]*([\u4e00-\u9fa5（）()股份有限公司有限责任公司集团]+)',
        r'委托方[:：\s]*([\u4e00-\u9fa5（）()股份有限公司有限责任公司集团]+)',
    ],
    "party_b": [
        r'乙方[:：\s]*([\u4e00-\u9fa5（）()股份有限公司有限责任公司集团]+)',
        r'卖方[:：\s]*([\u4e00-\u9fa5（）()股份有限公司有限责任公司集团]+)',
        r'承包方[:：\s]*([\u4e00-\u9fa5（）()股份有限公司有限责任公司集团]+)',
        r'受托方[:：\s]*([\u4e00-\u9fa5（）()股份有限公司有限责任公司集团]+)',
    ],
    "sign_date": [
        # 前缀引导的日期匹配（使用通用日期函数）
        r'签订日期[:：\s]*.*',
        r'签署日期[:：\s]*.*',
        r'签字日期[:：\s]*.*',
    ],
    "amount": [
        r'合同[金额总价][:：\s]*人民币[￥¥]?\s*([\d,]+\.?\d*)\s*[元圆]',
        r'合同[金额总价][:：\s]*[￥¥]\s*([\d,]+\.?\d*)',
        r'总[金额价][:：\s]*人民币[￥¥]?\s*([\d,]+\.?\d*)\s*[元圆]',
        r'金额[为大写]*[:：\s]*人民币[￥¥]?\s*([\d,]+\.?\d*)\s*[元圆]',
    ],
    "subject": [
        r'标的[:：\s]*([\u4e00-\u9fa5A-Za-z0-9\-/、，,。.（）()]+?)(?:\n|$)',
        r'项目名称[:：\s]*([\u4e00-\u9fa5A-Za-z0-9\-/、，,。.（）()]+?)(?:\n|$)',
        r'合同标的[:：\s]*([\u4e00-\u9fa5A-Za-z0-9\-/、，,。.（）()]+?)(?:\n|$)',
    ],
}

INVOICE_PATTERNS = {
    "invoice_no": [
        r'发票号码[:：\s]*([0-9]{8,20})',
        r'Invoice\s*No[:：\s]*([A-Z0-9]+)',
        r'发票代码[:：\s]*([0-9]{10,12})\s*发票号码[:：\s]*([0-9]{8,20})',
    ],
    "invoice_date": [
        r'开票日期[:：\s]*.*',
        r'开具日期[:：\s]*.*',
    ],
    "buyer_name": [
        r'购方[:：\s]*名称[:：\s]*([\u4e00-\u9fa5（）()股份有限公司有限责任公司集团]+)',
        r'买方[:：\s]*名称[:：\s]*([\u4e00-\u9fa5（）()股份有限公司有限责任公司集团]+)',
        r'名称[:：\s]*([\u4e00-\u9fa5（）()股份有限公司有限责任公司集团]+)\s*纳税人识别号',
    ],
    "seller_name": [
        r'销售方[:：\s]*名称[:：\s]*([\u4e00-\u9fa5（）()股份有限公司有限责任公司集团]+)',
        r'销方[:：\s]*名称[:：\s]*([\u4e00-\u9fa5（）()股份有限公司有限责任公司集团]+)',
        r'销售方信息\s*名称[:：\s]*([\u4e00-\u9fa5（）()股份有限公司有限责任公司集团]+)',
    ],
    "buyer_tax_id": [
        r'购方.*?纳税人识别号[:：\s]*([A-Z0-9]{15,20})',
        r'买方.*?纳税人识别号[:：\s]*([A-Z0-9]{15,20})',
    ],
    "seller_tax_id": [
        r'销售方.*?纳税人识别号[:：\s]*([A-Z0-9]{15,20})',
        r'销方.*?纳税人识别号[:：\s]*([A-Z0-9]{15,20})',
    ],
    "total_amount": [
        r'金额[:：\s]*[￥¥]?\s*([\d,]+\.?\d*)',
        r'合计金额[:：\s]*[￥¥]?\s*([\d,]+\.?\d*)',
    ],
    "total_tax": [
        r'税额[:：\s]*[￥¥]?\s*([\d,]+\.?\d*)',
        r'合计税额[:：\s]*[￥¥]?\s*([\d,]+\.?\d*)',
    ],
    "total_with_tax": [
        r'价税合计[（(]大写[)）][:：\s]*[￥¥]?\s*([\d,]+\.?\d*)',
        r'价税合计[:：\s]*[￥¥]?\s*([\d,]+\.?\d*)',
        r'价税合计[（(]小写[)）][￥¥]?\s*([\d,]+\.?\d*)',
    ],
}


# ============================================================
# 文档类型检测
# ============================================================
def detect_document_type(ocr_text, filename=""):
    """检测文档类型：contract / invoice / unknown"""
    text_lower = ocr_text.lower()
    name_lower = filename.lower()

    # 发票关键词
    invoice_keywords = ['发票号码', '发票代码', '开票日期', '价税合计', '税额', '纳税人识别号',
                        '增值税', '专用发票', '普通发票', '电子发票', 'invoice']
    invoice_score = sum(1 for kw in invoice_keywords if kw in ocr_text or kw.lower() in text_lower)
    if '发票' in filename or 'invoice' in name_lower or 'dzfp' in name_lower:
        invoice_score += 3

    # 合同关键词
    contract_keywords = ['合同编号', '甲方', '乙方', '签订日期', '合同金额', '合同标的',
                         '协议', 'contract', 'agreement', '发包方', '承包方', '委托方']
    contract_score = sum(1 for kw in contract_keywords if kw in ocr_text or kw.lower() in text_lower)
    if '合同' in filename or 'contract' in name_lower or '协议' in filename:
        contract_score += 3

    if invoice_score > contract_score and invoice_score >= 2:
        return "invoice"
    elif contract_score > invoice_score and contract_score >= 2:
        return "contract"
    elif invoice_score == contract_score and invoice_score > 0:
        # 平局时优先发票（发票有专属字段）
        return "invoice" if '发票' in filename else "contract"
    return "unknown"


# ============================================================
# 合同字段提取
# ============================================================
def extract_contract_fields(ocr_text, ocr_tables=None):
    """提取合同字段"""
    fields = {
        "document_type": "contract",
        "contract_no": None,
        "party_a": None,
        "party_b": None,
        "sign_date": None,
        "sign_year": None,
        "amount": None,
        "amount_numeric": None,
        "subject": None,
        "effective_date": None,
        "expiry_date": None,
        "confidence": 0.0,
    }

    matched = 0
    total = 0

    for field, patterns in CONTRACT_PATTERNS.items():
        total += 1
        # sign_date 和 effective_date 用通用日期提取
        if field in ("sign_date", "effective_date", "expiry_date"):
            for pattern in patterns:
                m = re.search(pattern, ocr_text, re.MULTILINE)
                if m:
                    prefix = m.group(0).split('：')[-1].split(':')[-1].strip()
                    if not prefix or len(prefix) < 2:
                        prefix = None
                    year, month, day = _find_date_in_text(ocr_text, [field.replace('_', '')])
                    if year:
                        fields[field] = f"{year}-{month:02d}-{day:02d}"
                        if field == "sign_date":
                            fields["sign_year"] = year
                        matched += 1
                    break
            continue

        for pattern in patterns:
            m = re.search(pattern, ocr_text, re.MULTILINE)
            if m:
                value = m.group(1).strip() if m.lastindex and m.lastindex >= 1 else m.group(0)
                if field == "amount":
                    amount_str = m.group(1).replace(',', '')
                    fields["amount"] = m.group(1)
                    try:
                        fields["amount_numeric"] = float(amount_str)
                    except ValueError:
                        pass
                else:
                    fields[field] = value
                matched += 1
                break

    fields["confidence"] = round(matched / total, 4) if total > 0 else 0
    return fields


# ============================================================
# 发票字段提取
# ============================================================
def extract_invoice_fields(ocr_text, ocr_tables=None):
    """提取发票字段"""
    fields = {
        "document_type": "invoice",
        "invoice_no": None,
        "invoice_date": None,
        "invoice_year": None,
        "buyer_name": None,
        "seller_name": None,
        "buyer_tax_id": None,
        "seller_tax_id": None,
        "items": [],
        "total_amount": None,
        "total_tax": None,
        "total_with_tax": None,
        "confidence": 0.0,
    }

    matched = 0
    total = 0

    for field, patterns in INVOICE_PATTERNS.items():
        if field in ("items",):
            continue
        total += 1
        # invoice_date 用通用日期提取
        if field == "invoice_date":
            for pattern in patterns:
                m = re.search(pattern, ocr_text, re.MULTILINE)
                if m:
                    year, month, day = _find_date_in_text(ocr_text, ['开票日期', '开具日期'])
                    if year:
                        fields["invoice_date"] = f"{year}-{month:02d}-{day:02d}"
                        fields["invoice_year"] = year
                        matched += 1
                    break
            continue

        for pattern in patterns:
            m = re.search(pattern, ocr_text, re.MULTILINE)
            if m:
                if field == "total_amount" or field == "total_tax" or field == "total_with_tax":
                    value_str = m.group(1).replace(',', '')
                    try:
                        fields[field] = float(value_str)
                    except ValueError:
                        fields[field] = m.group(1)
                else:
                    value = m.group(1).strip() if m.lastindex else m.group(0)
                    # 特殊处理：buyer_tax_id可能是group(2)
                    if field == "buyer_tax_id" and m.lastindex and m.lastindex >= 2:
                        value = m.group(2)
                    fields[field] = value
                matched += 1
                break

    # 提取货物明细（从表格或文本）
    fields["items"] = _extract_invoice_items(ocr_text, ocr_tables)

    fields["confidence"] = round(matched / total, 4) if total > 0 else 0
    return fields


def _extract_invoice_items(ocr_text, ocr_tables=None):
    """从表格或文本中提取发票货物明细"""
    items = []

    # 优先从表格提取
    if ocr_tables:
        for table in ocr_tables:
            data = table.get("data", [])
            for row in data:
                if len(row) >= 4:
                    # 尝试识别货物名称行
                    first_cell = str(row[0])
                    if any(kw in first_cell for kw in ['货物名称', '项目', '规格型号', '品名']):
                        continue  # 表头
                    # 简单提取：假设列顺序为 货物名称/规格/数量/单价/金额/税率/税额
                    item = {
                        "name": row[0] if len(row) > 0 else "",
                        "spec": row[1] if len(row) > 1 else "",
                        "unit": row[2] if len(row) > 2 else "",
                        "quantity": row[3] if len(row) > 3 else "",
                        "unit_price": row[4] if len(row) > 4 else "",
                        "amount": row[5] if len(row) > 5 else "",
                        "tax_rate": row[6] if len(row) > 6 else "",
                        "tax_amount": row[7] if len(row) > 7 else "",
                    }
                    items.append(item)

    # 如果表格没提取到，从文本中提取
    if not items:
        # 简化的货物行识别
        item_pattern = re.compile(r'([\u4e00-\u9fa5A-Za-z][\u4e00-\u9fa5A-Za-z0-9\-/]{2,})\s+(\d+\.?\d*)\s+(\d+\.?\d*)\s+(\d+\.?\d*)\s*(\d+%?)\s+(\d+\.?\d*)')
        for m in item_pattern.finditer(ocr_text):
            items.append({
                "name": m.group(1),
                "quantity": m.group(2),
                "unit_price": m.group(3),
                "amount": m.group(4),
                "tax_rate": m.group(5),
                "tax_amount": m.group(6),
            })

    return items


# ============================================================
# 保存提取结果
# ============================================================
def save_extraction_result(result, output_dir, output_name=None):
    """保存JSON+Excel双格式"""
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    if output_name is None:
        output_name = f"extract_{datetime.now().strftime('%Y%m%d_%H%M%S')}"

    # 1. 保存JSON
    json_path = output_dir / f"{output_name}.json"
    json_path.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding='utf-8')

    # 2. 保存Excel
    excel_path = output_dir / f"{output_name}.xlsx"
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "提取结果"

        # 写入字段
        ws['A1'] = "字段"
        ws['B1'] = "值"
        ws['A1'].font = Font(bold=True)
        ws['B1'].font = Font(bold=True)

        row = 2
        for k, v in result.items():
            if k in ("items", "tables"):
                continue
            ws.cell(row=row, column=1, value=k)
            ws.cell(row=row, column=2, value=str(v) if v is not None else "")
            row += 1

        # 写入货物明细
        if result.get("items"):
            row += 2
            ws.cell(row=row, column=1, value="货物明细").font = Font(bold=True)
            row += 1
            headers = ["货物名称", "规格", "单位", "数量", "单价", "金额", "税率", "税额"]
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=h)
                cell.font = Font(bold=True)
            row += 1
            for item in result["items"]:
                ws.cell(row=row, column=1, value=item.get("name", ""))
                ws.cell(row=row, column=2, value=item.get("spec", ""))
                ws.cell(row=row, column=3, value=item.get("unit", ""))
                ws.cell(row=row, column=4, value=item.get("quantity", ""))
                ws.cell(row=row, column=5, value=item.get("unit_price", ""))
                ws.cell(row=row, column=6, value=item.get("amount", ""))
                ws.cell(row=row, column=7, value=item.get("tax_rate", ""))
                ws.cell(row=row, column=8, value=item.get("tax_amount", ""))
                row += 1

        wb.save(excel_path)
    except ImportError:
        print("[extract] WARN: openpyxl未安装，仅保存JSON")

    print(f"[extract] JSON已保存: {json_path}")
    print(f"[extract] Excel已保存: {excel_path}")
    return {"json_path": str(json_path), "excel_path": str(excel_path)}


# ============================================================
# 命令行入口
# ============================================================
def cmd_extract(args):
    """提取单个文件的字段（支持 PDF、PNG、JPG、JPEG）

    扫描件PDF自动通过RapidOCR识别，无需手动指定OCR工具。
    使用 --force-ocr 可强制重新OCR（忽略缓存）。
    """
    file_path = Path(args.file)
    if not file_path.exists():
        print(f"[extract] ERROR: 文件不存在: {args.file}")
        return False

    print(f"[extract] 处理文件: {file_path.name}")
    ext = file_path.suffix.lower()

    ocr_text = ""
    ocr_tables = []

    if ext in ('.png', '.jpg', '.jpeg'):
        try:
            from ocr_engine import ocr_image
            result = ocr_image(str(file_path), project_name=args.project)
            ocr_text = result.get("text", "")
            print(f"[extract] OCR完成（{len(ocr_text)}字符，confidence={result.get('confidence', 0)}）")
        except ImportError:
            print("[extract] ERROR: ocr_engine不可用，无法处理图片")
            return False
        except Exception as e:
            print(f"[extract] ERROR: 图片OCR失败: {e}")
            return False
    else:
        try:
            from ocr_engine import ocr_with_table, ocr_pdf

            if args.force_ocr:
                ocr_result = ocr_pdf(
                    str(file_path), force_ocr=True,
                    use_cache=False, project_name=args.project
                )
            else:
                ocr_result = ocr_with_table(str(file_path), project_name=args.project)

            ocr_text = ocr_result.get("text", "")
            ocr_tables = ocr_result.get("tables", [])

            is_scanned = ocr_result.get("is_scanned", False)
            strategy = ocr_result.get("strategy", "unknown")
            cache_hit = ocr_result.get("cache_hit", False)

            print(f"[extract] OCR完成（{len(ocr_text)}字符，策略={strategy}，"
                  f"扫描件={is_scanned}，缓存={cache_hit}，"
                  f"表格={len(ocr_tables)}个）")
        except ImportError:
            print("[extract] WARN: ocr_engine不可用，尝试直接读取文本层")
            try:
                import fitz
                doc = fitz.open(str(file_path))
                ocr_text = "\n".join(page.get_text() for page in doc)
                doc.close()
                ocr_tables = []
            except Exception as e:
                print(f"[extract] ERROR: 无法读取PDF: {e}")
                return False

    # 2. 检测文档类型
    doc_type = detect_document_type(ocr_text, file_path.name)
    print(f"[extract] 文档类型: {doc_type}")

    # 3. 提取字段
    if doc_type == "contract":
        fields = extract_contract_fields(ocr_text, ocr_tables)
    elif doc_type == "invoice":
        fields = extract_invoice_fields(ocr_text, ocr_tables)
    else:
        print("[extract] WARN: 无法识别文档类型，尝试两种提取")
        contract = extract_contract_fields(ocr_text, ocr_tables)
        invoice = extract_invoice_fields(ocr_text, ocr_tables)
        if contract["confidence"] > invoice["confidence"]:
            fields = contract
        else:
            fields = invoice

    fields["source_file"] = str(file_path)
    fields["extracted_at"] = datetime.now().isoformat()

    print(f"[extract] 字段提取完成（confidence={fields['confidence']}）")
    print(json.dumps(fields, ensure_ascii=False, indent=2))

    # 4. 保存结果
    output_name = file_path.stem
    paths = save_extraction_result(fields, args.output_dir, output_name)
    return True


def cmd_batch(args):
    """批量提取（支持 PDF、PNG、JPG、JPEG）"""
    input_dir = Path(args.dir)
    if not input_dir.is_dir():
        print(f"[batch] ERROR: 不是目录: {args.dir}")
        return False

    # 收集所有文件
    all_files = []
    for ext in ('*.pdf', '*.png', '*.jpg', '*.jpeg'):
        all_files.extend(list(input_dir.rglob(ext)))
    print(f"[batch] 发现 {len(all_files)} 个文件（PDF + 图片）")

    success = 0
    failed = 0
    for i, fp in enumerate(all_files, 1):
        print(f"\n[batch] ({i}/{len(all_files)}) 处理: {fp.name}")
        args.file = str(fp)
        try:
            if cmd_extract(args):
                success += 1
            else:
                failed += 1
        except Exception as e:
            print(f"[batch] ERROR: {e}")
            failed += 1

    print(f"\n[batch] 完成: 成功 {success}，失败 {failed}")
    return failed == 0


def cmd_detect(args):
    """检测文档类型（支持 PDF、PNG、JPG、JPEG）

    扫描件PDF自动通过RapidOCR提取文本后检测，无需用户关心底层OCR实现。
    """
    file_path = Path(args.file)
    if not file_path.exists():
        print(f"[detect] ERROR: 文件不存在: {args.file}")
        return False

    ext = file_path.suffix.lower()

    if ext in ('.png', '.jpg', '.jpeg'):
        try:
            from ocr_engine import ocr_image
            result = ocr_image(str(file_path))
            text = result.get("text", "")
        except Exception as e:
            print(f"[detect] ERROR: {e}")
            return False
    else:
        try:
            import fitz
            doc = fitz.open(str(file_path))
            text = "\n".join(page.get_text() for page in doc)
            doc.close()

            if len(text.strip()) < 50:
                from ocr_engine import ocr_pdf
                print("[detect] 文本层几乎为空，启动RapidOCR扫描...")
                ocr_result = ocr_pdf(str(file_path), project_name=args.project)
                text = ocr_result.get("text", text)
                print(f"[detect] OCR完成（{len(text)}字符）")
        except Exception as e:
            print(f"[detect] ERROR: {e}")
            return False

    doc_type = detect_document_type(text, file_path.name)
    print(f"文件: {file_path.name}")
    print(f"类型: {doc_type}")
    return True


def cmd_quick_ocr(args):
    """快速OCR提取：直接对扫描件PDF/图片执行OCR并输出发票号码等关键字段

    这是 extract 的简化版，输出纯文本，适合 agent 直接调用。
    """
    file_path = Path(args.file)
    if not file_path.exists():
        print(json.dumps({"error": f"文件不存在: {args.file}"}, ensure_ascii=False))
        return False

    ext = file_path.suffix.lower()

    # 一步到位：OCR
    try:
        from ocr_engine import ocr_pdf, ocr_image

        if ext in ('.png', '.jpg', '.jpeg'):
            result = ocr_image(str(file_path), project_name=args.project)
        else:
            result = ocr_pdf(
                str(file_path), force_ocr=args.force_ocr,
                use_cache=not args.force_ocr, project_name=args.project
            )

        ocr_text = result.get("text", "")
        strategy = result.get("strategy", "unknown")
        processing_time = result.get("processing_time_seconds", 0)

        if not ocr_text.strip():
            print(json.dumps({
                "error": "OCR识别结果为空，该文件可能确实无文字内容",
                "strategy": strategy,
                "time_seconds": processing_time,
            }, ensure_ascii=False))
            return False

        # 提取关键字段
        doc_type = detect_document_type(ocr_text, file_path.name)

        if doc_type == "invoice":
            fields = extract_invoice_fields(ocr_text, result.get("tables", []))
        elif doc_type == "contract":
            fields = extract_contract_fields(ocr_text, result.get("tables", []))
        else:
            invoice_fields = extract_invoice_fields(ocr_text, result.get("tables", []))
            contract_fields = extract_contract_fields(ocr_text, result.get("tables", []))
            if invoice_fields["confidence"] >= contract_fields["confidence"]:
                fields = invoice_fields
                doc_type = "invoice"
            else:
                fields = contract_fields
                doc_type = "contract"

        # 干净输出
        output = {
            "file": str(file_path),
            "document_type": doc_type,
            "ocr_strategy": strategy,
            "ocr_time_seconds": processing_time,
            "text_length": len(ocr_text),
            "key_fields": {},
        }

        if doc_type == "invoice":
            output["key_fields"] = {
                "invoice_no": fields.get("invoice_no"),
                "invoice_date": fields.get("invoice_date"),
                "buyer_name": fields.get("buyer_name"),
                "seller_name": fields.get("seller_name"),
                "total_amount": fields.get("total_amount"),
                "total_tax": fields.get("total_tax"),
                "total_with_tax": fields.get("total_with_tax"),
                "confidence": fields.get("confidence"),
            }
        else:
            output["key_fields"] = {
                "contract_no": fields.get("contract_no"),
                "party_a": fields.get("party_a"),
                "party_b": fields.get("party_b"),
                "sign_date": fields.get("sign_date"),
                "amount": fields.get("amount"),
                "subject": fields.get("subject"),
                "confidence": fields.get("confidence"),
            }

        if args.show_text:
            output["ocr_text_preview"] = ocr_text[:2000]
        if args.show_all:
            output["ocr_full_text"] = ocr_text
            output["all_fields"] = fields

        print(json.dumps(output, ensure_ascii=False, indent=2))
        return True

    except ImportError:
        print(json.dumps({"error": "ocr_engine不可用，请安装: pip install rapidocr-openvino"}, ensure_ascii=False))
        return False
    except Exception as e:
        print(json.dumps({"error": str(e)}, ensure_ascii=False))
        return False


def main():
    parser = argparse.ArgumentParser(
        description='合同发票字段专用提取器（基于RapidOCR，扫描件PDF自动识别）',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog='''
示例：
  python contract_invoice_extractor.py quick-ocr --file "发票.pdf"
  python contract_invoice_extractor.py quick-ocr --file "扫描件.pdf" --show-text
  python contract_invoice_extractor.py extract --file "合同.pdf" --output-dir "结果/"
  python contract_invoice_extractor.py extract --file "扫描件.pdf" --force-ocr
  python contract_invoice_extractor.py batch --dir "发票目录/" --output-dir "结果/"
  python contract_invoice_extractor.py detect --file "未知文件.pdf"
        ''',
    )
    subparsers = parser.add_subparsers(dest="command", help="子命令")

    # quick-ocr 子命令（推荐：一键OCR提取）
    qocr_parser = subparsers.add_parser("quick-ocr", help="快速OCR提取关键字段（推荐）")
    qocr_parser.add_argument("--file", required=True, help="PDF/图片文件路径")
    qocr_parser.add_argument("--force-ocr", action="store_true", help="强制OCR（忽略缓存）")
    qocr_parser.add_argument("--project", default="default", help="项目名（用于缓存隔离）")
    qocr_parser.add_argument("--show-text", action="store_true", help="同时输出OCR文本预览（前2000字符）")
    qocr_parser.add_argument("--show-all", action="store_true", help="输出完整OCR文本和所有字段")

    # extract 子命令
    ext_parser = subparsers.add_parser("extract", help="提取单个文件字段（JSON+Excel双输出）")
    ext_parser.add_argument("--file", required=True, help="PDF/图片文件路径")
    ext_parser.add_argument("--output-dir", default=".trae/ocr_extract", help="输出目录")
    ext_parser.add_argument("--project", default="default", help="项目名（用于OCR缓存隔离）")
    ext_parser.add_argument("--force-ocr", action="store_true", help="强制OCR（忽略文本层和缓存）")

    # batch 子命令
    bat_parser = subparsers.add_parser("batch", help="批量提取")
    bat_parser.add_argument("--dir", required=True, help="输入目录")
    bat_parser.add_argument("--output-dir", default=".trae/ocr_extract", help="输出目录")
    bat_parser.add_argument("--project", default="default", help="项目名")

    # detect 子命令
    det_parser = subparsers.add_parser("detect", help="检测文档类型（扫描件自动OCR）")
    det_parser.add_argument("--file", required=True, help="PDF文件路径")
    det_parser.add_argument("--project", default="default", help="项目名")

    args = parser.parse_args()

    if args.command == "quick-ocr":
        success = cmd_quick_ocr(args)
    elif args.command == "extract":
        success = cmd_extract(args)
    elif args.command == "batch":
        success = cmd_batch(args)
    elif args.command == "detect":
        success = cmd_detect(args)
    else:
        parser.print_help()
        success = False

    sys.exit(0 if success else 1)


if __name__ == "__main__":
    main()
