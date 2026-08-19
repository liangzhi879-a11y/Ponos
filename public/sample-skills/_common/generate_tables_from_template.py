"""核心表格生成工具（模板写死版 v1.15.0）

解决"agent每次输出的字段名/列顺序不一致"问题。
直接以项目根目录下 00 高新模板（全） 目录的5个模板文件为基础（路径通过 path_config 推断），
复制模板保留所有字段名/列顺序/合并单元格/样式，仅填充数据行。

用法：
  python generate_tables_from_template.py --enterprise "深圳市XX公司" --year 2026 \\
    --ip-data ip.json --rd-data rd.json --ps-data ps.json --ach-data ach.json \\
    --output-dir "00_核心表格"

数据JSON格式示例：
  ip.json: [{"编号":"IP01","名称":"...","类别":"发明专利（非国防专利）",...}]
  rd.json: [{"编号":"RD01","名称":"...","技术领域一级":"...",...}]
  ps.json: [{"编号":"PS01","名称":"...",...}]
  ach.json: [{"序号":1,"名称":"...",...}]

依赖：openpyxl
"""
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from copy import copy
from datetime import datetime, date
import os
import sys
import json
import argparse
import shutil
from pathlib import Path

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from path_config import get_template_dir_str

# ============================================================
# 模板路径（基于 path_config 推断，不允许agent修改）
# ============================================================
TEMPLATE_DIR = get_template_dir_str()

TEMPLATES = {
    'ip': {
        'template_file': '前海云充科技-知识产权表（参与本次创新能力知识产权评价，汇总信息只统计此列表中的知识产权）.xlsx',
        'output_prefix': '-知识产权表.xlsx',
        'data_start_row': 2,  # 模板第1行表头，第2行起为数据
    },
    'rd': {
        'template_file': '前海云充科技-企业研究开发活动汇总表（近三年执行的活动）.xlsx',
        'output_prefix': '-研发活动汇总表.xlsx',
        'data_start_row': 2,
    },
    'ps': {
        'template_file': '前海云充科技-高新技术产品（服务）明细表.xlsx',
        'output_prefix': '-高新产品明细表.xlsx',
        'data_start_row': 2,
    },
    'ach': {
        'template_file': '前海云充科技-科技成果转化情况汇总表.xlsx',
        'output_prefix': '-科技成果转化情况汇总表.xlsx',
        'data_start_row': 2,
    },
    'rdps': {
        'template_file': '前海云充21-23年研发项目RD21个名称、高新产品PS？个.xls',
        'output_prefix': '-研发项目RD、高新产品PS-汇总表.xlsx',
        'data_start_row': 3,  # 第1行标题，第2行表头，第3行起数据
    },
    'toai': {
        'template_file': 'TO-AI(仅作示例，不关联模板公司).xlsx',
        'output_prefix': '-TO-AI.xlsx',
        'data_start_row': 2,
    },
    # v1.16.0新增：科技人员清单表格（三方对比，独立规格，不基于固定模板复制）
    # 参考：定稿-云充2023年科技人员信息表.xlsx（13列基础字段）+ 三方对比扩展字段
    # 数据源：花名册（基础）+ 上年度社保缴费记录 + 台账纳入计算人员
    'staff': {
        'template_file': None,  # 无固定模板，代码生成
        'output_prefix': '-科技人员清单（三方对比）.xlsx',
        'data_start_row': 2,
        'create_new': True,  # 标记为代码生成（非模板复制）
    },
}


# ============================================================
# 数据字段到模板列的映射（写死，确保数据写入正确的列）
# ============================================================
FIELD_MAPPING = {
    'ip': {
        # JSON数据字段名 → 模板列号(1-indexed)
        '编号': 1, '知识产权编号': 1,
        '名称': 2, '知识产权名称': 2,
        '类别': 3,
        '获得方式': 4,
        '专利号': 5, '专利号/著作权号': 5, '著作权号': 5,
        '授权日期': 6,
        '所属单位': 7, '知识产权所属单位或个人': 7,
        '摘要': 8, '国家知识产权局官方网站上公布的摘要': 8,
        '先进性说明': 9, '核心关键技术先进性说明': 9,
        '支持作用说明': 10, '核心技术的支持作用说明': 10,
    },
    'rd': {
        '编号': 1, '研发活动编号': 1,
        '名称': 2, '研发活动名称': 2,
        '技术领域一级': 3, '技术领域（一级）': 3,
        '技术领域二级': 4, '技术领域（二级）': 4,
        '技术领域三级': 5, '技术领域（三级）': 5,
        '开始时间': 6,
        '结束时间': 7,
        '技术来源': 8,
        '知识产权编号': 9,
        '研发经费总预算': 10, '研发经费总预算（万元）': 10,
        '研发经费近三年总支出': 11, '研发经费近三年总支出（万元）': 11,
        '第一年支出': 12, '其中第一年支出': 12, '其中：第一年（2021年）支出（万元）': 12,
        '第二年支出': 13, '其中第二年支出': 13, '其中：第二年（2022年）支出（万元）': 13,
        '第三年支出': 14, '其中第三年支出': 14, '其中：第三年（2023年）支出（万元）': 14,
        '研发活动人员数': 15, '研发人员数': 15,
    },
    'ps': {
        '编号': 1, '产品（服务）编号': 1,
        '名称': 2, '产品（服务）名称': 2,
        '技术领域一级': 3, '技术领域（一级）': 3,
        '技术领域二级': 4, '技术领域（二级）': 4,
        '技术领域三级': 5, '技术领域（三级）': 5,
        '技术来源': 6,
        '上年度销售收入': 7, '上年度销售收入 （万元）': 7,
        '是否主要产品': 8, '是否主要产品 （服务）': 8,
        '知识产权编号': 9,
        '关键技术及主要技术指标': 10, '关键技术及主要技术指标（限400字）': 10,
        '竞争优势': 11, '与同类产品（服务）的竞争优势（限400字）': 11,
        '知识产权支持作用': 12, '知识产权获得情况及其对产品（服务）在技术上发挥的支持作用（限400字）': 12,
    },
    'ach': {
        '序号': 1, '科技成果序号': 1,
        '名称': 2, '科技成果名称': 2,
        '成果类型': 3,
        '成果来源': 4,
        '转化结果': 5,
        '转化时间': 6,
        '关联IP': 7, '关联ip': 7,
        '关联RD': 8, '关联rd': 8,
        '关联PS': 9, '关联ps': 9,
        '转化形式': 10,
        '涉及关键技术': 11, '涉及关键技术（限400字）': 11,
        '成效': 12, '成效（限400字）': 12,
    },
    # v1.16.0新增：科技人员清单（三方对比）字段映射 - 16列
    # 参考：定稿-云充2023年科技人员信息表.xlsx（13列基础）+ 三方对比扩展字段
    'staff': {
        '序号': 1,
        '姓名': 2,
        '身份证号码': 3, '身份证号': 3,
        '性别': 4,
        '部门': 5,
        '岗位': 6, '职务': 6,
        '入职日期': 7, '入职时间': 7,
        '离职日期': 8, '离职时间': 8,
        '上年在职天数': 9, '在职天数': 9,
        '花名册': 10, '是否在花名册': 10,
        '社保清单': 11, '是否在社保清单': 11,
        '台账': 12, '是否在台账': 12,
        '是否符合科技人员条件': 13, '符合条件': 13,
        '不符合原因': 14, '原因': 14,
        '备注': 15, '说明': 15,
    },
}


# ============================================================
# 字数控制配置（写死，与提示词文件要求一致）
# ============================================================
# 每个表格的每个字段的字数范围 (min_chars, max_chars)
# None 表示不校验字数（如摘要严格按专利说明书原文，可短可长）
TEXT_LENGTH_LIMITS = {
    'ip': {
        8: None,  # 摘要：不校验字数（严格按专利说明书原文摘录）
        9: (300, 400),  # 先进性说明：300-400字
        10: (300, 400),  # 支持作用说明：300-400字
    },
    'rd': {
        # RD表主表只有15列基础字段，长文本字段在docx附件中（不在Excel中）
        # 这里不配置长文本字段
    },
    'ps': {
        10: (350, 400),  # 关键技术及主要技术指标
        11: (350, 400),  # 竞争优势
        12: (350, 400),  # 知识产权支持作用
    },
    'ach': {
        11: (370, 410),  # 涉及关键技术（允许超过400到410）
        12: (370, 410),  # 成效（允许超过400到410）
    },
    'staff': {
        # 科技人员清单（三方对比）无长文本字段，不校验字数
    },
}


DATE_COLS = {
    'rd': {6, 7},
    'ip': {6},
    'ach': {6},
    'staff': {7, 8},
}


def _parse_date(value):
    """解析日期字符串为 datetime 对象，支持多种常见格式"""
    if not value:
        return None
    if isinstance(value, (datetime, date)):
        return value
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d"):
        try:
            return datetime.strptime(str(value), fmt)
        except ValueError:
            continue
    return value


def validate_and_truncate_text(value, min_chars, max_chars):
    """校验字数（只告警不截断，由agent重新撰写）

    返回 (final_value, status, message)
    status: 'ok' / 'too_long' / 'too_short' / 'empty'
    - too_long: 字数超限，**不截断**，由agent重新撰写精简内容
    - too_short: 字数不足，**不补充**，由agent重新撰写扩充内容
    """
    if value is None or (isinstance(value, str) and value.strip() == ''):
        return ('', 'empty', '空值')

    text = str(value)
    length = len(text)

    if length > max_chars:
        # 超限：不截断，告警由agent重新撰写
        return (text, 'too_long', f'字数超限：{length} > {max_chars}，需agent重新精简到{max_chars}字以内')

    if length < min_chars:
        # 字数不足：不补充，告警由agent重新撰写
        return (text, 'too_short', f'字数不足：{length} < {min_chars}，需agent重新扩充到{min_chars}字以上')

    return (text, 'ok', f'字数合规：{length}')


def check_text_length(table_type, data_list):
    """检查数据列表中所有长文本字段的字数，返回报告（不修改数据）

    返回 {'passed': bool, 'too_long_count': int, 'too_short_count': int,
         'ok_count': int, 'details': [...]}
    """
    limits = TEXT_LENGTH_LIMITS.get(table_type, {})
    if not limits:
        return {'passed': True, 'too_long_count': 0, 'too_short_count': 0,
                'ok_count': 0, 'details': [], 'message': '该表格无字数校验配置'}

    field_map = FIELD_MAPPING.get(table_type, {})
    # 反向映射：列号 → 字段名候选
    col_to_fields = {}
    for fname, col in field_map.items():
        col_to_fields.setdefault(col, []).append(fname)

    details = []
    too_long_count = 0
    too_short_count = 0
    ok_count = 0
    empty_count = 0

    for row_idx, row_data in enumerate(data_list):
        for col, limit in limits.items():
            if limit is None:
                continue  # None 表示不校验字数（如IP摘要严格按专利说明书原文）
            min_chars, max_chars = limit
            # 找到该列对应的字段名
            field_names = col_to_fields.get(col, [])
            value = None
            matched_field = None
            for fname in field_names:
                if fname in row_data:
                    value = row_data[fname]
                    matched_field = fname
                    break

            if value is None:
                # 字段不存在，尝试模糊匹配
                for fname, c in field_map.items():
                    if c == col:
                        continue
                    if any(k in fname for k in field_names):
                        if fname in row_data:
                            value = row_data[fname]
                            matched_field = fname
                            break

            final_value, status, message = validate_and_truncate_text(
                value, min_chars, max_chars)

            if status == 'too_long':
                too_long_count += 1
            elif status == 'too_short':
                too_short_count += 1
            elif status == 'ok':
                ok_count += 1
            elif status == 'empty':
                empty_count += 1

            details.append({
                'row': row_idx + 1,
                'col': col,
                'field': matched_field or f'col{col}',
                'original_length': len(str(value)) if value else 0,
                'status': status,
                'message': message,
                'preview': str(value)[:50] + '...' if value and len(str(value)) > 50 else str(value),
            })

    passed = too_long_count == 0 and too_short_count == 0 and empty_count == 0
    return {
        'passed': passed,
        'too_long_count': too_long_count,
        'too_short_count': too_short_count,
        'ok_count': ok_count,
        'empty_count': empty_count,
        'details': details,
    }


def apply_text_length_limits(table_type, data_list):
    """对数据列表进行字数校验（只校验不截断，不修改数据）

    **重要**：本函数不会修改 data_list 中的任何字段值。
    字数超限或不足时只生成告警报告，由agent根据告警重新撰写内容。

    返回字数校验报告
    """
    limits = TEXT_LENGTH_LIMITS.get(table_type, {})
    if not limits:
        return {'passed': True, 'too_long_count': 0, 'too_short_count': 0,
                'ok_count': 0, 'details': [], 'message': '该表格无字数校验配置'}

    field_map = FIELD_MAPPING.get(table_type, {})
    col_to_fields = {}
    for fname, col in field_map.items():
        col_to_fields.setdefault(col, []).append(fname)

    details = []
    too_long_count = 0
    too_short_count = 0
    ok_count = 0
    empty_count = 0

    for row_idx, row_data in enumerate(data_list):
        for col, limit in limits.items():
            if limit is None:
                continue  # None 表示不校验字数（如IP摘要严格按专利说明书原文）
            min_chars, max_chars = limit
            field_names = col_to_fields.get(col, [])
            value = None
            matched_field = None
            for fname in field_names:
                if fname in row_data:
                    value = row_data[fname]
                    matched_field = fname
                    break

            final_value, status, message = validate_and_truncate_text(
                value, min_chars, max_chars)

            # 只统计，不修改数据
            if status == 'too_long':
                too_long_count += 1
            elif status == 'too_short':
                too_short_count += 1
            elif status == 'ok':
                ok_count += 1
            elif status == 'empty':
                empty_count += 1

            details.append({
                'row': row_idx + 1,
                'col': col,
                'field': matched_field or f'col{col}',
                'original_length': len(str(value)) if value else 0,
                'status': status,
                'message': message,
                'preview': str(value)[:50] + '...' if value and len(str(value)) > 50 else str(value),
            })

    passed = too_long_count == 0 and too_short_count == 0 and empty_count == 0
    return {
        'passed': passed,
        'too_long_count': too_long_count,
        'too_short_count': too_short_count,
        'ok_count': ok_count,
        'empty_count': empty_count,
        'details': details,
    }


def copy_template_style(src_cell, dst_cell):
    """复制单元格样式"""
    if src_cell.has_style:
        dst_cell.font = copy(src_cell.font)
        dst_cell.alignment = copy(src_cell.alignment)
        dst_cell.border = copy(src_cell.border)
        dst_cell.fill = copy(src_cell.fill)
        dst_cell.number_format = src_cell.number_format


# ============================================================
# v1.16.0新增：科技人员清单（三方对比）表头与样式定义
# ============================================================
STAFF_TABLE_HEADERS = [
    '序号', '姓名', '身份证号码', '性别', '部门', '岗位',
    '入职日期', '离职日期', '上年在职天数',
    '花名册', '社保清单', '台账',
    '是否符合科技人员条件', '不符合原因', '备注',
]

STAFF_SUMMARY_ROWS = [
    # 末尾统计行（按用户要求：根据社保计算科技人员占比）
    '花名册总人数', '社保清单总人数', '台账纳入总人数',
    '符合科技人员条件人数', '科技人员占比（基于社保）',
]


def _generate_table_create_new(table_type, data_list, enterprise_name, output_path, spec):
    """v1.16.0：代码生成新表格（不基于固定模板复制）

    用于无固定模板的新增表格类型（如科技人员三方对比清单）。
    按 FIELD_MAPPING 中定义的字段顺序生成表头，应用统一样式，写入数据。

    流程：
    1. 创建新 Workbook
    2. 写入表头（应用表头样式：宋体12号bold、左对齐、thin边框、浅色背景）
    3. 字数校验（如配置了 TEXT_LENGTH_LIMITS）
    4. 按字段映射写入数据行（应用数据样式：宋体11号、左对齐、thin边框）
    5. 追加统计行（仅 staff 表）
    6. 保存
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = table_type

    # 表头样式
    header_font = Font(name='宋体', size=12, bold=True)
    header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )

    # 数据样式
    data_font = Font(name='宋体', size=11)
    data_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

    # 1. 写入表头
    headers = STAFF_TABLE_HEADERS if table_type == 'staff' else []
    if not headers:
        # 通用回退：从 FIELD_MAPPING 提取列号排序的字段名
        fm = FIELD_MAPPING.get(table_type, {})
        cols = sorted(set(fm.values()))
        headers = [next((k for k, v in fm.items() if v == c), f'col{c}') for c in cols]

    for c, h in enumerate(headers, start=1):
        cell = ws.cell(1, c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # 2. 字数校验（科技人员清单无长文本字段，跳过）
    report = apply_text_length_limits(table_type, data_list)
    if not report.get('passed', True):
        # 配置了字数限制但不合规 → 停止写入（与模板复制分支保持一致）
        print(f'[字数校验警告] {table_type} 表（以下字段需agent重新撰写）:')
        print(f'  合规: {report["ok_count"]}  超限需精简: {report["too_long_count"]}  '
              f'字数不足需扩充: {report["too_short_count"]}  空值需补充: {report["empty_count"]}')
        for d in report['details']:
            if d['status'] != 'ok':
                action = '精简' if d['status'] == 'too_long' else ('扩充' if d['status'] == 'too_short' else '补充')
                print(f'    [行{d["row"]} 列{d["col"]} {d["field"]}] {d["status"]}: '
                      f'现{d["original_length"]}字 → {d["message"]} → 需agent{action}后重新生成')
        print(f'  ⚠ 警告：以上字段字数不符合要求，**已停止写入Excel**，agent需根据告警重新撰写后再次运行脚本')
        if os.path.exists(output_path):
            os.remove(output_path)
        return {'output_path': None, 'text_report': report, 'skipped': True}

    # 3. 写入数据行
    field_map = FIELD_MAPPING.get(table_type, {})
    date_cols = DATE_COLS.get(table_type, set())
    for idx, row_data in enumerate(data_list):
        r = 2 + idx
        for field_name, value in row_data.items():
            col = field_map.get(field_name)
            if col is None:
                for k, v in field_map.items():
                    if field_name in k or k in field_name:
                        col = v
                        break
            if col is None:
                continue

            if col in date_cols and value:
                parsed = _parse_date(value)
                if isinstance(parsed, datetime):
                    value = parsed

            cell = ws.cell(r, col, value=value)
            cell.font = data_font
            cell.alignment = data_align
            cell.border = thin_border
            if col in date_cols and isinstance(value, datetime):
                cell.number_format = "YYYY/MM/DD"

    # 4. 科技人员清单追加统计行（v1.16.0核心需求）
    if table_type == 'staff' and data_list:
        summary_start_row = 2 + len(data_list) + 1  # 数据行后空1行
        # 统计：花名册总人数、社保清单总人数、台账总人数、符合条件人数、占比
        roster_count = sum(1 for r in data_list if str(r.get('花名册', r.get('是否在花名册', ''))) in ('是', '√', 'Y', 'y'))
        ss_count = sum(1 for r in data_list if str(r.get('社保清单', r.get('是否在社保清单', ''))) in ('是', '√', 'Y', 'y'))
        ledger_count = sum(1 for r in data_list if str(r.get('台账', r.get('是否在台账', ''))) in ('是', '√', 'Y', 'y'))
        qualified_count = sum(1 for r in data_list if str(r.get('是否符合科技人员条件', r.get('符合条件', ''))) in ('是', '√', 'Y', 'y'))
        ratio = f'{(qualified_count / ss_count * 100):.2f}%' if ss_count > 0 else '社保清单为空，无法计算'

        summary_data = [
            ('花名册总人数', roster_count),
            ('社保清单总人数', ss_count),
            ('台账纳入总人数', ledger_count),
            ('符合科技人员条件人数', qualified_count),
            ('科技人员占比（基于社保）', ratio),
        ]
        summary_font = Font(name='宋体', size=11, bold=True)
        summary_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
        for i, (label, val) in enumerate(summary_data):
            r = summary_start_row + i
            cell_label = ws.cell(r, 1, value=label)
            cell_label.font = summary_font
            cell_label.fill = summary_fill
            cell_label.alignment = header_align
            cell_label.border = thin_border
            # 合并1-2列作为标签
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
            cell_val = ws.cell(r, 3, value=val)
            cell_val.font = summary_font
            cell_val.fill = summary_fill
            cell_val.alignment = header_align
            cell_val.border = thin_border

    # 5. 列宽设置（科技人员清单专用）
    if table_type == 'staff':
        col_widths = {1: 6, 2: 12, 3: 22, 4: 6, 5: 14, 6: 14, 7: 14, 8: 14, 9: 12,
                      10: 10, 11: 10, 12: 10, 13: 18, 14: 30, 15: 30}
        for c, w in col_widths.items():
            ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width = w
        # 表头行高
        ws.row_dimensions[1].height = 32

    # 6. 保存
    wb.save(output_path)
    print(f'[生成] {table_type}: {os.path.basename(output_path)} ({len(data_list)}行数据)')
    return {'output_path': output_path, 'text_report': report}


def generate_table_from_template(table_type, data_list, enterprise_name, output_dir):
    """从模板生成表格

    流程：
    1. 复制模板文件 → 输出文件（或代码生成新文件，当 create_new=True）
    2. 读取模板的表头行和样式
    3. 清空数据行（保留表头）
    4. 根据FIELD_MAPPING将data_list数据写入对应列
    5. 复制模板数据行样式到新数据行

    返回输出文件路径
    """
    spec = TEMPLATES[table_type]
    is_create_new = spec.get('create_new', False)

    # 输出文件名
    output_filename = f'{enterprise_name}{spec["output_prefix"]}'
    output_path = os.path.join(output_dir, output_filename)
    os.makedirs(output_dir, exist_ok=True)

    if is_create_new:
        # v1.16.0：代码生成新表格（无固定模板，如科技人员三方对比清单）
        return _generate_table_create_new(table_type, data_list, enterprise_name, output_path, spec)

    template_path = os.path.join(TEMPLATE_DIR, spec['template_file'])
    if not os.path.exists(template_path):
        raise FileNotFoundError(f'模板文件不存在: {template_path}')

    # 1. 复制模板文件
    shutil.copy2(template_path, output_path)

    # 2. 加载复制的文件
    wb = openpyxl.load_workbook(output_path)
    ws = wb.active

    # 3. 清空模板数据行（保留表头和样式作为参考）
    header_row = spec['data_start_row'] - 1 if spec['data_start_row'] > 1 else 1
    # 如果模板第1行是标题（如RDPS），保留第1-2行
    if table_type == 'rdps':
        header_row = 2  # 第2行是表头
        data_start = 3
    else:
        header_row = 1
        data_start = 2

    # 保存模板数据行的样式（第data_start行）
    template_row_styles = {}
    if ws.max_row >= data_start:
        for c in range(1, ws.max_column + 1):
            src_cell = ws.cell(data_start, c)
            template_row_styles[c] = {
                'font': copy(src_cell.font),
                'alignment': copy(src_cell.alignment),
                'border': copy(src_cell.border),
                'fill': copy(src_cell.fill),
                'number_format': src_cell.number_format,
            }

    # 清空所有数据行（从data_start开始）
    for r in range(data_start, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            ws.cell(r, c).value = None

    # 4. 写入数据前：字数校验（只告警不截断，由agent重新撰写）
    field_map = FIELD_MAPPING.get(table_type, {})
    report = apply_text_length_limits(table_type, data_list)

    # 打印字数校验报告
    if report['too_long_count'] > 0 or report['too_short_count'] > 0 or report['empty_count'] > 0:
        print(f'[字数校验警告] {table_type} 表（以下字段需agent重新撰写）:')
        print(f'  合规: {report["ok_count"]}  超限需精简: {report["too_long_count"]}  '
              f'字数不足需扩充: {report["too_short_count"]}  空值需补充: {report["empty_count"]}')
        for d in report['details']:
            if d['status'] != 'ok':
                action = '精简' if d['status'] == 'too_long' else ('扩充' if d['status'] == 'too_short' else '补充')
                print(f'    [行{d["row"]} 列{d["col"]} {d["field"]}] {d["status"]}: '
                      f'现{d["original_length"]}字 → {d["message"]} → 需agent{action}后重新生成')
                if d.get('preview'):
                    print(f'      预览: {d["preview"]}')
        print(f'  ⚠ 警告：以上字段字数不符合要求，**已停止写入Excel**，agent需根据告警重新撰写后再次运行脚本')
        # 删除已复制的模板文件（避免输出目录残留空Excel）
        if os.path.exists(output_path):
            os.remove(output_path)
        return {'output_path': None, 'text_report': report, 'skipped': True}

    print(f'[字数校验] {table_type} 表: 全部合规')

    # v1.18.2：IP表软著摘要预处理（软著行摘要为空时自动填"无"）
    if table_type == 'ip':
        for idx, data_row in enumerate(data_list):
            cat = data_row.get('类别') or data_row.get('category') or ''
            if '软件著作权' in str(cat):
                abstract_keys = ['摘要', '国家知识产权局官方网站上公布的摘要',
                                 '国家知识产权局官方网站上公布的摘要(限400字，软件著作权不用提供)']
                abstract_val = None
                matched_key = None
                for k in abstract_keys:
                    if k in data_row:
                        abstract_val = data_row[k]
                        matched_key = k
                        break
                if not abstract_val or (isinstance(abstract_val, str) and abstract_val.strip() == ''):
                    if matched_key:
                        data_row[matched_key] = '无'
                    else:
                        data_row['摘要'] = '无'
                    print(f'  [v1.18.2] 第{idx+1}行 软著摘要为空，已自动填"无"')
                elif str(abstract_val).strip() != '无':
                    print(f'  [v1.18.2警告] 第{idx+1}行 软著摘要应为"无"，实际为"{str(abstract_val)[:20]}"')

    # 5. 写入数据
    date_cols = DATE_COLS.get(table_type, set())
    for idx, data_row in enumerate(data_list):
        r = data_start + idx
        for field_name, value in data_row.items():
            col = field_map.get(field_name)
            if col is None:
                # 尝试模糊匹配
                for k, v in field_map.items():
                    if field_name in k or k in field_name:
                        col = v
                        break
            if col is None:
                continue

            if col in date_cols and value:
                parsed = _parse_date(value)
                if isinstance(parsed, datetime):
                    value = parsed

            cell = ws.cell(r, col)
            cell.value = value
            # 应用模板样式
            if col in template_row_styles:
                style = template_row_styles[col]
                cell.font = copy(style['font'])
                cell.alignment = copy(style['alignment'])
                cell.border = copy(style['border'])
                cell.fill = copy(style['fill'])
                cell.number_format = "YYYY/MM/DD" if (col in date_cols and isinstance(value, datetime)) else style['number_format']
            elif col in date_cols and isinstance(value, datetime):
                cell.number_format = "YYYY/MM/DD"

    # 6. 保存
    wb.save(output_path)
    print(f'[生成] {table_type}: {output_filename} ({len(data_list)}行数据)')
    return {'output_path': output_path, 'text_report': report}


def generate_all_tables(enterprise_name, output_dir, ip_data=None, rd_data=None,
                         ps_data=None, ach_data=None, rdps_data=None, staff_data=None):
    """生成所有核心表格

    参数：
      enterprise_name: 企业名称
      output_dir: 输出目录
      ip_data/rd_data/ps_data/ach_data/rdps_data: 数据列表（每项为dict）
      staff_data: v1.16.0新增 科技人员清单（三方对比）数据列表

    返回 {'ip': path, 'rd': path, ...}
    """
    results = {}
    if ip_data:
        results['ip'] = generate_table_from_template('ip', ip_data, enterprise_name, output_dir)
    if rd_data:
        results['rd'] = generate_table_from_template('rd', rd_data, enterprise_name, output_dir)
    if ps_data:
        results['ps'] = generate_table_from_template('ps', ps_data, enterprise_name, output_dir)
    if ach_data:
        results['ach'] = generate_table_from_template('ach', ach_data, enterprise_name, output_dir)
    if rdps_data:
        results['rdps'] = generate_table_from_template('rdps', rdps_data, enterprise_name, output_dir)
    if staff_data:
        results['staff'] = generate_table_from_template('staff', staff_data, enterprise_name, output_dir)
    return results


def main():
    parser = argparse.ArgumentParser(description='核心表格生成工具（模板写死版 v1.16.0）')
    parser.add_argument('--enterprise', required=True, help='企业名称')
    parser.add_argument('--output-dir', required=True, help='输出目录')
    parser.add_argument('--ip-data', help='IP数据JSON文件')
    parser.add_argument('--rd-data', help='RD数据JSON文件')
    parser.add_argument('--ps-data', help='PS数据JSON文件')
    parser.add_argument('--ach-data', help='成果转化数据JSON文件')
    parser.add_argument('--rdps-data', help='RD/PS汇总数据JSON文件')
    parser.add_argument('--staff-data', help='科技人员清单（三方对比）数据JSON文件（v1.16.0新增）')
    args = parser.parse_args()

    def load_json(path):
        if not path or not os.path.exists(path):
            return None
        with open(path, 'r', encoding='utf-8') as f:
            obj = json.load(f)
        # 兼容两种格式：
        #   1) 直接list: [{...}, {...}]
        #   2) 包装对象: {"table_type": "ip", "data": [...]} → 取data字段
        if isinstance(obj, dict) and 'data' in obj and isinstance(obj['data'], list):
            return obj['data']
        if isinstance(obj, list):
            return obj
        # 其他格式：返回None避免误用
        print(f'⚠ 警告: {path} 格式不符（应为list或含data字段的对象），已跳过')
        return None

    results = generate_all_tables(
        args.enterprise, args.output_dir,
        ip_data=load_json(args.ip_data),
        rd_data=load_json(args.rd_data),
        ps_data=load_json(args.ps_data),
        ach_data=load_json(args.ach_data),
        rdps_data=load_json(args.rdps_data),
        staff_data=load_json(args.staff_data),
    )

    print(f'\n生成 {len(results)} 个表格:')
    for ttype, result in results.items():
        if isinstance(result, dict):
            if result.get('skipped'):
                report = result.get('text_report', {})
                print(f'  {ttype}: ⚠已跳过（字数不合规，需agent重新撰写）'
                      f'  [超限{report.get("too_long_count",0)} 不足{report.get("too_short_count",0)} 空{report.get("empty_count",0)}]')
            else:
                output_path = result['output_path']
                report = result.get('text_report', {})
                status = '✓全部合规' if report.get('passed') else '⚠不合规'
                print(f'  {ttype}: {output_path}  [{status}]')
        else:
            print(f'  {ttype}: {result}')

    # 如果有字数不合规项（skipped），退出码非0
    has_issues = False
    for result in results.values():
        if isinstance(result, dict):
            if result.get('skipped'):
                has_issues = True
                break
            report = result.get('text_report', {})
            if not report.get('passed'):
                has_issues = True
                break
    if has_issues:
        print('\n⚠ 存在字数不合规项（已停止写入Excel），agent需根据上述告警重新撰写内容后再次运行脚本')
        sys.exit(1)


if __name__ == '__main__':
    main()
