# -*- coding: utf-8 -*-
"""
template_injector.py v2.0 - 官方模板注入引擎
将数据填入官方高新认定模板xlsx，而非从零生成。
复制模板→注入数据→保存输出。

模板查找优先级：项目目录模板 > 内置模板（_common/templates/）
内置模板随技能包分发，无需每个项目手动放置。
"""
import os
import re
import shutil
from pathlib import Path
from copy import copy
from datetime import datetime, date
import openpyxl
from openpyxl.utils import get_column_letter

RD_COLS: dict[int, str] = {
    1: "编号", 2: "名称", 3: "领域一级", 4: "领域二级", 5: "领域三级",
    6: "开始时间", 7: "结束时间", 8: "技术来源", 9: "IP编号", 10: "预算",
    11: "总支出", 12: "第一年", 13: "第二年", 14: "第三年", 15: "人员数",
    16: "目的", 17: "核心技术", 18: "阶段成果",
}

PS_COLS: dict[int, str] = {
    1: "编号", 2: "名称", 3: "领域一级", 4: "领域二级", 5: "领域三级",
    6: "技术来源", 7: "收入", 8: "是否主要", 9: "IP编号", 10: "关键技术",
    11: "竞争优势", 12: "支持作用",
}

IP_COLS: dict[int, str] = {
    1: "编号", 2: "名称", 3: "类别", 4: "获得方式", 5: "专利号",
    6: "授权日期", 7: "所属单位", 8: "摘要", 9: "先进性说明", 10: "支持作用说明",
}

ACH_COLS: dict[int, str] = {
    1: "序号", 2: "名称", 3: "成果类型", 4: "成果来源", 5: "转化结果",
    6: "转化时间", 7: "关联IP", 8: "关联RD", 9: "关联PS", 10: "转化形式",
    11: "涉及关键技术", 12: "成效",
}

RD_DATE_COLS: tuple[int, ...] = (6, 7)
IP_DATE_COLS: tuple[int, ...] = (6,)
ACH_DATE_COLS: tuple[int, ...] = (6,)

ACH_DROPDOWNS: dict[int, list[str]] = {
    3: ["专利", "版权", "集成电路布图设计", "其他"],
    4: ["自主研发", "受让、受赠、并购", "集成电路布图设计", "其他"],
    5: ["新产品", "新服务", "新设备", "新技术应用", "样品/样机", "其他"],
    10: [
        "许可他人使用该科技成果",
        "以该科技成果作为投资，折算股份或出资比例",
        "自行投资实施转化",
        "向他人转让该科技成果",
        "以该科技成果作为合作条件，与他人共同实施转化",
        "其他协商确定方式",
    ],
}

PS_DROPDOWNS: dict[int, list[str]] = {
    6: ["企业自有技术", "科研院所", "大专院校", "引进技术本企业消化创新", "国外技术", "其它企业技术"],
}

RD_DROPDOWNS: dict[int, list[str]] = {
    8: [
        "大专院校", "地方属科研院所", "其它企业技术", "引进技术本企业消化创新",
        "国外技术", "企业自有技术", "中央属科研院所",
    ],
}

IP_DROPDOWNS: dict[int, list[str]] = {
    3: [
        "实用新型专利", "外观设计专利", "软件著作权",
        "发明专利(非国防专利)", "发明专利(国防专利)",
        "植物新品种", "国家级农作物品种", "国家新药",
        "国家一级中药保护品种", "集成电路布图设计专有权",
    ],
    4: ["自主研发", "受让", "受赠", "并购", "其他"],
}

CHAR_LIMIT: int = 400


def _check_char_limit(value, field_name: str) -> None:
    if value is None:
        return
    text = str(value)
    length = len(text)
    if length > CHAR_LIMIT:
        print(f"  ⚠ WARNING: {field_name} 字数={length} 超标(限{CHAR_LIMIT}字)，不截断仅警告")


def _validate_dropdown(value, field_name: str, valid_values: list[str]) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    if text not in valid_values:
        print(f"  ⚠ WARNING: {field_name} 值='{text}' 不在有效选项 {valid_values} 中，已保留原值（请人工确认）")
        return None
    return text


class TemplateInjector:
    """模板注入引擎：复制官方模板xlsx → 填入数据 → 保存

    模板查找优先级：项目目录模板 > 内置模板（_common/templates/）
    内置模板随技能包分发，无需每个项目手动放置。
    """

    def __init__(self, template_dir: str, output_dir: str, enterprise_name: str):
        self.template_dir = Path(template_dir)
        self.output_dir = Path(output_dir)
        self.enterprise_name = enterprise_name
        self.output_dir.mkdir(parents=True, exist_ok=True)

    def _get_builtin_template_dir(self) -> Path:
        injector_dir = Path(__file__).parent
        return injector_dir / "templates"

    def _resolve_template(self, table_key: str, template_filename: str | None) -> tuple[str, str]:
        """查找模板文件，返回 (路径, 来源标识)

        来源标识: '项目目录' | '内置模板'
        """
        if template_filename:
            path = self.template_dir / template_filename
            if path.exists():
                return str(path), "项目目录"

        candidates = list(self.template_dir.glob(f"*{table_key}*.xlsx"))
        if candidates:
            return str(candidates[0]), "项目目录"

        builtin_dir = self._get_builtin_template_dir()
        if builtin_dir.exists():
            builtin_candidates = list(builtin_dir.glob(f"*{table_key}*.xlsx"))
            if builtin_candidates:
                path = str(builtin_candidates[0])
                print(f"  [模板] 使用内置模板: {builtin_candidates[0].name}")
                return path, "内置模板"

        raise FileNotFoundError(
            f"未找到模板文件，已搜索:\n"
            f"  项目目录: {self.template_dir}\n"
            f"  内置模板: {builtin_dir}\n"
            f"  关键词: {table_key}"
        )

    @staticmethod
    def _parse_date(value):
        """解析日期字符串为 datetime 对象，支持多种常见格式"""
        if not value:
            return None
        if isinstance(value, (datetime, date)):
            return value
        for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d"):
            try:
                return datetime.strptime(str(value), fmt)
            except ValueError:
                continue
        return value

    def _inject_rows(
        self,
        ws,
        data: list[dict],
        column_mapping: dict[int, str],
        start_row: int = 2,
        date_cols: tuple[int, ...] = (),
        char_limit_cols: dict[int, str] | None = None,
        dropdown_cols: dict[int, list[str]] | None = None,
    ) -> None:
        """通用注入方法"""
        header_row = start_row - 1
        style_row = start_row
        if ws.max_row < style_row:
            return

        template_styles: dict[int, dict] = {}
        for c in range(1, ws.max_column + 1):
            src = ws.cell(style_row, c)
            template_styles[c] = {
                "font": copy(src.font),
                "alignment": copy(src.alignment),
                "border": copy(src.border),
                "fill": copy(src.fill),
                "number_format": src.number_format,
            }

        existing_data_rows = ws.max_row - header_row
        needed_rows = len(data)
        if needed_rows > existing_data_rows:
            for _ in range(needed_rows - existing_data_rows):
                ws.insert_rows(start_row)
        elif needed_rows < existing_data_rows:
            ws.delete_rows(start_row + needed_rows, existing_data_rows - needed_rows)

        for idx, row_data in enumerate(data):
            r = start_row + idx
            for col_num, field_name in column_mapping.items():
                if field_name not in row_data:
                    continue
                value = row_data[field_name]

                if dropdown_cols and col_num in dropdown_cols:
                    validated = _validate_dropdown(value, field_name, dropdown_cols[col_num])
                    if validated is not None:
                        value = validated

                if col_num in date_cols and value:
                    parsed = self._parse_date(value)
                    if isinstance(parsed, datetime):
                        value = parsed

                cell = ws.cell(r, col_num, value=value)

                style = template_styles.get(col_num)
                if style:
                    cell.font = copy(style["font"])
                    cell.alignment = copy(style["alignment"])
                    cell.border = copy(style["border"])
                    cell.fill = copy(style["fill"])
                    cell.number_format = "YYYY/MM/DD" if (col_num in date_cols and isinstance(value, datetime)) else style["number_format"]
                elif col_num in date_cols and isinstance(value, datetime):
                    cell.number_format = "YYYY/MM/DD"

                if char_limit_cols and col_num in char_limit_cols:
                    _check_char_limit(value, char_limit_cols[col_num])

    def inject_rd_table(self, rd_data: list[dict], template_filename: str | None = None) -> str:
        """注入企业研究开发活动汇总表"""
        template_path, _ = self._resolve_template("研究开发活动汇总", template_filename)
        output_name = f"{self.enterprise_name}-企业研究开发活动汇总表（近三年执行的活动）.xlsx"
        output_path = str(self.output_dir / output_name)
        shutil.copy2(template_path, output_path)

        wb = openpyxl.load_workbook(output_path)
        ws = wb.active
        char_limit_cols = {16: "目的", 17: "核心技术", 18: "阶段成果"}
        self._inject_rows(ws, rd_data, RD_COLS, start_row=2, date_cols=RD_DATE_COLS, char_limit_cols=char_limit_cols, dropdown_cols=RD_DROPDOWNS)
        wb.save(output_path)
        wb.close()
        print(f"[注入] RD表: {output_name} ({len(rd_data)}行)")
        return output_path

    def inject_ps_table(self, ps_data: list[dict], template_filename: str | None = None) -> str:
        """注入高新技术产品（服务）明细表"""
        template_path, _ = self._resolve_template("高新技术产品", template_filename)
        output_name = f"{self.enterprise_name}-高新技术产品（服务）明细表.xlsx"
        output_path = str(self.output_dir / output_name)
        shutil.copy2(template_path, output_path)

        wb = openpyxl.load_workbook(output_path)
        ws = wb.active
        char_limit_cols = {10: "关键技术", 11: "竞争优势", 12: "支持作用"}
        self._inject_rows(ws, ps_data, PS_COLS, start_row=2, date_cols=(), char_limit_cols=char_limit_cols, dropdown_cols=PS_DROPDOWNS)
        wb.save(output_path)
        wb.close()
        print(f"[注入] PS表: {output_name} ({len(ps_data)}行)")
        return output_path

    def inject_ip_table(self, ip_data: list[dict], template_filename: str | None = None) -> str:
        """注入知识产权表"""
        template_path, _ = self._resolve_template("知识产权表", template_filename)
        output_name = f"{self.enterprise_name}-知识产权表（参与本次创新能力知识产权评价，汇总信息只统计此列表中的知识产权）.xlsx"
        output_path = str(self.output_dir / output_name)
        shutil.copy2(template_path, output_path)

        wb = openpyxl.load_workbook(output_path)
        ws = wb.active
        char_limit_cols = {8: "摘要", 9: "先进性说明", 10: "支持作用说明"}
        self._inject_rows(ws, ip_data, IP_COLS, start_row=2, date_cols=IP_DATE_COLS, char_limit_cols=char_limit_cols, dropdown_cols=IP_DROPDOWNS)
        wb.save(output_path)
        wb.close()
        print(f"[注入] IP表: {output_name} ({len(ip_data)}行)")
        return output_path

    def inject_achievement_table(self, ach_data: list[dict], template_filename: str | None = None) -> str:
        """注入科技成果转化情况汇总表"""
        template_path, _ = self._resolve_template("科技成果转化", template_filename)
        output_name = f"{self.enterprise_name}-科技成果转化情况汇总表.xlsx"
        output_path = str(self.output_dir / output_name)
        shutil.copy2(template_path, output_path)

        wb = openpyxl.load_workbook(output_path)
        ws = wb.active
        char_limit_cols = {11: "涉及关键技术", 12: "成效"}
        self._inject_rows(ws, ach_data, ACH_COLS, start_row=2, date_cols=ACH_DATE_COLS, char_limit_cols=char_limit_cols, dropdown_cols=ACH_DROPDOWNS)
        wb.save(output_path)
        wb.close()
        print(f"[注入] 成果表: {output_name} ({len(ach_data)}行)")
        return output_path


def _create_minimal_template(headers: list[str], num_data_rows: int = 3) -> str:
    """创建最小模板xlsx（含表头行 + num_data_rows行空数据），返回文件路径"""
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    header_font = Font(name="宋体", size=10, bold=True)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    header_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(1, c, value=h)
        cell.font = header_font
        cell.alignment = header_align
        cell.fill = header_fill
        cell.border = thin_border

    data_font = Font(name="宋体", size=10)
    data_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    for r in range(2, 2 + num_data_rows):
        for c in range(1, len(headers) + 1):
            cell = ws.cell(r, c)
            cell.font = data_font
            cell.alignment = data_align
            cell.border = thin_border

    col_widths = {1: 10, 2: 20, 3: 14, 4: 14, 5: 14, 6: 14, 7: 14, 8: 14,
                  9: 14, 10: 12, 11: 12, 12: 12, 13: 12, 14: 12, 15: 10,
                  16: 40, 17: 40, 18: 40}
    for c, w in col_widths.items():
        if c <= len(headers):
            ws.column_dimensions[get_column_letter(c)].width = w
    ws.row_dimensions[1].height = 28

    import tempfile
    fd, path = tempfile.mkstemp(suffix=".xlsx")
    os.close(fd)
    wb.save(path)
    wb.close()
    return path


if __name__ == "__main__":
    ENTERPRISE = "宏日嘉"
    OUTPUT_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "_test_injector_output")
    TEMPLATE_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "_test_injector_templates")

    os.makedirs(TEMPLATE_DIR, exist_ok=True)

    RD_HEADERS = [
        "研发活动编号", "研发活动名称", "技术领域（一级）", "技术领域（二级）", "技术领域（三级）",
        "开始时间", "结束时间", "技术来源", "知识产权编号",
        "研发经费总预算（万元）", "研发经费近三年总支出（万元）",
        "第一年（2023年）支出", "第二年（2024年）支出", "第三年（2025年）支出",
        "研发活动人员数", "目的及组织实施方式（限400字）",
        "核心技术及创新点（限400字）", "取得的阶段性成果（限400字）",
    ]
    PS_HEADERS = [
        "产品（服务）编号", "产品（服务）名称", "技术领域（一级）", "技术领域（二级）", "技术领域（三级）",
        "技术来源", "上年度销售收入（万元）", "是否主要产品（服务）", "知识产权编号",
        "关键技术及主要技术指标（限400字）", "与同类产品的竞争优势（限400字）",
        "知识产权获得情况及其支持作用（限400字）",
    ]
    IP_HEADERS = [
        "知识产权编号", "知识产权名称", "类别", "获得方式", "专利号/著作权号",
        "授权日期", "知识产权所属单位或个人", "摘要(限400字)",
        "核心关键技术先进性说明(限400字)", "对产品核心技术的支持作用说明(限400字)",
    ]
    ACH_HEADERS = [
        "科技成果序号", "科技成果名称", "成果类型", "成果来源", "转化结果",
        "转化时间", "关联IP", "关联RD", "关联PS", "转化形式",
        "涉及关键技术（限400字）", "成效（限400字）",
    ]

    shutil.copy(_create_minimal_template(RD_HEADERS), os.path.join(TEMPLATE_DIR, "企业研究开发活动汇总表.xlsx"))
    shutil.copy(_create_minimal_template(PS_HEADERS), os.path.join(TEMPLATE_DIR, "高新技术产品明细表.xlsx"))
    shutil.copy(_create_minimal_template(IP_HEADERS), os.path.join(TEMPLATE_DIR, "知识产权表.xlsx"))
    shutil.copy(_create_minimal_template(ACH_HEADERS), os.path.join(TEMPLATE_DIR, "科技成果转化情况汇总表.xlsx"))

    rd_test = [
        {
            "编号": "RD01", "名称": "高效节能充电模块研发",
            "领域一级": "新能源与节能", "领域二级": "高效节能技术", "领域三级": "工业节能技术",
            "开始时间": "2023-06-01", "结束时间": "2025-05-31",
            "技术来源": "企业自有技术", "IP编号": "IP01,IP02,IP03",
            "预算": 180.0, "总支出": 172.5, "第一年": 60.0, "第二年": 55.0, "第三年": 57.5,
            "人员数": 6,
            "目的": "本项目旨在研发一种高效节能充电模块，通过优化功率拓扑结构和控制算法，实现充电效率提升至95%以上，降低待机功耗至0.5W以下，满足新能源汽车充电桩及工业充电设备对高能效、高可靠性的核心组件需求。项目采用LLC谐振变换拓扑结合DSP数字控制方案，开发宽输入电压范围（180-480VAC）的高频开关电源模块。",
            "核心技术": "（1）LLC谐振变换器参数优化设计技术：建立谐振槽路参数多目标优化模型，通过频域分析法确定谐振电感、励磁电感与谐振电容的最优配比，实现全负载范围内零电压开关。（2）自适应死区时间控制技术：基于实时检测的谐振电流过零点，动态调节上下管死区时间，最小化体二极管导通损耗。（3）数字均流并联技术：采用CAN总线通信的数字化均流策略，实现多模块并联时电流不均衡度<3%。",
            "阶段成果": "已完成LLC谐振变换器原理样机研制并通过测试，满载效率达95.8%，待机功耗0.42W，获授权发明专利2项。正在开展EMC测试和可靠性验证，预计2025年Q2完成小批量试产。",
        },
        {
            "编号": "RD02", "名称": "智能电池管理系统开发",
            "领域一级": "新能源与节能", "领域二级": "新型动力电池", "领域三级": "电池管理系统技术",
            "开始时间": "2023-03-01", "结束时间": "2025-02-28",
            "技术来源": "企业自有技术", "IP编号": "IP04,IP05",
            "预算": 150.0, "总支出": 145.0, "第一年": 50.0, "第二年": 48.0, "第三年": 47.0,
            "人员数": 5,
            "目的": "本项目旨在开发面向锂离子电池组的智能电池管理系统（BMS），实现单体电池电压/温度/电流的实时高精度采集与均衡管理，SOC估算精度达到±3%，支持电池组健康状态（SOH）在线诊断。",
            "核心技术": "（1）基于扩展卡尔曼滤波的SOC估算算法：建立锂离子电池二阶RC等效电路模型，通过递推最小二乘法在线辨识模型参数，再结合EKF实现动态工况下的SOC精确估算。（2）主动均衡与被动均衡混合策略：根据电池组不一致度自动切换均衡模式，单体压差较大时采用变压器隔离型主动均衡。",
            "阶段成果": "已完成BMS硬件电路设计和嵌入式软件开发，SOC估算精度实测±2.7%，均衡效率>85%，获得软件著作权1项、实用新型专利1项。",
        },
        {
            "编号": "RD03", "名称": "直流快充桩功率分配系统",
            "领域一级": "新能源与节能", "领域二级": "新能源汽车充电", "领域三级": "直流快充技术",
            "开始时间": "2023-09-01", "结束时间": "2025-08-31",
            "技术来源": "企业自有技术", "IP编号": "IP06,IP07,IP08",
            "预算": 200.0, "总支出": 195.0, "第一年": 80.0, "第二年": 60.0, "第三年": 55.0,
            "人员数": 8,
            "目的": "本项目针对多枪直流快充桩的功率动态分配需求，研发一套基于矩阵开关拓扑的柔性功率分配系统，支持120kW总功率在1-4路输出间按需动态分配，单枪最大输出功率120kW，功率分配响应时间<50ms。",
            "核心技术": "（1）矩阵式直流接触器功率拓扑：设计4×4矩阵开关网络，通过优化接触器投切逻辑实现任意输出回路与任意功率模块的柔性连接，所需接触器数量较传统方案减少40%。（2）优先级动态功率调度算法：建立车辆需求功率-电池SOC-充电时间多目标优化模型，实时解算最优功率分配方案。（3）接触器状态在线检测与保护：通过辅助触点电压检测实现接触器粘连、卡滞故障的快速诊断。",
            "阶段成果": "已完成功率分配系统原理验证和接触器矩阵样机搭建，调度算法仿真通过，获得发明专利1项。正在开展60kW工程样机联调测试。",
        },
    ]

    ps_test = [
        {
            "编号": "PS01", "名称": "高效节能充电模块",
            "领域一级": "新能源与节能", "领域二级": "高效节能技术", "领域三级": "工业节能技术",
            "技术来源": "企业自有技术", "收入": 680.0, "是否主要": "是",
            "IP编号": "IP01,IP02,IP03",
            "关键技术": "本产品采用LLC谐振变换拓扑与DSP数字控制方案，关键技术包括：（1）谐振槽路参数多目标优化设计，实现全负载范围ZVS；（2）自适应死区时间控制技术，最小化体二极管导通损耗；（3）数字化均流并联技术，支持最多8模块并联输出。输入电压范围180-480VAC，满载效率≥95%，待机功耗≤0.5W。",
            "竞争优势": "（1）效率领先：满载效率95%以上，较同类产品提升2-3个百分点；（2）宽输入电压适应性：180-480VAC全覆盖，适配国内外电网标准而无需切换；（3）模块化并联能力：支持8模块并联，单机柜功率可达240kW，系统扩容便捷且维护成本低；（4）全数字化控制：支持远程监控、OTA固件升级，运维智能化水平行业领先。",
            "支持作用": "本产品核心技术由自有三项知识产权（IP01、IP02发明专利，IP03软件著作权）提供支撑。IP01的LLC谐振变换器参数优化方法保障了全负载高效运行，IP02的自适应死区时间控制专利将体二极管损耗降至最低，IP03的均流控制软件实现多模块可靠并联。三项知识产权共同构成产品核心技术壁垒。",
        },
        {
            "编号": "PS02", "名称": "智能电池管理系统",
            "领域一级": "新能源与节能", "领域二级": "新型动力电池", "领域三级": "电池管理系统技术",
            "技术来源": "企业自有技术", "收入": 420.0, "是否主要": "是",
            "IP编号": "IP04,IP05",
            "关键技术": "本产品基于锂电池二阶RC等效电路模型与扩展卡尔曼滤波算法，实现SOC估算精度±3%以内。关键技术包括：（1）在线参数辨识算法，实时更新电池模型参数以适配老化状态；（2）主被动混合均衡策略，根据压差自动切换模式，均衡效率≥85%；（3）多级安全保护机制，涵盖过充/过放/过温/短路保护。",
            "竞争优势": "（1）SOC精度高：动态工况下SOC估算误差<3%，优于行业平均5%的水平；（2）均衡效率优异：主被动混合均衡较纯被动均衡效率提升40%以上；（3）诊断功能完善：支持SOH在线评估、内阻异常检测、热失控预警等智能诊断功能；（4）通信接口丰富：同时支持CAN2.0B、RS485和蓝牙BLE通信。",
            "支持作用": "IP04软件著作权保护了BMS嵌入式控制软件的全部核心算法实现，包括SOC估算、均衡控制和故障诊断逻辑；IP05实用新型专利保护了BMS硬件电路板的关键设计，包括多通道高精度电压采集前端和隔离通信接口。两项知识产权形成了从硬件到软件的完整保护。",
        },
    ]

    ip_test = [
        {
            "编号": "IP01", "名称": "一种基于LLC谐振变换器参数优化设计方法",
            "类别": "发明专利（非国防专利）", "获得方式": "自主研发", "专利号": "ZL202310123456.7",
            "授权日期": "2024-03-15", "所属单位": "宏日嘉科技有限公司",
            "摘要": "本发明公开了一种基于LLC谐振变换器参数优化设计方法，包括以下步骤：建立谐振槽路参数的频域模型；定义多目标优化函数，约束条件包括增益范围、频率范围、谐振电流有效值等；采用粒子群优化算法求解最优参数组合；输出谐振电感、励磁电感和谐振电容的推荐值。本发明的方法能够在全负载范围内实现零电压开关。",
            "先进性说明": "本发明通过多目标优化模型同时考虑增益范围、频率范围和损耗三大约束，解决了传统试凑法设计效率低的痛点。采用粒子群算法全局寻优，相较传统频域分析法的单点设计，优化后的谐振参数可提升满载效率1.5%以上。该方法已通过实测验证，将理论设计到实际样机的参数偏差控制在3%以内。",
            "支持作用说明": "本发明是PS01高效节能充电模块的核心技术支撑，直接决定了LLC变换器在全负载范围内实现零电压开关的能力，是产品达到95%满载效率的技术基础。该专利保护了从参数建模到优化求解的完整技术路线，阻止竞争对手通过参数逆向破解。",
        },
        {
            "编号": "IP02", "名称": "一种自适应死区时间控制电路及控制方法",
            "类别": "发明专利（非国防专利）", "获得方式": "自主研发", "专利号": "ZL202310234567.8",
            "授权日期": "2024-06-20", "所属单位": "宏日嘉科技有限公司",
            "摘要": "本发明公开了一种自适应死区时间控制电路及控制方法，包括谐振电流过零检测模块、死区时间计算模块和PWM生成模块。通过实时检测谐振电流过零点，动态计算最优死区时间，在保证安全跨导的前提下最小化体二极管导通时间。本发明可有效降低体二极管反向恢复损耗，提升LLC变换器整体效率。",
            "先进性说明": "传统固定死区时间方案在全负载范围内效率差异大，轻载时死区时间占比过高导致效率骤降。本发明的自适应方案通过实时谐振电流过零检测，将死区时间精确匹配当前工况，轻载效率较固定死区方案提升2个百分点，且无需额外高压检测电路，成本无增加。",
            "支持作用说明": "本发明是PS01高效节能充电模块实现待机功耗≤0.5W的关键技术，通过最小化体二极管导通时间，将待机损耗降低30%以上。同时该专利保护的自适应控制电路设计难度大，构成了产品的核心技术壁垒，竞争对手难以绕过。",
        },
        {
            "编号": "IP03", "名称": "LLC谐振变换器数字均流控制软件V1.0",
            "类别": "软件著作权", "获得方式": "自主研发", "专利号": "2024SR0567890",
            "授权日期": "2024-02-10", "所属单位": "宏日嘉科技有限公司",
            "摘要": "无",
            "先进性说明": "本软件实现了基于CAN总线的数字化均流控制算法，支持最多8模块并联运行。算法通过周期性广播各模块输出电流，基于加权平均法计算均流基准值，各模块通过PID控制器调节自身输出跟踪基准值。相较于传统模拟均流方案，数字均流精度更高（电流不均衡度<3%），且支持模块热插拔和故障自动隔离。",
            "支持作用说明": "本软件著作权保护了PS01高效节能充电模块的并联均流控制方案，是实现多模块并联扩容的核心技术。随着充电功率需求的不断增长，模块化并联是行业主流技术路线，本软件确保了产品在并联应用场景下的技术竞争力。",
        },
    ]

    ach_test = [
        {
            "序号": 1, "名称": "LLC谐振变换器参数优化设计技术成果",
            "成果类型": "发明专利", "成果来源": "RD01", "转化结果": "PS01",
            "转化时间": "2024-06-01", "关联IP": "IP01", "关联RD": "RD01", "关联PS": "PS01",
            "转化形式": "自行投资实施转化",
            "涉及关键技术": "本成果涉及LLC谐振变换器谐振槽路参数多目标优化设计方法，通过建立包含增益范围、频率范围和损耗约束同时考虑的优化模型，采用粒子群算法实现全局寻优，确保全负载范围零电压开关。该技术实现了满载效率≥95%、参数偏差<3%的核心性能指标。",
            "成效": "该成果已成功应用于PS01高效节能充电模块产品，支撑产品于2024年6月实现量产转化。2024年度实现销售收入680万元，产品已通过多家充电桩厂商的供应商认证。相较上一代产品，效率提升3个百分点，客户运维成本降低约15%，获得下游客户一致好评。",
        },
        {
            "序号": 2, "名称": "自适应死区时间控制技术成果",
            "成果类型": "发明专利", "成果来源": "RD01", "转化结果": "PS01",
            "转化时间": "2024-08-01", "关联IP": "IP02", "关联RD": "RD01", "关联PS": "PS01",
            "转化形式": "自行投资实施转化",
            "涉及关键技术": "本成果涉及自适应死区时间控制技术，通过实时检测LLC谐振电流过零点，动态计算最优死区时间，在保证安全跨导的前提下最小化体二极管导通时间。包含谐振电流过零检测电路、数字死区时间计算模块和PWM生成模块三个核心子技术。",
            "成效": "该成果已集成至PS01产品中并于2024年8月完成产品升级。升级后待机功耗从0.8W降至0.5W以下，轻载效率提升2个百分点。以年出货1000台计算，每年可为终端用户节省电费约12万元，产品能效指标达到行业领先水平。",
        },
        {
            "序号": 3, "名称": "数字均流控制软件技术成果",
            "成果类型": "软件著作权", "成果来源": "RD01", "转化结果": "PS01",
            "转化时间": "2024-04-01", "关联IP": "IP03", "关联RD": "RD01", "关联PS": "PS01",
            "转化形式": "自行投资实施转化",
            "涉及关键技术": "本成果涉及基于CAN总线的数字化均流控制算法，采用加权平均法计算均流基准值，各模块通过PID闭环调节输出电流跟踪基准值。支持8模块并联运行、模块热插拔和故障自动隔离功能，电流不均衡度控制在3%以内。",
            "成效": "该成果作为PS01产品并联扩容的核心软件已投入使用，支持单机柜从30kW灵活扩容至240kW。已有3家客户采购多模块并联系列产品用于大型充电场站建设，2024年实现相关销售收入200万元。模块化架构显著降低了客户的初始投资和后续扩容成本。",
        },
    ]

    injector = TemplateInjector(TEMPLATE_DIR, OUTPUT_DIR, ENTERPRISE)
    injector.inject_rd_table(rd_test)
    injector.inject_ps_table(ps_test)
    injector.inject_ip_table(ip_test)
    injector.inject_achievement_table(ach_test)

    import glob as _glob
    for f in _glob.glob(os.path.join(TEMPLATE_DIR, "*.xlsx")):
        try:
            os.remove(f)
        except OSError:
            pass
    try:
        os.rmdir(TEMPLATE_DIR)
    except OSError:
        pass

    print(f"\n所有表格已生成至: {OUTPUT_DIR}")
