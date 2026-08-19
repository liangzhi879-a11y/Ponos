#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""全量发票 PS 筛选脚本

核心原则：
    PS 名称必须从历史申报材料（高新申报书）中提取，而非从发票货物名称反推。
    只有历史 PS 与当前发票无法匹配时，才进行扩展优化。

功能说明：
    从全量发票 Excel 中按 PS（高新技术产品/服务）筛选发票，生成标注文件和统计表。
    全流程分四步：
      1. extract-ps-baseline：从申请书 PDF 提取 PS 基线（名称/编号/技术领域）
      2. load-invoices：读取全量发票 Excel
      3. match：以申请书 PS 为基准进行匹配
      4. generate-report：生成 PS 发票标注表、PS 统计表、占比分析报告

CLI 用法：
    # 第一步：从申请书 PDF 提取 PS 基线
    python invoice_ps_matcher.py extract-ps-baseline \\
        --application-pdf "申请书.pdf" \\
        --output "PS基线.json"

    # 第二步：加载全量发票 Excel
    python invoice_ps_matcher.py load-invoices \\
        --invoice-file "全量发票.xlsx" \\
        --output "发票明细.json"

    # 第三步：以申请书 PS 为基准匹配发票
    python invoice_ps_matcher.py match \\
        --ps-baseline "PS基线.json" \\
        --invoices "发票明细.json" \\
        --output "匹配结果.json"

    # 第四步：生成报表
    python invoice_ps_matcher.py generate-report \\
        --match-result "匹配结果.json" \\
        --output-dir "输出目录" \\
        --enterprise "企业名称"

匹配算法（按优先级）：
    - 直接匹配（1.0）：货物名称包含完整 PS 名称
    - 关键词匹配（0.8）：货物名称包含 PS 名称的核心关键词
    - 近义匹配（0.6）：货物名称与 PS 名称的近义词匹配
    - 技术领域匹配（0.4）：货物名称属于 PS 技术领域
    - 不匹配（0）：以上均不命中

依赖：
    - openpyxl（必选，读写 xlsx）
    - PyMuPDF/fitz（必选，读取申请书 PDF）
    - jieba（可选，中文分词，提升关键词提取效果）

输出 JSON 结构详见各子命令 docstring。
"""

import os
import re
import sys
import json
import argparse
from datetime import datetime

# v2.6: 权威术语核验模块
try:
    from verify_authoritative_terms import scan_and_correct, verify_tech_field
    _TERM_VERIFY_AVAILABLE = True
except ImportError:
    _TERM_VERIFY_AVAILABLE = False

# ============================================================
# 依赖加载
# ============================================================

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    _OPENPYXL_AVAILABLE = True
except ImportError:
    _OPENPYXL_AVAILABLE = False

try:
    import fitz  # PyMuPDF
    _FITZ_AVAILABLE = True
except ImportError:
    _FITZ_AVAILABLE = False

try:
    import jieba
    _JIEBA_AVAILABLE = True
except ImportError:
    _JIEBA_AVAILABLE = False


# ============================================================
# 常量定义
# ============================================================

# 全量发票 Excel 支持的 sheet 名（按优先级）
INVOICE_SHEET_NAMES = ['发票明细', '开票明细', '全量发票', '发票', 'Sheet1', '明细']

# 发票列名候选（用于自动识别列）
INVOICE_COLUMN_CANDIDATES = {
    'invoice_no': ['发票号码', '发票号', '发票代码', '号码'],
    'invoice_date': ['开票日期', '发票日期', '日期', '开票时间'],
    'buyer_name': ['购方名称', '购货单位', '购买方', '购方', '客户名称', '买方'],
    'seller_name': ['销方名称', '销货单位', '销售方', '销方', '供应商'],
    'goods_name': ['货物名称', '商品名称', '货物或应税劳务名称', '品名', '货物', '商品'],
    'spec': ['规格型号', '规格', '型号'],
    'unit': ['单位'],
    'quantity': ['数量'],
    'unit_price': ['单价'],
    'amount': ['金额', '不含税金额'],
    'tax_rate': ['税率', '征收率'],
    'tax_amount': ['税额', '税金'],
    'total_amount': ['价税合计', '合计', '总金额', '含税金额'],
}

# 匹配度阈值
MATCH_THRESHOLD = 0.4  # 低于此值视为不匹配

# 高新收入占比达标线
HIGH_TECH_RATIO_THRESHOLD = 0.60

# PS 编号正则
PS_NUMBER_PATTERN = re.compile(r'PS\s*0*(\d{1,3})', re.IGNORECASE)

# 申请书中 PS 表的表头关键词（用于定位表格）
PS_TABLE_HEADER_KEYWORDS = ['产品', '编号', '名称', '技术领域', '收入', 'PS']

# 8 大高新技术领域关键词（一级）
TECH_FIELD_KEYWORDS = {
    '电子信息': ['电子', '信息', '软件', '通信', '集成电路', '半导体', '计算机', '网络'],
    '生物与新医药': ['生物', '医药', '制药', '医疗器械', '基因', '疫苗', '中药', '医疗'],
    '航空航天': ['航空', '航天', '飞机', '导弹', '卫星', '火箭', '飞行器'],
    '新材料': ['材料', '合金', '复合材料', '纳米材料', '高分子', '陶瓷', '金属'],
    '高技术服务': ['服务', '咨询', '研发', '外包', '技术服', '检验', '检测'],
    '新能源与节能': ['新能源', '太阳能', '风能', '电池', '节能', '光伏', '储能', '充电'],
    '资源与环境': ['环境', '环保', '污水', '废气', '固废', '资源', '生态', '治理'],
    '先进制造与自动化': ['制造', '自动化', '机器人', '数控', '机械', '智能装备', '产线', '加工'],
}

# 近义词词典（PS 名称 ↔ 货物名称常见近义）
SYNONYM_DICT = {
    '系统': ['平台', '软件', '应用'],
    '软件': ['系统', '平台', '应用', '程序'],
    '设备': ['装置', '机器', '装备', '机'],
    '装置': ['设备', '机器', '装备'],
    '平台': ['系统', '软件'],
    '产品': ['制品', '货物', '商品'],
    '技术': ['工艺', '方法'],
    '服务': ['技术服务', '咨询', '支持'],
    '材料': ['原料', '制品'],
    '组件': ['部件', '零件', '模块'],
    '模块': ['组件', '部件', '单元'],
}


# ============================================================
# 工具函数
# ============================================================

def _to_str(value):
    """安全转字符串，处理 None/float 等类型"""
    if value is None:
        return ''
    if isinstance(value, float):
        # 整数浮点去小数
        if value == int(value):
            return str(int(value))
        return str(value)
    return str(value).strip()


def _to_float(value):
    """安全转浮点，处理 None/字符串/逗号分隔"""
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip().replace(',', '').replace('，', '').replace('¥', '').replace('￥', '')
    s = re.sub(r'[^\d.\-]', '', s)
    try:
        return float(s) if s else 0.0
    except ValueError:
        return 0.0


def _normalize_text(text):
    """文本规范化：去空白、统一全角、转小写"""
    if not text:
        return ''
    s = str(text)
    # 全角转半角（字母数字）
    s = s.translate(str.maketrans('０１２３４５６７８９ａｂｃｄｅｆｇｈｉｊｋｌｍｎｏｐｑｒｓｔｕｖｗｘｙｚＡＢＣＤＥＦＧＨＩＪＫＬＭＮＯＰＱＲＳＴＵＶＷＸＹＺ',
                                   '0123456789abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ'))
    s = re.sub(r'\s+', '', s)
    return s.lower()


def _extract_ps_number(value):
    """从字符串中提取 PS 编号（如 PS01 → 1，PS1 → 1）"""
    if value is None:
        return None
    m = PS_NUMBER_PATTERN.search(str(value))
    return int(m.group(1)) if m else None


def _extract_keywords(text):
    """从文本中提取核心关键词

    优先使用 jieba 分词；不可用时退化为按字数≥2 的切片。
    """
    if not text:
        return []
    text = str(text)
    if _JIEBA_AVAILABLE:
        words = jieba.lcut(text)
        # 过滤停用词、单字、纯数字
        return [w for w in words if len(w) >= 2 and not w.isdigit() and w not in {'有限', '公司', '技术', '有限公', '司'}]
    # 退化：2-3 字滑窗
    return [text[i:i + 2] for i in range(len(text) - 1)]


def _detect_tech_field(text):
    """根据文本识别所属的一级技术领域（v2.6: 结果经权威术语矫正）"""
    if not text:
        return ''
    text_lower = str(text)
    best_field = ''
    best_score = 0
    for field, keywords in TECH_FIELD_KEYWORDS.items():
        score = sum(1 for kw in keywords if kw in text_lower)
        if score > best_score:
            best_score = score
            best_field = field
    # v2.6: 权威术语强制矫正
    if best_field and _TERM_VERIFY_AVAILABLE:
        corrected, corrections = scan_and_correct(best_field)
        if corrections:
            return corrected
    return best_field


# ============================================================
# 第一步：从申请书 PDF 提取 PS 基线
# ============================================================

def extract_ps_from_application(pdf_path):
    """从高新申报书 PDF 中提取 PS 基线（名称、编号、技术领域）

    核心原则：PS 名称来源于申请书，而非发票反推。

    Args:
        pdf_path: 申请书 PDF 路径

    Returns:
        dict: {"ps_items": [{"ps_id": "PS01", "ps_name": "...", "tech_field": "..."}],
               "source": "申请书", "pdf_path": "...", "extracted_at": "ISO时间"}
        失败时返回 {"ps_items": [], "source": "申请书", "error": "..."}
    """
    if not _FITZ_AVAILABLE:
        return {'ps_items': [], 'source': '申请书', 'error': 'PyMuPDF(fitz) 未安装，无法读取 PDF'}
    if not os.path.exists(pdf_path):
        return {'ps_items': [], 'source': '申请书', 'error': f'PDF 不存在: {pdf_path}'}

    try:
        doc = fitz.open(pdf_path)
    except Exception as e:
        return {'ps_items': [], 'source': '申请书', 'error': f'打开 PDF 失败: {e}'}

    # 策略1：优先用 PyMuPDF 的 find_tables 提取表格
    all_tables = []
    for page_no in range(len(doc)):
        page = doc[page_no]
        try:
            tables = page.find_tables()
            for tbl in tables:
                rows = tbl.extract()
                if rows:
                    all_tables.append({'page': page_no + 1, 'rows': rows})
        except Exception:
            pass

    ps_items = []
    # 在所有表格中定位 PS 表（表头包含"产品"和"编号"或"PS"）
    for tbl in all_tables:
        rows = tbl['rows']
        if not rows:
            continue
        # 检测表头行
        header_idx = -1
        for i, row in enumerate(rows[:5]):
            row_text = ' '.join(_to_str(c) for c in row)
            if '产品' in row_text and ('编号' in row_text or 'PS' in row_text.upper() or '名称' in row_text):
                header_idx = i
                break
        if header_idx < 0:
            continue
        header = [_normalize_text(_to_str(c)) for c in rows[header_idx]]
        # 定位列索引
        col_map = {}
        for idx, h in enumerate(header):
            if not h:
                continue
            if '编号' in h and 'ps_id_col' not in col_map:
                col_map['ps_id_col'] = idx
            elif 'ps' in h and 'ps_id_col' not in col_map:
                col_map['ps_id_col'] = idx
            if '名称' in h and 'ps_name_col' not in col_map:
                col_map['ps_name_col'] = idx
            if '技术领域' in h and 'tech_col' not in col_map:
                col_map['tech_col'] = idx
            if '收入' in h and 'revenue_col' not in col_map:
                col_map['revenue_col'] = idx
        if 'ps_name_col' not in col_map:
            continue
        # 提取数据行
        for row in rows[header_idx + 1:]:
            if not row or all(_to_str(c) == '' for c in row):
                continue
            row_text = ' '.join(_to_str(c) for c in row)
            if '总计' in row_text or '合计' in row_text or '小计' in row_text:
                continue
            ps_name = _to_str(row[col_map['ps_name_col']]) if col_map['ps_name_col'] < len(row) else ''
            if not ps_name:
                continue
            # PS 编号
            ps_id = ''
            if 'ps_id_col' in col_map and col_map['ps_id_col'] < len(row):
                ps_id = _to_str(row[col_map['ps_id_col']])
                if not ps_id:
                    ps_id = f'PS{len(ps_items) + 1:02d}'
            else:
                ps_id = f'PS{len(ps_items) + 1:02d}'
            # 标准化 PS 编号格式
            ps_no = _extract_ps_number(ps_id)
            if ps_no is not None:
                ps_id = f'PS{ps_no:02d}'
            # 技术领域
            tech_field = ''
            if 'tech_col' in col_map and col_map['tech_col'] < len(row):
                tech_field = _to_str(row[col_map['tech_col']])
            if not tech_field:
                tech_field = _detect_tech_field(ps_name)
            # v2.6: 权威术语强制矫正
            if tech_field and _TERM_VERIFY_AVAILABLE:
                corrected_field, _ = scan_and_correct(tech_field)
                if corrected_field != tech_field:
                    tech_field = corrected_field
            # 收入（可选）
            revenue = 0.0
            if 'revenue_col' in col_map and col_map['revenue_col'] < len(row):
                revenue = _to_float(row[col_map['revenue_col']])

            ps_items.append({
                'ps_id': ps_id,
                'ps_name': ps_name,
                'tech_field': tech_field,
                'revenue': revenue,
            })

    doc.close()

    # 策略2：若表格提取失败，退化用正则在文本中匹配
    if not ps_items:
        ps_items = _extract_ps_from_text(pdf_path)

    # 去重（按 ps_id）
    seen = set()
    deduped = []
    for item in ps_items:
        if item['ps_id'] not in seen:
            seen.add(item['ps_id'])
            deduped.append(item)

    return {
        'ps_items': deduped,
        'source': '申请书',
        'pdf_path': os.path.abspath(pdf_path),
        'extracted_at': datetime.now().isoformat(timespec='seconds'),
        'count': len(deduped),
    }


def _extract_ps_from_text(pdf_path):
    """退化策略：从 PDF 全文本中用正则提取 PS 条目

    适用于 PDF 无可解析表格的情况。
    """
    if not _FITZ_AVAILABLE:
        return []
    try:
        doc = fitz.open(pdf_path)
    except Exception:
        return []

    full_text = ''
    for page_no in range(len(doc)):
        page = doc[page_no]
        full_text += page.get_text() + '\n'
    doc.close()

    # 匹配 PS01xxx：xxx 或 PS01 xxx 等
    pattern = re.compile(r'(PS\s*0*\d{1,3})\s*[:：、\.\s]*([^\n PS]{2,80})', re.IGNORECASE)
    items = []
    seen = set()
    for m in pattern.finditer(full_text):
        ps_id_raw = m.group(1)
        ps_name = m.group(2).strip().rstrip('。，；、')
        ps_no = _extract_ps_number(ps_id_raw)
        if ps_no is None:
            continue
        ps_id = f'PS{ps_no:02d}'
        if ps_id in seen:
            continue
        # 过滤掉明显不像产品名称的（如纯数字、太短）
        if len(ps_name) < 2 or ps_name.isdigit():
            continue
        seen.add(ps_id)
        items.append({
            'ps_id': ps_id,
            'ps_name': ps_name,
            'tech_field': _detect_tech_field(ps_name),
            'revenue': 0.0,
        })
    return items


# ============================================================
# 第二步：加载全量发票 Excel
# ============================================================

def load_full_invoices(xlsx_path):
    """加载全量发票 Excel

    支持多种 sheet 名：发票明细、开票明细、全量发票等。
    自动识别列：发票号码、开票日期、购方名称、货物名称、价税合计、金额、税率等。

    Args:
        xlsx_path: 全量发票 Excel 路径

    Returns:
        dict: {"invoices": [...], "total_count": N, "total_amount": M,
               "sheet_name": "...", "columns": {...}}
    """
    if not _OPENPYXL_AVAILABLE:
        return {'invoices': [], 'total_count': 0, 'total_amount': 0.0,
                'error': 'openpyxl 未安装'}
    if not os.path.exists(xlsx_path):
        return {'invoices': [], 'total_count': 0, 'total_amount': 0.0,
                'error': f'文件不存在: {xlsx_path}'}

    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    except Exception as e:
        return {'invoices': [], 'total_count': 0, 'total_amount': 0.0,
                'error': f'加载 Excel 失败: {e}'}

    # 选择 sheet
    ws = None
    sheet_name = ''
    for name in INVOICE_SHEET_NAMES:
        if name in wb.sheetnames:
            ws = wb[name]
            sheet_name = name
            break
    if ws is None:
        # 取第一个 sheet
        ws = wb.worksheets[0]
        sheet_name = ws.title

    # 自动检测表头行（前 5 行中包含至少 2 个候选列名）
    header_row = 1
    header_map = {}
    for r in range(1, min(8, ws.max_row + 1)):
        detected = {}
        for c in range(1, ws.max_column + 1):
            cell_val = _to_str(ws.cell(r, c).value)
            if not cell_val:
                continue
            for field, candidates in INVOICE_COLUMN_CANDIDATES.items():
                if field in detected:
                    continue
                for cand in candidates:
                    if cand in cell_val:
                        detected[field] = c
                        break
        if len(detected) >= 2:
            header_row = r
            header_map = detected
            break

    if not header_map:
        # 无表头，按列序号默认映射
        header_map = {
            'invoice_no': 1, 'invoice_date': 2, 'buyer_name': 3,
            'goods_name': 4, 'spec': 5, 'unit': 6, 'quantity': 7,
            'unit_price': 8, 'amount': 9, 'tax_rate': 10,
            'tax_amount': 11, 'total_amount': 12,
        }

    # 读取数据行
    invoices = []
    total_amount = 0.0
    for r in range(header_row + 1, ws.max_row + 1):
        # 跳过空行
        row_has_data = False
        for c in header_map.values():
            if c <= ws.max_column and ws.cell(r, c).value not in (None, ''):
                row_has_data = True
                break
        if not row_has_data:
            continue

        # 跳过合计/总计行
        first_cell = _to_str(ws.cell(r, 1).value)
        if first_cell in ('合计', '总计', '小计') or '合计' in first_cell and len(first_cell) <= 4:
            continue

        inv = {}
        for field, col in header_map.items():
            if col <= ws.max_column:
                inv[field] = ws.cell(r, col).value
            else:
                inv[field] = None

        # 标准化字段
        inv['invoice_no'] = _to_str(inv.get('invoice_no'))
        inv['invoice_date'] = _to_str(inv.get('invoice_date'))
        inv['buyer_name'] = _to_str(inv.get('buyer_name'))
        inv['seller_name'] = _to_str(inv.get('seller_name'))
        inv['goods_name'] = _to_str(inv.get('goods_name'))
        inv['spec'] = _to_str(inv.get('spec'))
        inv['unit'] = _to_str(inv.get('unit'))
        inv['quantity'] = _to_float(inv.get('quantity'))
        inv['unit_price'] = _to_float(inv.get('unit_price'))
        inv['amount'] = _to_float(inv.get('amount'))
        inv['tax_rate'] = _to_str(inv.get('tax_rate'))
        inv['tax_amount'] = _to_float(inv.get('tax_amount'))
        inv['total_amount'] = _to_float(inv.get('total_amount'))

        # 若价税合计为 0 但金额+税额>0，自动求和
        if inv['total_amount'] == 0 and (inv['amount'] + inv['tax_amount']) > 0:
            inv['total_amount'] = inv['amount'] + inv['tax_amount']

        inv['row_index'] = r
        total_amount += inv['total_amount']
        invoices.append(inv)

    wb.close()

    return {
        'invoices': invoices,
        'total_count': len(invoices),
        'total_amount': round(total_amount, 2),
        'sheet_name': sheet_name,
        'columns': {k: v for k, v in header_map.items()},
        'source_file': os.path.abspath(xlsx_path),
        'loaded_at': datetime.now().isoformat(timespec='seconds'),
    }


# ============================================================
# 第三步：发票与 PS 匹配
# ============================================================

def match_invoice_to_ps(invoice_row, ps_item):
    """单张发票与单个 PS 的匹配度计算

    匹配规则（按优先级，命中即返回）：
        1. 直接匹配（1.0）：货物名称包含完整 PS 名称
        2. 关键词匹配（0.8）：货物名称包含 PS 名称的核心关键词
        3. 近义匹配（0.6）：货物名称与 PS 名称的近义词匹配
        4. 技术领域匹配（0.4）：货物名称属于 PS 技术领域
        5. 不匹配（0）

    Args:
        invoice_row: 发票字典（含 goods_name 等字段）
        ps_item: PS 字典（含 ps_name, tech_field 等）

    Returns:
        tuple: (匹配度 float, 匹配类型 str, 命中关键词 str)
    """
    goods_name = _normalize_text(invoice_row.get('goods_name', ''))
    ps_name = _normalize_text(ps_item.get('ps_name', ''))
    tech_field = ps_item.get('tech_field', '')

    if not goods_name or not ps_name:
        # 若货物名为空但技术领域命中，给 0.4
        if goods_name and tech_field:
            if _is_tech_field_match(goods_name, tech_field):
                return 0.4, 'tech_field', tech_field
        return 0.0, 'no_match', ''

    # 1. 直接匹配：货物名称包含完整 PS 名称
    if ps_name in goods_name:
        return 1.0, 'direct', ps_item.get('ps_name', '')

    # PS 名称的子串分段匹配（PS 名称可能由多个词组成，任一词段命中视为直接匹配）
    ps_segments = _split_segments(ps_item.get('ps_name', ''))
    direct_hits = [seg for seg in ps_segments if len(seg) >= 3 and _normalize_text(seg) in goods_name]
    if direct_hits and len(''.join(direct_hits)) >= len(ps_name) * 0.6:
        return 1.0, 'direct', '|'.join(direct_hits)

    # 2. 关键词匹配：货物名称包含 PS 名称的核心关键词
    ps_keywords = _extract_keywords(ps_item.get('ps_name', ''))
    if ps_keywords:
        goods_lower = goods_name
        hit_kws = [kw for kw in ps_keywords if _normalize_text(kw) in goods_lower]
        if hit_kws:
            # 命中关键词数 / 总关键词数 ≥ 0.5 → 0.8；否则按比例
            ratio = len(hit_kws) / len(ps_keywords)
            if ratio >= 0.5:
                return 0.8, 'keyword', '|'.join(hit_kws)
            if ratio >= 0.3:
                return 0.6, 'synonym', '|'.join(hit_kws)

    # 3. 近义匹配：通过近义词词典
    synonym_hits = _check_synonym_match(goods_name, ps_name)
    if synonym_hits:
        return 0.6, 'synonym', '|'.join(synonym_hits)

    # 4. 技术领域匹配
    if tech_field and _is_tech_field_match(goods_name, tech_field):
        return 0.4, 'tech_field', tech_field

    return 0.0, 'no_match', ''


def _split_segments(text):
    """将 PS 名称切分为有意义的词段（用于直接匹配的子串检查）"""
    if not text:
        return []
    text = str(text)
    # 按常见分隔符切分
    parts = re.split(r'[（）\(\)、，,/\\\s·\-·]+', text)
    return [p.strip() for p in parts if p.strip() and len(p.strip()) >= 2]


def _check_synonym_match(goods_norm, ps_norm):
    """检查货物名称与 PS 名称的近义词匹配"""
    hits = []
    for src, syns in SYNONYM_DICT.items():
        if src in ps_norm:
            for syn in syns:
                if syn in goods_norm and syn != src:
                    hits.append(f'{src}→{syn}')
    return hits


def _is_tech_field_match(goods_name, tech_field):
    """检查货物名称是否属于指定技术领域"""
    if not tech_field or not goods_name:
        return False
    kws = TECH_FIELD_KEYWORDS.get(tech_field, [])
    return any(kw in goods_name for kw in kws)


def match_invoices_to_ps_list(invoices, ps_baseline):
    """批量匹配发票与 PS 列表

    以申请书 PS 基线为基准，每张发票分配到匹配度最高的 PS（≥阈值）。
    未达阈值的发票进入 unmatched_invoices。

    Args:
        invoices: 发票列表
        ps_baseline: PS 基线 dict（含 ps_items）

    Returns:
        dict: {
            "matched": [{"ps_id": "...", "ps_name": "...", "invoices": [...]}],
            "unmatched_invoices": [...],
            "unmatched_ps": [...],
            "coverage": X.XX,
            "stats": {...}
        }
    """
    ps_items = ps_baseline.get('ps_items', [])
    matched = []
    unmatched_invoices = []
    unmatched_ps = []

    # 初始化每个 PS 的发票桶
    ps_buckets = {}
    for ps in ps_items:
        ps_buckets[ps['ps_id']] = {
            'ps_id': ps['ps_id'],
            'ps_name': ps['ps_name'],
            'tech_field': ps.get('tech_field', ''),
            'revenue': ps.get('revenue', 0.0),
            'invoices': [],
            'total_amount': 0.0,
        }

    # 逐张发票匹配
    for inv in invoices:
        best_score = 0.0
        best_type = 'no_match'
        best_keyword = ''
        best_ps_id = None
        # 每张发票只匹配一个 PS（取最高分）
        for ps in ps_items:
            score, mtype, kw = match_invoice_to_ps(inv, ps)
            if score > best_score:
                best_score = score
                best_type = mtype
                best_keyword = kw
                best_ps_id = ps['ps_id']

        if best_score >= MATCH_THRESHOLD and best_ps_id is not None:
            inv_record = dict(inv)
            inv_record['matched_ps_id'] = best_ps_id
            inv_record['match_score'] = round(best_score, 2)
            inv_record['match_type'] = best_type
            inv_record['match_keyword'] = best_keyword
            ps_buckets[best_ps_id]['invoices'].append(inv_record)
            ps_buckets[best_ps_id]['total_amount'] += inv.get('total_amount', 0.0)
        else:
            unmatched_inv = dict(inv)
            unmatched_inv['match_score'] = round(best_score, 2)
            unmatched_inv['match_type'] = best_type
            unmatched_inv['match_keyword'] = best_keyword
            unmatched_invoices.append(unmatched_inv)

    # 汇总
    for ps in ps_items:
        bucket = ps_buckets[ps['ps_id']]
        if bucket['invoices']:
            bucket['total_amount'] = round(bucket['total_amount'], 2)
            matched.append(bucket)
        else:
            unmatched_ps.append({
                'ps_id': ps['ps_id'],
                'ps_name': ps['ps_name'],
                'tech_field': ps.get('tech_field', ''),
                'reason': '无任何发票匹配到该 PS',
            })

    # 覆盖率 = 已匹配发票金额 / 全量发票金额
    total_invoice_amount = sum(inv.get('total_amount', 0.0) for inv in invoices)
    matched_amount = sum(b['total_amount'] for b in matched)
    coverage = round(matched_amount / total_invoice_amount, 4) if total_invoice_amount > 0 else 0.0

    return {
        'matched': matched,
        'unmatched_invoices': unmatched_invoices,
        'unmatched_ps': unmatched_ps,
        'coverage': coverage,
        'stats': {
            'total_invoices': len(invoices),
            'matched_invoices': sum(len(b['invoices']) for b in matched),
            'unmatched_invoices_count': len(unmatched_invoices),
            'total_ps': len(ps_items),
            'matched_ps': len(matched),
            'unmatched_ps_count': len(unmatched_ps),
            'total_invoice_amount': round(total_invoice_amount, 2),
            'matched_amount': round(matched_amount, 2),
            'unmatched_amount': round(total_invoice_amount - matched_amount, 2),
        },
        'source': ps_baseline.get('source', '申请书'),
        'matched_at': datetime.now().isoformat(timespec='seconds'),
    }


# ============================================================
# 第四步：生成报表
# ============================================================

def calculate_high_tech_ratio(match_result):
    """计算高新收入占比

    高新收入 = 所有已匹配 PS 的发票价税合计之和
    全量发票合计 = 所有发票价税合计之和
    高新收入占比 = 高新收入 / 全量发票合计

    Args:
        match_result: match 子命令返回的结果

    Returns:
        dict: {
            "high_tech_revenue": 高新收入,
            "total_invoice_amount": 全量发票合计,
            "high_tech_ratio": 占比,
            "threshold": 0.60,
            "qualified": bool,
        }
    """
    stats = match_result.get('stats', {})
    high_tech_revenue = stats.get('matched_amount', 0.0)
    total_amount = stats.get('total_invoice_amount', 0.0)
    ratio = round(high_tech_revenue / total_amount, 4) if total_amount > 0 else 0.0
    return {
        'high_tech_revenue': round(high_tech_revenue, 2),
        'total_invoice_amount': round(total_amount, 2),
        'high_tech_ratio': ratio,
        'threshold': HIGH_TECH_RATIO_THRESHOLD,
        'qualified': ratio >= HIGH_TECH_RATIO_THRESHOLD,
    }


def generate_ps_invoice_summary(match_result, output_path, enterprise='企业'):
    """生成 PS 统计表 xlsx

    列：PS编号/PS名称/发票数量/金额合计/价税合计/占高新收入比例/占全量发票比例

    Args:
        match_result: match 子命令返回的结果
        output_path: 输出 xlsx 路径
        enterprise: 企业名称（写入 sheet 名/标题）
    """
    if not _OPENPYXL_AVAILABLE:
        raise RuntimeError('openpyxl 未安装，无法生成 xlsx')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'PS统计表'

    # 样式
    header_font = Font(name='宋体', size=12, bold=True)
    cell_font = Font(name='宋体', size=11)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left = Alignment(horizontal='left', vertical='center', wrap_text=True)

    headers = ['PS编号', 'PS名称', '发票数量', '金额合计(元)', '价税合计(元)',
               '占高新收入比例', '占全量发票比例']
    for c, h in enumerate(headers, 1):
        cell = ws.cell(1, c, h)
        cell.font = header_font
        cell.alignment = center
        cell.border = thin_border

    stats = match_result.get('stats', {})
    total_invoice_amount = stats.get('total_invoice_amount', 0.0)
    high_tech_total = stats.get('matched_amount', 0.0)

    r = 2
    for bucket in match_result.get('matched', []):
        inv_count = len(bucket['invoices'])
        amount_sum = round(sum(inv.get('amount', 0.0) for inv in bucket['invoices']), 2)
        total_sum = round(bucket['total_amount'], 2)
        ratio_in_high = round(total_sum / high_tech_total, 4) if high_tech_total > 0 else 0.0
        ratio_in_all = round(total_sum / total_invoice_amount, 4) if total_invoice_amount > 0 else 0.0
        row_data = [
            bucket['ps_id'], bucket['ps_name'], inv_count, amount_sum, total_sum,
            f'{ratio_in_high * 100:.2f}%', f'{ratio_in_all * 100:.2f}%',
        ]
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(r, c, val)
            cell.font = cell_font
            cell.alignment = left if c == 2 else center
            cell.border = thin_border
        r += 1

    # 合计行
    total_count = sum(len(b['invoices']) for b in match_result.get('matched', []))
    total_amount_sum = round(sum(inv.get('amount', 0.0) for b in match_result.get('matched', []) for inv in b['invoices']), 2)
    total_total_sum = round(sum(b['total_amount'] for b in match_result.get('matched', [])), 2)
    row_data = ['合计', '', total_count, total_amount_sum, total_total_sum, '100.00%',
                f'{(total_total_sum / total_invoice_amount * 100) if total_invoice_amount > 0 else 0:.2f}%']
    for c, val in enumerate(row_data, 1):
        cell = ws.cell(r, c, val)
        cell.font = Font(name='宋体', size=11, bold=True)
        cell.alignment = center
        cell.border = thin_border

    # 列宽
    col_widths = [10, 32, 10, 16, 16, 14, 14]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # 标题行（在第0行插入）
    ws.insert_rows(1)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    title_cell = ws.cell(1, 1, f'{enterprise}-PS发票统计表')
    title_cell.font = Font(name='宋体', size=14, bold=True)
    title_cell.alignment = center

    wb.save(output_path)
    wb.close()


def _generate_ps_invoice_marked_table(match_result, output_path, enterprise='企业'):
    """生成 PS 发票标注表 xlsx

    每个 PS 一个 Sheet，含列：序号/发票号码/开票日期/购方名称/货物名称/价税合计/匹配度
    """
    if not _OPENPYXL_AVAILABLE:
        raise RuntimeError('openpyxl 未安装，无法生成 xlsx')

    wb = openpyxl.Workbook()
    # 删除默认 sheet
    default_ws = wb.active
    wb.remove(default_ws)

    header_font = Font(name='宋体', size=12, bold=True)
    cell_font = Font(name='宋体', size=11)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left = Alignment(horizontal='left', vertical='center', wrap_text=True)

    headers = ['序号', '发票号码', '开票日期', '购方名称', '货物名称', '价税合计(元)', '匹配度', '匹配类型']

    for bucket in match_result.get('matched', []):
        # sheet 名不能超过 31 字符，不能含特殊字符
        sheet_name = f"{bucket['ps_id']}_{bucket['ps_name']}"[:28]
        sheet_name = re.sub(r'[\\/\?\*\[\]:]', '_', sheet_name)
        ws = wb.create_sheet(sheet_name)

        # 标题行
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
        title_cell = ws.cell(1, 1, f"{bucket['ps_id']} {bucket['ps_name']}（{enterprise}）")
        title_cell.font = Font(name='宋体', size=14, bold=True)
        title_cell.alignment = center

        # 表头
        for c, h in enumerate(headers, 1):
            cell = ws.cell(2, c, h)
            cell.font = header_font
            cell.alignment = center
            cell.border = thin_border

        # 数据行
        for idx, inv in enumerate(bucket['invoices'], 1):
            row_data = [
                idx, inv.get('invoice_no', ''), inv.get('invoice_date', ''),
                inv.get('buyer_name', ''), inv.get('goods_name', ''),
                inv.get('total_amount', 0.0),
                inv.get('match_score', 0.0), inv.get('match_type', ''),
            ]
            for c, val in enumerate(row_data, 1):
                cell = ws.cell(2 + idx, c, val)
                cell.font = cell_font
                cell.alignment = left if c in (4, 5) else center
                cell.border = thin_border

        # 列宽
        col_widths = [6, 20, 14, 28, 36, 14, 8, 10]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # 若无任何 PS 匹配，保留一个空 sheet 提示
    if not wb.sheetnames:
        ws = wb.create_sheet('无匹配')
        ws.cell(1, 1, '未匹配到任何 PS（请检查申请书 PS 提取与发票匹配阈值）')

    wb.save(output_path)
    wb.close()


def _generate_ratio_report(match_result, output_path, enterprise='企业'):
    """生成占比分析报告 JSON"""
    ratio_info = calculate_high_tech_ratio(match_result)
    stats = match_result.get('stats', {})
    report = {
        'enterprise': enterprise,
        'generated_at': datetime.now().isoformat(timespec='seconds'),
        'source': match_result.get('source', '申请书'),
        'high_tech_revenue': ratio_info['high_tech_revenue'],
        'total_invoice_amount': ratio_info['total_invoice_amount'],
        'high_tech_ratio': ratio_info['high_tech_ratio'],
        'threshold': ratio_info['threshold'],
        'qualified': ratio_info['qualified'],
        'ps_count': stats.get('matched_ps', 0),
        'unmatched_ps_count': stats.get('unmatched_ps_count', 0),
        'unmatched_invoice_count': stats.get('unmatched_invoices_count', 0),
        'unmatched_invoice_amount': stats.get('unmatched_amount', 0.0),
        'conclusion': (
            f"高新收入占比 {ratio_info['high_tech_ratio'] * 100:.2f}% "
            f"{'达到' if ratio_info['qualified'] else '未达到'} "
            f"{ratio_info['threshold'] * 100:.0f}% 达标线"
        ),
        'unmatched_ps_list': [
            {'ps_id': p['ps_id'], 'ps_name': p['ps_name']}
            for p in match_result.get('unmatched_ps', [])
        ],
    }
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(report, f, ensure_ascii=False, indent=2)


# ============================================================
# CLI 命令实现
# ============================================================

def _cmd_extract_ps_baseline(args):
    """子命令：extract-ps-baseline"""
    result = extract_ps_from_application(args.application_pdf)
    if args.output:
        os.makedirs(os.path.dirname(os.path.abspath(args.output)) or '.', exist_ok=True)
        with open(args.output, 'w', encoding='utf-8') as f:
            json.dump(result, f, ensure_ascii=False, indent=2)
        print(f'[OK] PS 基线已保存：{args.output}')
        print(f'     共提取 {result.get("count", 0)} 个 PS')
    else:
        print(json.dumps(result, ensure_ascii=False, indent=2))
    return result


def _cmd_load_invoices(args):
    """子命令：load-invoices"""
    result = load_full_invoices(args.invoice_file)
    if args.output:
        os.makedirs(os.path.dirname(os.path.abspath(args.output)) or '.', exist_ok=True)
        with open(args.output, 'w', encoding='utf-8') as f:
            json.dump(result, f, ensure_ascii=False, indent=2)
        print(f'[OK] 发票明细已保存：{args.output}')
        print(f'     共加载 {result.get("total_count", 0)} 张发票，'
              f'价税合计 {result.get("total_amount", 0):.2f} 元')
    else:
        print(json.dumps(result, ensure_ascii=False, indent=2))
    return result


def _cmd_match(args):
    """子命令：match"""
    with open(args.ps_baseline, 'r', encoding='utf-8') as f:
        ps_baseline = json.load(f)
    with open(args.invoices, 'r', encoding='utf-8') as f:
        inv_data = json.load(f)
    invoices = inv_data.get('invoices', [])
    result = match_invoices_to_ps_list(invoices, ps_baseline)
    if args.output:
        os.makedirs(os.path.dirname(os.path.abspath(args.output)) or '.', exist_ok=True)
        with open(args.output, 'w', encoding='utf-8') as f:
            json.dump(result, f, ensure_ascii=False, indent=2)
        print(f'[OK] 匹配结果已保存：{args.output}')
        print(f'     匹配发票 {result["stats"]["matched_invoices"]}/{result["stats"]["total_invoices"]}，'
              f'覆盖率 {result["coverage"] * 100:.2f}%')
    else:
        print(json.dumps(result, ensure_ascii=False, indent=2))
    return result


def _cmd_generate_report(args):
    """子命令：generate-report"""
    with open(args.match_result, 'r', encoding='utf-8') as f:
        match_result = json.load(f)

    os.makedirs(args.output_dir, exist_ok=True)
    enterprise = args.enterprise or '企业'

    # 1. PS 发票标注表
    marked_path = os.path.join(args.output_dir, 'PS发票标注表.xlsx')
    _generate_ps_invoice_marked_table(match_result, marked_path, enterprise)

    # 2. PS 统计表
    summary_path = os.path.join(args.output_dir, 'PS统计表.xlsx')
    generate_ps_invoice_summary(match_result, summary_path, enterprise)

    # 3. 占比分析报告
    ratio_path = os.path.join(args.output_dir, '占比分析报告.json')
    _generate_ratio_report(match_result, ratio_path, enterprise)

    result = {
        'marked_table': marked_path,
        'summary_table': summary_path,
        'ratio_report': ratio_path,
        'enterprise': enterprise,
        'generated_at': datetime.now().isoformat(timespec='seconds'),
    }
    print(f'[OK] 报表已生成于：{args.output_dir}')
    print(f'     - PS发票标注表：{marked_path}')
    print(f'     - PS统计表：{summary_path}')
    print(f'     - 占比分析报告：{ratio_path}')
    return result


def build_parser():
    """构建 argparse 解析器"""
    parser = argparse.ArgumentParser(
        prog='invoice_ps_matcher',
        description='全量发票 PS 筛选脚本（PS 名称来源于申请书，非发票反推）',
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    subparsers = parser.add_subparsers(dest='command', help='子命令')

    # extract-ps-baseline
    p1 = subparsers.add_parser('extract-ps-baseline', help='从申请书 PDF 提取 PS 基线')
    p1.add_argument('--application-pdf', required=True, help='申请书 PDF 路径')
    p1.add_argument('--output', help='输出 JSON 路径（省略则输出到 stdout）')
    p1.set_defaults(func=_cmd_extract_ps_baseline)

    # load-invoices
    p2 = subparsers.add_parser('load-invoices', help='加载全量发票 Excel')
    p2.add_argument('--invoice-file', required=True, help='全量发票 Excel 路径')
    p2.add_argument('--output', help='输出 JSON 路径（省略则输出到 stdout）')
    p2.set_defaults(func=_cmd_load_invoices)

    # match
    p3 = subparsers.add_parser('match', help='以申请书 PS 为基准匹配发票')
    p3.add_argument('--ps-baseline', required=True, help='PS 基线 JSON 路径')
    p3.add_argument('--invoices', required=True, help='发票明细 JSON 路径')
    p3.add_argument('--output', help='输出匹配结果 JSON 路径')
    p3.set_defaults(func=_cmd_match)

    # generate-report
    p4 = subparsers.add_parser('generate-report', help='生成 PS 发票标注表/统计表/占比分析报告')
    p4.add_argument('--match-result', required=True, help='匹配结果 JSON 路径')
    p4.add_argument('--output-dir', required=True, help='输出目录')
    p4.add_argument('--enterprise', default='企业', help='企业名称（默认"企业"）')
    p4.set_defaults(func=_cmd_generate_report)

    return parser


def main():
    """CLI 主入口"""
    parser = build_parser()
    args = parser.parse_args()
    if not args.command:
        parser.print_help()
        return 1
    try:
        args.func(args)
    except Exception as e:
        print(f'[ERROR] {type(e).__name__}: {e}', file=sys.stderr)
        return 1
    return 0


if __name__ == '__main__':
    sys.exit(main())
