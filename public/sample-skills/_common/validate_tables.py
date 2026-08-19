"""核心表格模板对齐校验工具（v1.14.0）

对照 00 高新模板（全） 目录下的5个模板文件，校验agent生成的核心表格是否符合：
1. 字段数、字段名、字段顺序与模板一致
2. 样式（字体/字号/边框/对齐）与模板一致
3. 长文本字段字数在300-400字范围
4. 禁止合并标题行
5. 必填字段非空

用法：
  python validate_tables.py --dir "00_核心表格目录路径"
  python validate_tables.py --file "单个xlsx路径" --type ip|rd|ps|ach|rdps
  python validate_tables.py --enterprise "深圳市XX公司" --year 2026

依赖：openpyxl
"""
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
import os
import sys
import json
import argparse
import re
import glob
from pathlib import Path

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from path_config import get_template_dir_str, get_enterprise_data_dir

# ============================================================
# 模板定义（严格对齐 00 高新模板（全） 目录的5个文件，路径基于 path_config 推断）
# ============================================================

TEMPLATE_DIR = get_template_dir_str()

# 5种表格的模板定义
TABLE_SPECS = {
    'ip': {
        'name': '知识产权表',
        'template_file': '前海云充科技-知识产权表（参与本次创新能力知识产权评价，汇总信息只统计此列表中的知识产权）.xlsx',
        'expected_columns': 10,
        'expected_headers': [
            '知识产权编号', '知识产权名称', '类别', '获得方式',
            '专利号/著作权号', '授权日期', '知识产权所属单位或个人',
            '国家知识产权局官方网站上公布的摘要(限400字，软件著作权不用提供)',
            '与本知识产权相关的核心关键技术先进性说明(限400字)',
            '该知识产权与本企业产品（服务）核心技术的支持作用说明(限400字)',
        ],
        # 摘要(第8列)严格按专利说明书原文摘录，不校验字数下限
        # 先进性说明(第9列)、支持作用(第10列) 要求300-400字
        # v1.18.2：软著摘要必须填"无"（禁止空值），由专门的校验逻辑处理
        'long_text_columns': [8, 9, 10],  # 1-indexed，第8列软著摘要走特殊校验
        'allow_empty_long_text': {},  # v1.18.2：不再允许空值，软著必须填"无"
        'header_row': 1,
    },
    'rd': {
        'name': '企业研究开发活动汇总表',
        'template_file': '前海云充科技-企业研究开发活动汇总表（近三年执行的活动）.xlsx',
        'expected_columns': 15,
        'expected_headers': [
            '研发活动编号', '研发活动名称',
            '技术领域（一级）', '技术领域（二级）', '技术领域（三级）',
            '开始时间', '结束时间', '技术来源', '知识产权编号',
            '研发经费总预算（万元）', '研发经费近三年总支出（万元）',
            '其中：第一年（2021年）支出（万元）',
            '其中：第二年（2022年）支出（万元）',
            '其中：第三年（2023年）支出（万元）',
            '研发活动人员数',
        ],
        # RD模板实际只有15列基础字段，但 RD表还会附加3个长文本字段（在另一个工作表或文档中）
        'long_text_columns': [],  # 主表无长文本字段，长文本字段在 .docx 附件中
        'header_row': 1,
    },
    'ps': {
        'name': '高新技术产品（服务）明细表',
        'template_file': '前海云充科技-高新技术产品（服务）明细表.xlsx',
        'expected_columns': 12,
        'expected_headers': [
            '产品（服务）编号', '产品（服务）名称',
            '技术领域（一级）', '技术领域（二级）', '技术领域（三级）',
            '技术来源', '上年度销售收入 （万元）', '是否主要产品 （服务）',
            '知识产权编号',
            '关键技术及主要技术指标（限400字）',
            '与同类产品（服务）的竞争优势（限400字）',
            '知识产权获得情况及其对产品（服务）在技术上发挥的支持作用（限400字）',
        ],
        'long_text_columns': [10, 11, 12],
        'header_row': 1,
    },
    'ach': {
        'name': '科技成果转化情况汇总表',
        'template_file': '前海云充科技-科技成果转化情况汇总表.xlsx',
        'expected_columns': 12,
        'expected_headers': [
            '科技成果序号', '科技成果名称', '成果类型', '成果来源',
            '转化结果', '转化时间', '关联IP', '关联RD', '关联PS',
            '转化形式', '涉及关键技术（限400字）', '成效（限400字）',
        ],
        'long_text_columns': [11, 12],
        'min_chars': 370,  # 成果转化字段字数下限370（提示词要求370-410字）
        'max_chars': 410,  # 允许超过400字到410字
        'header_row': 1,
    },
    'rdps': {
        'name': 'RD/PS汇总表',
        'template_file': '前海云充21-23年研发项目RD21个名称、高新产品PS？个.xls',
        'expected_columns': 9,
        'expected_headers': [
            '编号', '研发项目名称', '项目起止日期', '研发人数',
            '实际研发费（万元）', '研发预算', '知识产权', '高新产品',
        ],
        'long_text_columns': [],
        'header_row': 2,  # 模板第1行是标题，第2行才是表头
    },
    # v1.16.0新增：科技人员清单（三方对比）独立规格
    # 不基于固定模板，由 generate_tables_from_template.py 代码生成
    # 参考：定稿-云充2023年科技人员信息表.xlsx（13列基础字段）+ 三方对比扩展字段
    'staff': {
        'name': '科技人员清单（三方对比）',
        'template_file': None,  # 无固定模板，代码生成
        'expected_columns': 15,
        'expected_headers': [
            '序号', '姓名', '身份证号码', '性别', '部门', '岗位',
            '入职日期', '离职日期', '上年在职天数',
            '花名册', '社保清单', '台账',
            '是否符合科技人员条件', '不符合原因', '备注',
        ],
        'long_text_columns': [],  # 无长文本字段
        'header_row': 1,
        # v1.16.0新增校验项：末尾统计行
        'require_summary_rows': True,
        'summary_labels': [
            '花名册总人数', '社保清单总人数', '台账纳入总人数',
            '符合科技人员条件人数', '科技人员占比（基于社保）',
        ],
    },
}


def validate_text_length(text, min_chars=300, max_chars=400, allow_empty=False):
    """校验文本字数是否在指定范围内"""
    if text is None or (isinstance(text, str) and text.strip() == ''):
        if allow_empty:
            return (True, '空值（允许）')
        return (False, '空值')
    length = len(str(text))
    if length < min_chars:
        return (False, f'字数不足：{length} < {min_chars}')
    if length > max_chars:
        return (False, f'字数超限：{length} > {max_chars}')
    return (True, f'字数合规：{length}')


def validate_table_format(file_path, table_type, min_chars=300, max_chars=400):
    """校验单个表格是否符合模板规范

    返回 {'passed': bool, 'errors': [], 'warnings': [], 'stats': {}}
    """
    result = {
        'file': os.path.basename(file_path),
        'table_type': table_type,
        'table_name': TABLE_SPECS[table_type]['name'],
        'passed': True,
        'errors': [],
        'warnings': [],
        'stats': {},
    }

    if not os.path.exists(file_path):
        result['passed'] = False
        result['errors'].append(f'文件不存在: {file_path}')
        return result

    spec = TABLE_SPECS[table_type]
    expected_cols = spec['expected_columns']
    expected_headers = spec['expected_headers']
    header_row = spec['header_row']

    try:
        wb = openpyxl.load_workbook(file_path, data_only=False)
    except Exception as e:
        result['passed'] = False
        result['errors'].append(f'加载Excel失败: {e}')
        return result

    ws = wb.active
    actual_rows = ws.max_row
    actual_cols = ws.max_column
    result['stats']['rows'] = actual_rows
    result['stats']['cols'] = actual_cols

    # 1. 列数校验
    if actual_cols < expected_cols:
        result['passed'] = False
        result['errors'].append(f'列数不足: 实际{actual_cols} < 期望{expected_cols}')
    elif actual_cols > expected_cols:
        result['warnings'].append(f'列数超出: 实际{actual_cols} > 期望{expected_cols}')

    # 2. 合并单元格校验（禁止合并标题行）
    if ws.merged_cells.ranges:
        for r in ws.merged_cells.ranges:
            # 检查是否合并了第1行（禁止）
            if r.min_row == 1 and r.max_row == 1:
                result['passed'] = False
                result['errors'].append(f'禁止合并标题行: {r}（第1行必须是表头，不允许合并单元格）')
            else:
                result['warnings'].append(f'存在合并单元格: {r}')

    # 3. 表头校验
    actual_headers = []
    for c in range(1, min(actual_cols + 1, expected_cols + 1)):
        v = ws.cell(header_row, c).value
        actual_headers.append(str(v) if v is not None else '')

    result['stats']['actual_headers'] = actual_headers

    # 3.1 检查简化字段名
    simplified_patterns = [
        (r'^编号$', '编号（应使用完整字段名如"知识产权编号"或"研发活动编号"）'),
        (r'^技术领域$', '技术领域（应分一级/二级/三级 3列）'),
        (r'^起止时间$', '起止时间（应分"开始时间"+"结束时间" 2列）'),
        (r'^摘要', '摘要（应使用完整字段名"国家知识产权局官方网站上公布的摘要(限400字，软件著作权不用提供)"）'),
        (r'^所属单位$', '所属单位（应使用"知识产权所属单位或个人"）'),
        (r'^专利号/软著号', '专利号/软著号（应使用"专利号/著作权号"）'),
    ]
    for h in actual_headers:
        for pattern, msg in simplified_patterns:
            if re.match(pattern, h):
                result['passed'] = False
                result['errors'].append(f'字段名简化: "{h}" → {msg}')
                break

    # 3.2 检查技术来源是否误用"自主研发"（RD/PS表）
    if table_type in ['rd', 'ps']:
        tech_source_col = 8 if table_type == 'rd' else 6
        for r in range(header_row + 1, actual_rows + 1):
            v = ws.cell(r, tech_source_col).value
            if v and '自主研发' in str(v):
                result['passed'] = False
                result['errors'].append(f'第{r}行 技术来源误用"自主研发"，应改为"企业自有技术"')
                break

    # 3.3 检查表头是否完全匹配模板
    if actual_headers and actual_headers[0]:
        # 第一个字段必须是模板要求的（不允许是合并标题）
        if actual_headers[0] != expected_headers[0]:
            # 检查是否是"XX公司 XX表"合并标题
            if '公司' in actual_headers[0] and ('表' in actual_headers[0] or '汇总' in actual_headers[0]):
                result['passed'] = False
                result['errors'].append(
                    f'禁止合并标题行: 第1行第1列"{actual_headers[0]}"（应直接是表头"{expected_headers[0]}"）'
                )

    # 4. 表头字段名完整匹配
    for i, (actual, expected) in enumerate(zip(actual_headers, expected_headers)):
        if actual != expected:
            # 部分匹配（实际是简写）
            if actual and expected and (actual in expected or expected.startswith(actual)):
                result['warnings'].append(f'第{i+1}列字段名简写: 实际"{actual}" → 期望"{expected}"')
            else:
                result['errors'].append(f'第{i+1}列字段名不符: 实际"{actual}" → 期望"{expected}"')
                result['passed'] = False

    # 5. 长文本字段字数校验
    long_text_cols = spec['long_text_columns']
    # 优先使用spec中定义的字数范围，否则用参数传入的默认值
    spec_min = spec.get('min_chars', min_chars)
    spec_max = spec.get('max_chars', max_chars)
    text_validations = []
    for r in range(header_row + 1, actual_rows + 1):
        for col in long_text_cols:
            if col > actual_cols:
                continue
            v = ws.cell(r, col).value
            # v1.18.2修订：IP表第8列（摘要）特殊处理
            # 专利/实用新型/外观设计：严格按专利说明书原文，不校验字数，只校验非空
            # 软件著作权：必须填"无"（禁止空值，禁止填其他内容）
            if table_type == 'ip' and col == 8:
                cat_col_val = ws.cell(r, 3).value  # 类别列
                is_software_copyright = cat_col_val and '软件著作权' in str(cat_col_val)
                if is_software_copyright:
                    # 软著摘要必须填"无"
                    if v is None or (isinstance(v, str) and v.strip() == ''):
                        text_validations.append({
                            'row': r, 'col': col, 'passed': False,
                            'msg': '软件著作权摘要为空（必须填"无"）',
                            'text_preview': '',
                        })
                        result['passed'] = False
                        result['errors'].append(
                            f'第{r}行第{col}列 软件著作权摘要为空，必须填"无"（v1.18.2强制）'
                        )
                        continue
                    if str(v).strip() != '无':
                        result['warnings'].append(
                            f'第{r}行第{col}列 软件著作权摘要应为"无"，实际为"{str(v)[:20]}"'
                        )
                    # 软著摘要填"无"，跳过字数校验
                    text_validations.append({
                        'row': r, 'col': col, 'passed': True,
                        'msg': '软件著作权摘要填"无"（合规）',
                        'text_preview': '无',
                    })
                    continue
                else:
                    # 非软著：严格按专利说明书原文，只校验非空，不校验字数
                    if v is None or (isinstance(v, str) and v.strip() == ''):
                        text_validations.append({
                            'row': r, 'col': col, 'passed': False,
                            'msg': '摘要为空（专利/实用新型/外观设计必须按专利文献原文填写）',
                            'text_preview': '',
                        })
                        result['passed'] = False
                        result['errors'].append(
                            f'第{r}行第{col}列 摘要为空，必须严格按专利文献中的摘要原文填写（v1.18.2强制）'
                        )
                        continue
                    # 非软著摘要有内容，跳过字数校验（字数长短由专利文献决定）
                    text_validations.append({
                        'row': r, 'col': col, 'passed': True,
                        'msg': f'摘要已填写（{len(str(v))}字，不校验字数）',
                        'text_preview': str(v)[:50] + '...' if v and len(str(v)) > 50 else str(v),
                    })
                    continue
            # 其他长文本字段（第9/10列）走标准字数校验
            passed, msg = validate_text_length(v, spec_min, spec_max, allow_empty=False)
            text_validations.append({
                'row': r, 'col': col, 'passed': passed, 'msg': msg,
                'text_preview': str(v)[:50] + '...' if v and len(str(v)) > 50 else str(v),
            })
            if not passed:
                result['passed'] = False
                result['errors'].append(
                    f'第{r}行第{col}列长文本字段{msg}（预览: {str(v)[:30]}...）'
                )

    result['stats']['text_validations'] = text_validations
    result['stats']['long_text_count'] = len(text_validations)
    result['stats']['long_text_pass'] = sum(1 for v in text_validations if v['passed'])
    result['stats']['long_text_fail'] = sum(1 for v in text_validations if not v['passed'])

    # 6. 样式校验（采样）
    style_issues = []
    if actual_rows >= header_row and actual_cols >= 1:
        # 表头样式
        h_cell = ws.cell(header_row, 1)
        font = h_cell.font
        if font.name != '宋体':
            style_issues.append(f'表头字体应为"宋体"，实际"{font.name}"')
        if font.size != 12:
            style_issues.append(f'表头字号应为12，实际{font.size}')
        if not font.bold:
            style_issues.append('表头应为bold')
        # 边框
        b = h_cell.border
        if not (b.top.style == 'thin' and b.bottom.style == 'thin'
                and b.left.style == 'thin' and b.right.style == 'thin'):
            style_issues.append('表头缺少thin边框')
        # 数据行样式（采样第1行数据）
        if actual_rows > header_row:
            d_cell = ws.cell(header_row + 1, 1)
            df = d_cell.font
            if df.name != '宋体':
                style_issues.append(f'数据字体应为"宋体"，实际"{df.name}"')
            if df.size and df.size > 12:
                style_issues.append(f'数据字号应≤11，实际{df.size}')
            db = d_cell.border
            if not (db.top.style == 'thin' and db.bottom.style == 'thin'
                    and db.left.style == 'thin' and db.right.style == 'thin'):
                style_issues.append('数据行缺少thin边框')

    for issue in style_issues:
        result['warnings'].append(f'样式: {issue}')

    # 7. v1.16.0新增：科技人员清单（staff）统计行校验
    if table_type == 'staff' and spec.get('require_summary_rows'):
        summary_labels = spec.get('summary_labels', [])
        # 在数据行之后查找统计行（合并A:B列的标签）
        found_summaries = []
        for r in range(header_row + 2, actual_rows + 1):  # 跳过表头和首行数据
            v = ws.cell(r, 1).value
            if v and isinstance(v, str) and v in summary_labels:
                found_summaries.append(v)
        missing = [s for s in summary_labels if s not in found_summaries]
        if missing:
            result['passed'] = False
            result['errors'].append(f'科技人员清单缺少统计行: {missing}（应在数据行末尾追加）')

        # 校验"科技人员占比（基于社保）"行是否有数值
        for r in range(header_row + 2, actual_rows + 1):
            v = ws.cell(r, 1).value
            if v == '科技人员占比（基于社保）':
                ratio_val = ws.cell(r, 3).value
                if not ratio_val:
                    result['passed'] = False
                    result['errors'].append(f'第{r}行 "科技人员占比（基于社保）" 缺少数值')
                elif isinstance(ratio_val, str) and '%' not in ratio_val and '社保清单为空' not in ratio_val:
                    result['warnings'].append(f'第{r}行 占比值可能格式不正确: {ratio_val}（应为 "X.XX%" 格式）')
                result['stats']['staff_ratio'] = str(ratio_val)
                break

        # 校验"上年在职天数"是否≥183天（仅对符合条件的人员）
        days_col = 9  # 上年在职天数列
        staff_qualified_col = 13  # 是否符合科技人员条件列
        insufficient_days = []
        for r in range(header_row + 1, actual_rows + 1):
            label = ws.cell(r, 1).value
            # 跳过统计行
            if label and isinstance(label, str) and label in summary_labels:
                continue
            days_val = ws.cell(r, days_col).value
            qual_val = ws.cell(r, staff_qualified_col).value
            if qual_val and str(qual_val).strip() in ('是', '√', 'Y', 'y'):
                # 符合条件的人员，必须≥183天
                try:
                    days = int(days_val) if days_val is not None else 0
                    if days < 183:
                        insufficient_days.append((r, days))
                except (ValueError, TypeError):
                    result['warnings'].append(f'第{r}行 "上年在职天数" 非数字: {days_val}')
        if insufficient_days:
            result['passed'] = False
            for r, d in insufficient_days:
                result['errors'].append(f'第{r}行 符合科技人员条件但上年在职天数{d}天<183天（不符合认定要求）')

    return result


def validate_all_tables(table_dir, min_chars=300, max_chars=400):
    """校验目录下所有核心表格"""
    results = []
    files = os.listdir(table_dir) if os.path.isdir(table_dir) else []

    # 文件名匹配规则
    match_rules = [
        ('ip', [r'知识产权表\.xlsx$', r'知识产权.*\.xlsx$']),
        ('rd', [r'研发活动.*汇总.*\.xlsx$', r'企业研究开发活动.*\.xlsx$']),
        ('ps', [r'高新技术产品.*明细.*\.xlsx$', r'高新产品.*明细.*\.xlsx$']),
        ('ach', [r'科技成果转化.*汇总.*\.xlsx$']),
        ('rdps', [r'研发项目RD.*PS.*\.xlsx$', r'RD.*PS.*汇总.*\.xlsx$']),
        # v1.16.0新增：科技人员清单（三方对比）匹配规则
        ('staff', [r'科技人员清单.*三方对比.*\.xlsx$', r'科技人员清单.*\.xlsx$']),
    ]

    matched = {}
    for f in files:
        if not f.endswith('.xlsx'):
            continue
        for ttype, patterns in match_rules:
            for p in patterns:
                if re.search(p, f):
                    if ttype not in matched:
                        matched[ttype] = f
                    break

    print(f'匹配到的文件:')
    for ttype, fname in matched.items():
        print(f'  {ttype}: {fname}')

    for ttype, fname in matched.items():
        fpath = os.path.join(table_dir, fname)
        result = validate_table_format(fpath, ttype, min_chars, max_chars)
        results.append(result)

    return results


def print_report(results):
    """打印校验报告"""
    print()
    print('=' * 80)
    print('核心表格模板对齐校验报告')
    print('=' * 80)
    print()

    all_passed = True
    for r in results:
        status = '✓ PASS' if r['passed'] else '✗ FAIL'
        print(f'[{status}] {r["table_name"]} ({r["table_type"]})')
        print(f'  文件: {r["file"]}')
        print(f'  规模: {r["stats"].get("rows", 0)}行 × {r["stats"].get("cols", 0)}列')
        if r['errors']:
            for e in r['errors']:
                print(f'  ERROR: {e}')
            all_passed = False
        if r['warnings']:
            for w in r['warnings'][:5]:  # 只显示前5条警告
                print(f'  WARN: {w}')
            if len(r['warnings']) > 5:
                print(f'  ... ({len(r["warnings"]) - 5} 条更多警告)')
        # 字数校验统计
        if r['stats'].get('long_text_count', 0) > 0:
            total = r['stats']['long_text_count']
            ok = r['stats']['long_text_pass']
            fail = r['stats']['long_text_fail']
            print(f'  字数校验: {ok}/{total} 通过, {fail} 失败')
        print()

    print('=' * 80)
    if all_passed:
        print('最终结论: ✓ 全部通过')
    else:
        print('最终结论: ✗ 存在不合规项，需整改')
    print('=' * 80)

    return all_passed


def main():
    parser = argparse.ArgumentParser(description='核心表格模板对齐校验工具 v1.14.0')
    parser.add_argument('--dir', help='核心表格所在目录')
    parser.add_argument('--file', help='单个xlsx文件路径')
    parser.add_argument('--type', choices=['ip', 'rd', 'ps', 'ach', 'rdps'], help='表格类型')
    parser.add_argument('--enterprise', help='企业名称（自动定位目录）')
    parser.add_argument('--year', type=int, default=2026, help='申报年份')
    parser.add_argument('--min-chars', type=int, default=300, help='长文本最小字数（默认300）')
    parser.add_argument('--max-chars', type=int, default=400, help='长文本最大字数（默认400）')
    parser.add_argument('--json', action='store_true', help='输出JSON格式报告')
    args = parser.parse_args()

    if args.dir:
        results = validate_all_tables(args.dir, args.min_chars, args.max_chars)
    elif args.file and args.type:
        result = validate_table_format(args.file, args.type, args.min_chars, args.max_chars)
        results = [result]
    elif args.enterprise:
        # 自动定位目录（v1.14.1 改造：从 path_config 获取企业数据根目录，移除硬编码 D:\OneDrive 个人路径）
        enterprise_data_dir = get_enterprise_data_dir()
        if enterprise_data_dir is None:
            print('[ERROR] 未设置 GXTZ_ENTERPRISE_DATA_DIR 环境变量，无法自动定位企业核心表格目录', file=sys.stderr)
            print('        请使用以下任一方式：', file=sys.stderr)
            print('        1. 设置环境变量 GXTZ_ENTERPRISE_DATA_DIR 指向企业数据根目录', file=sys.stderr)
            print('        2. 使用 --dir 参数显式指定 00_核心表格 目录路径', file=sys.stderr)
            sys.exit(1)
        candidates = [
            str(Path(enterprise_data_dir) / f"【国高】*{args.enterprise}*" / f"*_高新认定材料_{args.year}" / "00_核心表格"),
        ]
        found = None
        for pattern in candidates:
            matches = glob.glob(pattern)
            if matches:
                found = matches[0]
                break
        if not found:
            print(f'未找到企业 {args.enterprise} 的核心表格目录（搜索根: {enterprise_data_dir}）')
            sys.exit(1)
        print(f'定位目录: {found}')
        results = validate_all_tables(found, args.min_chars, args.max_chars)
    else:
        parser.print_help()
        sys.exit(1)

    if args.json:
        print(json.dumps(results, ensure_ascii=False, indent=2, default=str))
    else:
        all_passed = print_report(results)
        sys.exit(0 if all_passed else 1)


if __name__ == '__main__':
    main()
