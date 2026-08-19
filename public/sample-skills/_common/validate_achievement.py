"""科技成果转化材料审核脚本

用途：审核高新技术企业认定申报中的科技成果转化证明材料是否合规，包括：
  1. 成果转化表关联完整性：RD→IP→PS 链条是否完整
  2. 每 RD 至少 1 项成果转化
  3. 近三年每年至少 1 项成果转化
  4. 转化形式、转化时间完整性
  5. 成果类型分类校验
  6. 成果转化证明材料完整性（合同/发票/检测报告等）

用法：
  python validate_achievement.py --dir "成果转化材料目录"
  python validate_achievement.py --dir "成果转化材料目录" --project-root "项目根目录"
  python validate_achievement.py --file "科技成果转化情况汇总表.xlsx" --project-root "项目根目录"

输出：JSON 格式审核报告，结构如下：
  {
    "passed": bool,
    "errors": [{"file": "...", "row": 0, "field": "...", "reason": "..."}],
    "warnings": [...],
    "stats": {"total_achievement": N, "valid_achievement": N, ...}
  }

退出码：审核通过 0，存在错误 1

依赖：openpyxl（用于读取 xlsx 汇总表）；若无 openpyxl 则跳过表格校验
"""

import os
import re
import sys
import json
import argparse
from datetime import datetime


# ============================================================
# 常量定义
# ============================================================

# 证明材料文件支持的扩展名
PROOF_EXTENSIONS = ('.pdf', '.jpg', '.jpeg', '.png', '.doc', '.docx', '.xls', '.xlsx')

# 成果转化表模板表头（与 validate_tables.py 中 ach 规格一致）
ACHIEVEMENT_HEADERS = [
    '科技成果序号', '科技成果名称', '成果类型', '成果来源',
    '转化结果', '转化时间', '关联IP', '关联RD', '关联PS',
    '转化形式', '涉及关键技术（限400字）', '成效（限400字）',
]

# 成果类型有效值
VALID_ACHIEVEMENT_TYPES = {
    '自主研发', '受让', '受赠', '并购', '其他',
}

# 转化形式有效值
VALID_TRANSFORM_FORMS = {
    '自行投资实施转化', '向他人转让该技术', '许可他人使用该技术',
    '以该技术作为合作条件', '以该技术作价投资', '其他',
}

# RD/IP/PS 编号正则
RD_NUMBER_PATTERN = re.compile(r'RD(\d{1,3})', re.IGNORECASE)
IP_NUMBER_PATTERN = re.compile(r'IP(\d{1,3})', re.IGNORECASE)
PS_NUMBER_PATTERN = re.compile(r'PS(\d{1,3})', re.IGNORECASE)


def _load_project_index(project_root):
    """加载 project_index.json，用于 RD→IP→PS 链条校验"""
    if not project_root:
        return None
    candidates = [
        os.path.join(project_root, '.trae', 'project_knowledge', 'project_index.json'),
        os.path.join(project_root, '.trae', 'project_index.json'),
    ]
    for path in candidates:
        if os.path.exists(path):
            try:
                with open(path, 'r', encoding='utf-8') as f:
                    return json.load(f)
            except (json.JSONDecodeError, OSError):
                return None
    return None


def _get_recent_three_years(project_index, default=None):
    """从 project_index.json 获取近三年年份列表"""
    if default is None:
        default = [2023, 2024, 2025]
    if project_index and project_index.get('recent_three_years'):
        years = project_index['recent_three_years']
        if isinstance(years, list) and len(years) == 3:
            return [int(y) for y in years]
    return default


def _extract_year(value):
    """从字符串中提取 4 位年份"""
    if value is None:
        return None
    m = re.search(r'(19|20)\d{2}', str(value))
    return int(m.group(0)) if m else None


def _extract_numbers(value, pattern):
    """从单元格值中提取所有匹配的编号（如 RD01, RD02 → [1, 2]）"""
    if value is None:
        return []
    return [int(m.group(1)) for m in pattern.finditer(str(value))]


def _scan_proof_files(proof_dir):
    """扫描成果转化证明材料目录下的文件

    返回：{序号: [文件名列表], ...}（按成果序号分组）
    """
    proofs = {}
    if not os.path.isdir(proof_dir):
        return proofs
    # 成果序号正则：1_、01_、1、01（文件名开头）
    seq_pattern = re.compile(r'^(\d{1,3})[_\-.\s]')
    for entry in sorted(os.listdir(proof_dir)):
        ext = os.path.splitext(entry)[1].lower()
        if ext not in PROOF_EXTENSIONS:
            continue
        m = seq_pattern.match(entry)
        if m:
            seq = int(m.group(1))
            proofs.setdefault(seq, []).append(entry)
    return proofs


def _read_achievement_table(file_path):
    """读取科技成果转化情况汇总表

    返回：[{'row': 行号, 'seq': 序号, 'name': 名称, 'type': 类型, 'source': 来源,
           'result': 转化结果, 'transform_time': 转化时间, 'ip': [IP编号],
           'rd': [RD编号], 'ps': [PS编号], 'form': 转化形式, ...}, ...]
    """
    rows = []
    try:
        import openpyxl
    except ImportError:
        return None, 'openpyxl 未安装，跳过表格读取'

    if not os.path.exists(file_path):
        return None, f'文件不存在: {file_path}'

    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
    except Exception as e:
        return None, f'加载 Excel 失败: {e}'

    ws = wb.active
    header_row = 1
    # 自动检测表头行（查找包含"科技成果序号"的行）
    for r in range(1, min(5, ws.max_row + 1)):
        v = ws.cell(r, 1).value
        if v and '序号' in str(v):
            header_row = r
            break

    # 读取数据行
    for r in range(header_row + 1, ws.max_row + 1):
        seq_val = ws.cell(r, 1).value
        # 跳过空行和统计行
        if seq_val is None or str(seq_val).strip() == '':
            continue
        if isinstance(seq_val, str) and ('总计' in seq_val or '合计' in seq_val or '说明' in seq_val):
            continue

        row_data = {
            'row': r,
            'seq': seq_val,
            'name': ws.cell(r, 2).value,
            'type': ws.cell(r, 3).value,
            'source': ws.cell(r, 4).value,
            'result': ws.cell(r, 5).value,
            'transform_time': ws.cell(r, 6).value,
            'ip_raw': ws.cell(r, 7).value,
            'rd_raw': ws.cell(r, 8).value,
            'ps_raw': ws.cell(r, 9).value,
            'form': ws.cell(r, 10).value,
            'tech_key': ws.cell(r, 11).value,
            'effect': ws.cell(r, 12).value,
            'ip': _extract_numbers(ws.cell(r, 7).value, IP_NUMBER_PATTERN),
            'rd': _extract_numbers(ws.cell(r, 8).value, RD_NUMBER_PATTERN),
            'ps': _extract_numbers(ws.cell(r, 9).value, PS_NUMBER_PATTERN),
        }
        rows.append(row_data)

    return rows, None


def validate_achievement_directory(ach_dir, project_root=None):
    """审核成果转化材料目录

    参数：
      ach_dir: 成果转化材料目录路径（含汇总表 + 证明材料）
      project_root: 项目根目录（用于读取 project_index.json）

    返回：审核报告 dict
    """
    errors = []
    warnings = []
    stats = {
        'total_achievement': 0,
        'valid_achievement': 0,
        'invalid_achievement': 0,
        'with_full_chain': 0,        # RD→IP→PS 链条完整
        'with_partial_chain': 0,     # 链条部分缺失
        'rd_without_achievement': 0, # 无成果转化的 RD 数
        'proof_files_count': 0,
        'by_year': {},               # 按转化年份统计
    }

    # ============================================================
    # 校验 1：目录结构校验
    # ============================================================
    if not os.path.exists(ach_dir):
        errors.append({
            'file': ach_dir, 'row': 0, 'field': 'directory',
            'reason': f'成果转化材料目录不存在: {ach_dir}',
        })
        return {'passed': False, 'errors': errors, 'warnings': warnings, 'stats': stats}

    # 查找成果转化汇总表
    table_file = None
    table_patterns = [r'科技成果转化.*汇总.*\.xlsx$', r'科技成果转化.*\.xlsx$']
    for entry in os.listdir(ach_dir):
        if not entry.endswith('.xlsx'):
            continue
        for p in table_patterns:
            if re.search(p, entry):
                table_file = os.path.join(ach_dir, entry)
                break
        if table_file:
            break

    # 加载 project_index.json
    project_index = _load_project_index(project_root)
    recent_years = _get_recent_three_years(project_index)

    # 从 project_index 获取 RD 列表
    rd_ids_in_index = set()
    if project_index and isinstance(project_index.get('knowledge_graph'), dict):
        for node in project_index['knowledge_graph'].get('nodes', []) or []:
            if not isinstance(node, dict):
                continue
            node_id = str(node.get('id', ''))
            m = RD_NUMBER_PATTERN.search(node_id)
            if m:
                rd_ids_in_index.add(int(m.group(1)))

    # ============================================================
    # 校验汇总表
    # ============================================================
    table_rows = []
    if table_file is None:
        warnings.append({
            'file': '', 'row': 0, 'field': 'table',
            'reason': '未找到科技成果转化情况汇总表（xlsx），跳过表格关联性校验',
        })
    else:
        rows, err = _read_achievement_table(table_file)
        if err:
            warnings.append({
                'file': os.path.basename(table_file), 'row': 0, 'field': 'table',
                'reason': err,
            })
        else:
            table_rows = rows or []
            stats['total_achievement'] = len(table_rows)

    # ============================================================
    # 逐行校验成果转化记录
    # ============================================================
    rd_with_achievement = set()
    years_with_achievement = set()

    for row_data in table_rows:
        r = row_data['row']
        fname = os.path.basename(table_file) if table_file else ''
        has_error = False

        # 校验 1：RD→IP→PS 链条完整性
        has_rd = len(row_data['rd']) > 0
        has_ip = len(row_data['ip']) > 0
        has_ps = len(row_data['ps']) > 0

        if has_rd and has_ip and has_ps:
            stats['with_full_chain'] += 1
        elif has_rd or has_ip or has_ps:
            stats['with_partial_chain'] += 1
            missing = []
            if not has_rd:
                missing.append('RD')
            if not has_ip:
                missing.append('IP')
            if not has_ps:
                missing.append('PS')
            warnings.append({
                'file': fname, 'row': r, 'field': 'chain',
                'reason': f'第{r}行 成果转化链条不完整，缺少: {",".join(missing)}',
            })
        else:
            stats['with_partial_chain'] += 1
            errors.append({
                'file': fname, 'row': r, 'field': 'chain',
                'reason': f'第{r}行 RD→IP→PS 链条全部缺失',
            })
            has_error = True

        # 记录有关联成果转化的 RD
        for rd_no in row_data['rd']:
            rd_with_achievement.add(rd_no)

        # 校验 4：转化形式、转化时间完整性
        if not row_data['form'] or str(row_data['form']).strip() == '':
            errors.append({
                'file': fname, 'row': r, 'field': 'form',
                'reason': f'第{r}行 转化形式为空',
            })
            has_error = True
        else:
            # 转化形式分类校验
            form_str = str(row_data['form']).strip()
            if form_str not in VALID_TRANSFORM_FORMS and not any(
                v in form_str for v in VALID_TRANSFORM_FORMS
            ):
                warnings.append({
                    'file': fname, 'row': r, 'field': 'form',
                    'reason': f'第{r}行 转化形式"{form_str}"不在标准分类中',
                })

        if not row_data['transform_time'] or str(row_data['transform_time']).strip() == '':
            errors.append({
                'file': fname, 'row': r, 'field': 'transform_time',
                'reason': f'第{r}行 转化时间为空',
            })
            has_error = True
        else:
            # 提取转化年份
            year = _extract_year(row_data['transform_time'])
            if year:
                years_with_achievement.add(year)
                stats['by_year'][year] = stats['by_year'].get(year, 0) + 1
                # 校验转化时间是否在近三年内
                if year not in recent_years:
                    warnings.append({
                        'file': fname, 'row': r, 'field': 'transform_time',
                        'reason': f'第{r}行 转化时间年份 {year} 不在近三年 {recent_years} 内',
                    })

        # 校验 5：成果类型分类校验
        if not row_data['type'] or str(row_data['type']).strip() == '':
            errors.append({
                'file': fname, 'row': r, 'field': 'type',
                'reason': f'第{r}行 成果类型为空',
            })
            has_error = True
        else:
            type_str = str(row_data['type']).strip()
            if type_str not in VALID_ACHIEVEMENT_TYPES and not any(
                v in type_str for v in VALID_ACHIEVEMENT_TYPES
            ):
                warnings.append({
                    'file': fname, 'row': r, 'field': 'type',
                    'reason': f'第{r}行 成果类型"{type_str}"不在标准分类中',
                })

        # 转化结果完整性
        if not row_data['result'] or str(row_data['result']).strip() == '':
            warnings.append({
                'file': fname, 'row': r, 'field': 'result',
                'reason': f'第{r}行 转化结果为空',
            })

        if has_error:
            stats['invalid_achievement'] += 1
        else:
            stats['valid_achievement'] += 1

    # ============================================================
    # 校验 2：每 RD 至少 1 项成果转化
    # ============================================================
    if rd_ids_in_index:
        rd_without = rd_ids_in_index - rd_with_achievement
        stats['rd_without_achievement'] = len(rd_without)
        if rd_without:
            missing_list = sorted(rd_without)
            errors.append({
                'file': '', 'row': 0, 'field': 'rd_coverage',
                'reason': f'共 {len(rd_without)} 个 RD 无成果转化: RD{missing_list[:10]}'
                          + ('...' if len(missing_list) > 10 else ''),
            })
    elif table_rows:
        # 无 project_index 时，从表格内部统计
        stats['rd_without_achievement'] = 0

    # ============================================================
    # 校验 3：近三年每年至少 1 项成果转化
    # ============================================================
    for year in recent_years:
        if year not in years_with_achievement:
            errors.append({
                'file': '', 'row': 0, 'field': 'yearly_coverage',
                'reason': f'近三年 {year} 年无成果转化记录',
            })

    # ============================================================
    # 校验 6：成果转化证明材料完整性
    # ============================================================
    proofs = _scan_proof_files(ach_dir)
    stats['proof_files_count'] = sum(len(v) for v in proofs.values())

    if table_rows:
        for row_data in table_rows:
            seq = row_data['seq']
            # 尝试将序号转为整数
            try:
                seq_int = int(seq)
            except (ValueError, TypeError):
                continue
            if seq_int not in proofs:
                warnings.append({
                    'file': '', 'row': row_data['row'], 'field': 'proof_material',
                    'reason': f'第{row_data["row"]}行 成果序号 {seq} 未找到对应证明材料文件',
                })

    passed = len(errors) == 0
    return {
        'passed': passed,
        'errors': errors,
        'warnings': warnings,
        'stats': stats,
    }


def validate_single_achievement_file(file_path, project_root=None):
    """审核单个成果转化汇总表文件"""
    errors = []
    warnings = []
    stats = {
        'total_achievement': 0,
        'valid_achievement': 0,
        'invalid_achievement': 0,
        'with_full_chain': 0,
        'with_partial_chain': 0,
    }

    if not os.path.exists(file_path):
        errors.append({
            'file': file_path, 'row': 0, 'field': 'file',
            'reason': f'文件不存在: {file_path}',
        })
        return {'passed': False, 'errors': errors, 'warnings': warnings, 'stats': stats}

    rows, err = _read_achievement_table(file_path)
    if err:
        errors.append({
            'file': os.path.basename(file_path), 'row': 0, 'field': 'table',
            'reason': err,
        })
        return {'passed': False, 'errors': errors, 'warnings': warnings, 'stats': stats}

    rows = rows or []
    stats['total_achievement'] = len(rows)
    fname = os.path.basename(file_path)

    project_index = _load_project_index(project_root)
    recent_years = _get_recent_three_years(project_index)
    years_with_achievement = set()

    for row_data in rows:
        r = row_data['row']
        has_error = False

        # 链条完整性
        has_rd = len(row_data['rd']) > 0
        has_ip = len(row_data['ip']) > 0
        has_ps = len(row_data['ps']) > 0
        if has_rd and has_ip and has_ps:
            stats['with_full_chain'] += 1
        else:
            stats['with_partial_chain'] += 1
            missing = [x for x, has in [('RD', has_rd), ('IP', has_ip), ('PS', has_ps)] if not has]
            if not has_rd and not has_ip and not has_ps:
                errors.append({
                    'file': fname, 'row': r, 'field': 'chain',
                    'reason': f'第{r}行 RD→IP→PS 链条全部缺失',
                })
                has_error = True
            else:
                warnings.append({
                    'file': fname, 'row': r, 'field': 'chain',
                    'reason': f'第{r}行 链条不完整，缺少: {",".join(missing)}',
                })

        # 转化形式与时间
        if not row_data['form'] or str(row_data['form']).strip() == '':
            errors.append({
                'file': fname, 'row': r, 'field': 'form',
                'reason': f'第{r}行 转化形式为空',
            })
            has_error = True

        if not row_data['transform_time'] or str(row_data['transform_time']).strip() == '':
            errors.append({
                'file': fname, 'row': r, 'field': 'transform_time',
                'reason': f'第{r}行 转化时间为空',
            })
            has_error = True
        else:
            year = _extract_year(row_data['transform_time'])
            if year:
                years_with_achievement.add(year)

        # 成果类型
        if not row_data['type'] or str(row_data['type']).strip() == '':
            errors.append({
                'file': fname, 'row': r, 'field': 'type',
                'reason': f'第{r}行 成果类型为空',
            })
            has_error = True

        if has_error:
            stats['invalid_achievement'] += 1
        else:
            stats['valid_achievement'] += 1

    # 近三年覆盖校验
    for year in recent_years:
        if year not in years_with_achievement:
            errors.append({
                'file': fname, 'row': 0, 'field': 'yearly_coverage',
                'reason': f'近三年 {year} 年无成果转化记录',
            })

    passed = len(errors) == 0
    return {
        'passed': passed,
        'errors': errors,
        'warnings': warnings,
        'stats': stats,
    }


def main():
    """CLI 入口：解析参数并执行成果转化材料审核"""
    parser = argparse.ArgumentParser(
        description='科技成果转化材料审核脚本 - 校验高企认定成果转化材料合规性',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog='''示例：
  python validate_achievement.py --dir "科技成果转化证明资料"
  python validate_achievement.py --dir "科技成果转化证明资料" --project-root "项目根目录"
  python validate_achievement.py --file "科技成果转化情况汇总表.xlsx" --project-root "项目根目录"
''',
    )
    parser.add_argument('--dir', help='成果转化材料所在目录')
    parser.add_argument('--file', help='单个成果转化汇总表 xlsx 文件路径')
    parser.add_argument('--project-root', dest='project_root',
                        help='项目根目录（用于读取 project_index.json 进行关联校验）')
    args = parser.parse_args()

    if not args.dir and not args.file:
        parser.print_help()
        sys.exit(1)

    if args.dir:
        report = validate_achievement_directory(args.dir, args.project_root)
    else:
        report = validate_single_achievement_file(args.file, args.project_root)

    # 输出 JSON 格式审核报告
    print(json.dumps(report, ensure_ascii=False, indent=2, default=str))
    sys.exit(0 if report['passed'] else 1)


if __name__ == '__main__':
    main()
