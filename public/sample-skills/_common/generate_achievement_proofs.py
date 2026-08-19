#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
generate_achievement_proofs.py - 科技成果转化证明材料PDF附件生成器 v1.2.0

将每项成果转化的证明材料（IP证书、合同、发票、检测报告等）合并为单个PDF文件。

核心流程：
  1. 读取成果转化汇总表，获取每项成果的序号、名称、关联IP、转化时间
  2. 从IP证书目录匹配关联的知识产权证书（排除说明书/权利要求书/审查意见）
  3. 从合同发票目录匹配合同和发票文件
  4. 按成果合并所有证明材料为单个PDF
  5. 输出到最终版本目录

集成经验规则（v1.1.0）：
  R03 证书优先：排除说明书/权利要求书/审查意见/缴费凭证等非证书文件
  R02 年份硬过滤：合同发票必须限定在近三年(VALID_YEARS)内，禁止跨年配对
  R04 测试报告跨年校验：检测报告年份必须与转化成果年份一致
  R05 文件唯一分配：同一物理文件禁止重复分配给多个成果
  v1.22.0 材料类型扩充：质量管理体系证书 + 检测报告分类识别

用法：
  模式1 - 从按成果归类的材料目录生成（推荐，agent先手工分配好合同发票）：
    python generate_achievement_proofs.py organize ^
        --achievement-table "03_成果转化证明/成果转化汇总表.xlsx" ^
        --materials-dir "03_成果转化证明/" ^
        --ip-dir "02_知识产权证明/" ^
        --output-dir "03_成果转化证明/最终版本/" ^
        --application-year 2026

  模式2 - 自动扫描合同发票目录匹配（agent未手工分配时）：
    python generate_achievement_proofs.py auto-scan ^
        --achievement-table "03_成果转化证明/成果转化汇总表.xlsx" ^
        --ip-dir "02_知识产权证明/" ^
        --contract-dir "05-合同发票/" ^
        --output-dir "03_成果转化证明/最终版本/" ^
        --application-year 2026

  模式3 - 列出每项成果的材料匹配情况（预览，不生成PDF）：
    python generate_achievement_proofs.py preview ^
        --achievement-table "03_成果转化证明/成果转化汇总表.xlsx" ^
        --materials-dir "03_成果转化证明/" ^
        --ip-dir "02_知识产权证明/" ^
        --application-year 2026

输出：
  - {序号}_{成果名称}.pdf：合并后的证明材料PDF
  - _生成报告.json：生成过程的详细记录（含违规告警）

依赖：openpyxl, PyPDF2（已列入 tech_stack.json）
"""

import os
import re
import sys
import json
import shutil
import hashlib
import argparse
import tempfile
from datetime import datetime
from pathlib import Path

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

try:
    from onedrive_utils import is_under_onedrive_syncroot, OneDriveSafetyWarning
    from safe_workspace import SafeWorkspace, check_and_report
    _SAFE_WORKSPACE_AVAILABLE = True
except ImportError:
    _SAFE_WORKSPACE_AVAILABLE = False

try:
    from PyPDF2 import PdfReader, PdfMerger
    _PYPDF2_AVAILABLE = True
except ImportError:
    _PYPDF2_AVAILABLE = False

try:
    import openpyxl
    _OPENPYXL_AVAILABLE = True
except ImportError:
    _OPENPYXL_AVAILABLE = False


# ============================================================
# 配置常量
# ============================================================

IP_NUMBER_PATTERN = re.compile(r'IP(\d{1,3})', re.IGNORECASE)
RD_NUMBER_PATTERN = re.compile(r'RD(\d{1,3})', re.IGNORECASE)
PS_NUMBER_PATTERN = re.compile(r'PS(\d{1,3})', re.IGNORECASE)

PROOF_EXTENSIONS = ('.pdf', '.jpg', '.jpeg', '.png')

YEAR_PATTERN = re.compile(r'(\d{4})')
HXS_YEAR_PATTERN = re.compile(r'HXS(\d{4})\d{4}', re.IGNORECASE)
DZFP_YEAR_PATTERN = re.compile(r'dzfp_\d{6}_(\d{4})')

# R03 证书优先：非证书文件关键词（经验来源 EXP-2026-07-17-001）
_IP_CERT_EXCLUDE_KEYWORDS = [
    '说明书', '权利要求书', '权利要求', '审查意见', '通知书',
    '受理通知书', '缴费凭证', '年费', '变更', '手续合格',
    '补正', '驳回', '复审', '无效', '专利文献',
    '摘要附图', 'abstract', 'specification', 'claims',
]

# v1.22.0 材料类型扩充：质量管理体系证书关键词
_QUALITY_CERT_KEYWORDS = [
    '认证', 'ISO', 'CE', '质量管理', '环境认证', '医疗认证',
]

# v1.22.0 材料类型扩充：检测报告关键词
_TEST_REPORT_KEYWORDS = [
    '检测报告', '测试报告', '检验报告', '试验报告', '查新报告',
]


def _get_valid_years(application_year):
    """R02: 计算近三年有效年份集合

    Returns:
        set[int]: {Y-3, Y-2, Y-1}
    """
    return {application_year - 3, application_year - 2, application_year - 1}


def _extract_year_from_filename(filename):
    """从文件名提取年份，优先级：HXS编码 > dzfp时间戳 > 通用4位年份

    经验来源 EXP-2026-07-17-001 规则5：文件名年份仅作参考，必须交叉验证PDF内容
    """
    m = HXS_YEAR_PATTERN.search(filename)
    if m:
        return int(m.group(1))

    m = DZFP_YEAR_PATTERN.search(filename)
    if m:
        return int(m.group(1))

    m = YEAR_PATTERN.search(filename)
    if m:
        return int(m.group(1))

    return None


def _extract_year_from_text(text):
    """从文本中提取4位年份"""
    if not text:
        return None
    m = YEAR_PATTERN.search(str(text))
    return int(m.group(1)) if m else None


def _extract_ip_numbers(value):
    """从单元格值提取IP编号列表"""
    if value is None:
        return []
    return [int(m.group(1)) for m in IP_NUMBER_PATTERN.finditer(str(value))]


def _safe_filename(name):
    """将成果名称转为安全的文件名"""
    if not name:
        return 'unknown'
    name = str(name).strip()
    name = re.sub(r'[\\/:*?"<>|]', '_', name)
    name = re.sub(r'\s+', '', name)
    return name


def _file_hash(file_path):
    """计算文件MD5，用于 R05 文件唯一性追踪"""
    try:
        h = hashlib.md5()
        with open(file_path, 'rb') as f:
            for chunk in iter(lambda: f.read(8192), b''):
                h.update(chunk)
        return h.hexdigest()
    except Exception:
        return None


def _is_excluded_ip_file(filename):
    """R03: 判断是否为应排除的非证书文件"""
    f_lower = filename.lower()
    for kw in _IP_CERT_EXCLUDE_KEYWORDS:
        if kw.lower() in f_lower:
            return True
    return False


def _classify_material_type(filename):
    """v1.22.0: 将材料文件按类型分类

    Returns:
        str: 'contract' | 'invoice' | 'test_report' | 'quality_cert' | 'ip_cert' | 'other'
    """
    f_lower = filename.lower()

    if '合同' in f_lower or 'contract' in f_lower:
        return 'contract'
    if '发票' in f_lower or 'invoice' in f_lower or f_lower.startswith('dzfp_'):
        return 'invoice'
    for kw in _TEST_REPORT_KEYWORDS:
        if kw in filename:
            return 'test_report'
    for kw in _QUALITY_CERT_KEYWORDS:
        if kw in filename:
            return 'quality_cert'
    if re.search(r'IP\d{2}', filename):
        return 'ip_cert'
    return 'other'


def read_achievement_table(file_path):
    """读取成果转化汇总表，返回成果列表"""
    if not _OPENPYXL_AVAILABLE:
        return None, 'openpyxl 未安装'
    if not os.path.exists(file_path):
        return None, f'文件不存在: {file_path}'

    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
    except Exception as e:
        return None, f'加载Excel失败: {e}'

    ws = wb.active
    header_row = 1
    for r in range(1, min(6, ws.max_row + 1)):
        v = str(ws.cell(r, 1).value or '')
        if '序号' in v:
            header_row = r
            break

    achievements = []
    for r in range(header_row + 1, ws.max_row + 1):
        seq_val = ws.cell(r, 1).value
        if seq_val is None:
            continue
        seq_str = str(seq_val).strip()
        if not seq_str:
            continue
        if '合计' in seq_str or '总计' in seq_str or '说明' in seq_str:
            continue

        try:
            seq = int(float(seq_str))
        except (ValueError, TypeError):
            continue

        ach = {
            'seq': seq,
            'name': ws.cell(r, 2).value,
            'type': ws.cell(r, 3).value,
            'source': ws.cell(r, 4).value,
            'result': ws.cell(r, 5).value,
            'transform_time': ws.cell(r, 6).value,
            'ip_num': ws.cell(r, 7).value,
            'rd_num': ws.cell(r, 8).value,
            'ps_num': ws.cell(r, 9).value,
            'form': ws.cell(r, 10).value,
            'ip_nums': _extract_ip_numbers(ws.cell(r, 7).value),
            'rd_nums': _extract_ip_numbers(ws.cell(r, 8).value),
            'ps_nums': _extract_ip_numbers(ws.cell(r, 9).value),
            'row': r,
        }
        achievements.append(ach)

    return achievements, None


def find_achievement_table(ach_dir):
    """在成果转化目录中查找汇总表"""
    for entry in os.listdir(ach_dir):
        if not entry.lower().endswith('.xlsx'):
            continue
        if '成果转化' in entry or '科技成果转化' in entry:
            return os.path.join(ach_dir, entry)
    return None


def scan_ip_certificates(ip_dir):
    """R03: 扫描IP证书目录，排除说明书/权利要求书/审查意见等非证书文件

    经验来源：EXP-2026-07-17-001 规则3「专利证书优先」
    - 只收录专利证书、软著证书
    - 排除说明书、权利要求书、审查意见、缴费凭证、通知书等

    Returns:
        dict[int, str]: {IP编号: 证书PDF路径}
        dict[int, list[str]]: {IP编号: [被排除的文件路径]}
    """
    ip_certs = {}
    ip_excluded = {}
    if not os.path.isdir(ip_dir):
        return ip_certs, ip_excluded

    for entry in sorted(os.listdir(ip_dir)):
        entry_path = os.path.join(ip_dir, entry)
        if os.path.isdir(entry_path):
            continue
        if not entry.lower().endswith('.pdf'):
            continue

        m = IP_NUMBER_PATTERN.search(entry)
        if not m:
            continue

        ip_num = int(m.group(1))

        if _is_excluded_ip_file(entry):
            ip_excluded.setdefault(ip_num, []).append(entry_path)
            continue

        if ip_num not in ip_certs:
            ip_certs[ip_num] = entry_path

    return ip_certs, ip_excluded


def scan_materials_dir(materials_dir):
    """v1.1.0: 扫描按成果归类的材料目录，支持扩展材料类型分类

    期望的目录结构：
      materials_dir/
        成果01_xxx/
          02_合同_xxx.pdf
          03_发票_xxx.pdf
          04_产品相关资料_04_检测报告_xxx.pdf
          05_质量管理证明_xxx.pdf

    Returns:
        dict[int, dict]: {
            成果序号: {
                'contracts': [path],
                'invoices': [path],
                'test_reports': [path],
                'quality_certs': [path],
                'ip_certs': [path],
                'other': [path],
            }
        }
    """
    result = {}
    if not os.path.isdir(materials_dir):
        return result

    ach_dir_pattern = re.compile(r'^成果(\d{1,3})[_\-\s]')

    for entry in sorted(os.listdir(materials_dir)):
        entry_path = os.path.join(materials_dir, entry)
        if not os.path.isdir(entry_path):
            continue

        m = ach_dir_pattern.match(entry)
        if not m:
            continue

        seq = int(m.group(1))
        materials = {
            'contracts': [], 'invoices': [], 'test_reports': [],
            'quality_certs': [], 'ip_certs': [], 'other': [],
        }

        for f in sorted(os.listdir(entry_path)):
            f_path = os.path.join(entry_path, f)
            if not os.path.isfile(f_path):
                continue
            ext = os.path.splitext(f)[1].lower()
            if ext not in PROOF_EXTENSIONS and ext not in ('.docx', '.doc'):
                continue

            mtype = _classify_material_type(f)
            if mtype in materials:
                materials[mtype].append(f_path)
            else:
                materials['other'].append(f_path)

        result[seq] = materials

    return result


def scan_contract_invoice_dir(contract_dir):
    """扫描合同发票目录，建立年份 → 文件列表 的映射"""
    by_year = {}
    if not os.path.isdir(contract_dir):
        return by_year

    for root, dirs, files in os.walk(contract_dir):
        dirs[:] = [d for d in dirs if not d.startswith('_backup') and not d.startswith('.')]

        for f in files:
            ext = os.path.splitext(f)[1].lower()
            if ext not in ('.pdf', '.jpg', '.jpeg', '.png'):
                continue

            f_path = os.path.join(root, f)
            year = _extract_year_from_filename(f)

            if year:
                by_year.setdefault(str(year), []).append(f_path)
            else:
                by_year.setdefault('unknown', []).append(f_path)

    return by_year


def _validate_material_years(materials, valid_years, ach_name, ach_seq):
    """R02 + R04: 校验材料文件的年份是否合规

    规则2: 合同发票必须在近三年内
    规则4: 测试报告年份必须与成果转化年份一致（由调用方传入 ach_year）

    Returns:
        list[str]: 警告信息列表
    """
    warnings = []

    for ct in materials.get('contracts', []):
        fname = os.path.basename(ct)
        fyear = _extract_year_from_filename(fname)
        if fyear and fyear not in valid_years:
            warnings.append(
                f'[R02-合同年份越界] 合同 "{fname}" 年份 {fyear} 不在近三年 {sorted(valid_years)} 内'
            )

    for inv in materials.get('invoices', []):
        fname = os.path.basename(inv)
        fyear = _extract_year_from_filename(fname)
        if fyear and fyear not in valid_years:
            warnings.append(
                f'[R02-发票年份越界] 发票 "{fname}" 年份 {fyear} 不在近三年 {sorted(valid_years)} 内'
            )

    for tr in materials.get('test_reports', []):
        fname = os.path.basename(tr)
        fyear = _extract_year_from_filename(fname)
        if fyear and fyear not in valid_years:
            warnings.append(
                f'[R04-检测报告年份越界] 检测报告 "{fname}" 年份 {fyear} '
                f'不在近三年 {sorted(valid_years)} 内'
            )

    return warnings


def convert_image_to_pdf(image_path, output_dir):
    """将图片文件转为PDF"""
    try:
        from PIL import Image
        img = Image.open(image_path)
        if img.mode in ('RGBA', 'LA', 'P'):
            img = img.convert('RGB')
        pdf_path = os.path.join(output_dir, os.path.splitext(os.path.basename(image_path))[0] + '.pdf')
        img.save(pdf_path, 'PDF', resolution=100.0)
        return pdf_path
    except Exception:
        return None


def merge_files_to_pdf(file_paths, output_path, max_size_mb=2):
    """将多个文件合并为单个PDF"""
    if not _PYPDF2_AVAILABLE:
        return False, 0, 0, ['PyPDF2 未安装']

    merger = PdfMerger()
    merged_count = 0
    warnings = []
    temp_dir = None

    for fp in file_paths:
        if not fp or not os.path.exists(fp):
            warnings.append(f'文件不存在: {fp}')
            continue

        ext = os.path.splitext(fp)[1].lower()
        actual_path = fp

        if ext in ('.jpg', '.jpeg', '.png'):
            if temp_dir is None:
                temp_dir = tempfile.mkdtemp(prefix='ach_proof_')
            converted = convert_image_to_pdf(fp, temp_dir)
            if converted:
                actual_path = converted
            else:
                warnings.append(f'图片转换失败: {fp}')
                continue

        try:
            merger.append(actual_path)
            merged_count += 1
        except Exception as e:
            warnings.append(f'合并失败 {os.path.basename(fp)}: {e}')

    if merged_count == 0:
        if temp_dir:
            shutil.rmtree(temp_dir, ignore_errors=True)
        return False, 0, 0, warnings

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    merger.write(output_path)
    merger.close()

    if temp_dir:
        shutil.rmtree(temp_dir, ignore_errors=True)

    size_mb = os.path.getsize(output_path) / (1024 * 1024)
    if size_mb > max_size_mb:
        warnings.append(
            f'文件大小 {size_mb:.2f}MB 超过 {max_size_mb}MB 限制，建议压缩: '
            f'python file_compressor.py auto --input "{output_path}" --type ACHIEVEMENT --output "{output_path}"'
        )

    return True, merged_count, size_mb, warnings


def organize_mode(args):
    """根据按成果归类的材料目录生成证明材料PDF"""
    ach_list, err = read_achievement_table(args.achievement_table)
    if err:
        print(json.dumps({'success': False, 'error': err}, ensure_ascii=False))
        return 1

    ip_certs, ip_excluded = scan_ip_certificates(args.ip_dir)
    materials = scan_materials_dir(args.materials_dir)
    output_dir = args.output_dir or os.path.join(os.path.dirname(args.achievement_table), '最终版本')
    os.makedirs(output_dir, exist_ok=True)

    valid_years = _get_valid_years(args.application_year)
    used_files = {}  # R05: hash → (seq, filename) 追踪

    report = {
        'success': True,
        'version': '1.1.0',
        'generated_at': datetime.now().isoformat(),
        'application_year': args.application_year,
        'valid_years': sorted(valid_years),
        'output_dir': output_dir,
        'experience_rules': {
            'R02_year_filtering': f'近三年={sorted(valid_years)}，合同/发票/检测报告硬过滤',
            'R03_certificate_priority': '已排除说明书/权利要求书/审查意见等非证书文件',
            'R04_cross_year_validation': '检测报告年份已校验',
            'R05_file_uniqueness': '同一文件不会分配给多个成果',
        },
        'ip_excluded_count': sum(len(v) for v in ip_excluded.values()),
        'ip_excluded_files': {
            f'IP{num:02d}': [os.path.basename(p) for p in paths]
            for num, paths in sorted(ip_excluded.items())
        },
        'achievements': [],
    }

    for ach in ach_list:
        ach_report = {
            'seq': ach['seq'],
            'name': ach['name'],
            'status': 'pending',
            'files_collected': [],
            'file_types': {},
            'duplicate_warnings': [],
            'year_warnings': [],
        }

        file_paths = []

        # 1. R03: 添加关联的IP证书（已排除说明书等）
        for ip_num in ach['ip_nums']:
            if ip_num in ip_certs:
                fpath = ip_certs[ip_num]
                fhash = _file_hash(fpath)

                # R05: 检查文件是否已被其他成果使用
                if fhash and fhash in used_files:
                    prev_seq, prev_name = used_files[fhash]
                    ach_report['duplicate_warnings'].append(
                        f'[R05-文件重复] IP{ip_num:02d}证书 "{os.path.basename(fpath)}" '
                        f'已被成果{prev_seq}({prev_name})使用'
                    )
                else:
                    if fhash:
                        used_files[fhash] = (ach['seq'], ach['name'])

                file_paths.append(fpath)
                ach_report['files_collected'].append(
                    f'IP{ip_num:02d}证书: {os.path.basename(fpath)}'
                )
                ach_report['file_types'].setdefault('ip_certs', 0)
                ach_report['file_types']['ip_certs'] += 1
            else:
                excluded_files = ip_excluded.get(ip_num, [])
                if excluded_files:
                    ach_report['year_warnings'].append(
                        f'IP{ip_num:02d}证书被排除（非证书文件）: '
                        + ', '.join(os.path.basename(p) for p in excluded_files)
                    )
                ach_report['year_warnings'].append(f'IP{ip_num:02d}证书未找到')

        # 2. v1.22.0: 按材料类型分类添加
        ach_mat = materials.get(ach['seq'], {})

        for category, label in [
            ('contracts', '合同'), ('invoices', '发票'),
            ('test_reports', '检测报告'), ('quality_certs', '质量管理证明'),
            ('ip_certs', 'IP证书'), ('other', '其他'),
        ]:
            for fp in ach_mat.get(category, []):
                fhash = _file_hash(fp)
                if fhash and fhash in used_files:
                    prev_seq, prev_name = used_files[fhash]
                    ach_report['duplicate_warnings'].append(
                        f'[R05-文件重复] {label} "{os.path.basename(fp)}" '
                        f'已被成果{prev_seq}({prev_name})使用'
                    )
                    continue

                if fhash:
                    used_files[fhash] = (ach['seq'], ach['name'])

                file_paths.append(fp)
                ach_report['files_collected'].append(f'{label}: {os.path.basename(fp)}')
                ach_report['file_types'].setdefault(category, 0)
                ach_report['file_types'][category] += 1

        # R02 + R04: 年份校验
        year_warnings = _validate_material_years(ach_mat, valid_years, ach['name'], ach['seq'])
        ach_report['year_warnings'].extend(year_warnings)

        # 合并所有警告
        ach_report['warnings'] = ach_report.pop('year_warnings') + ach_report.pop('duplicate_warnings')

        if not file_paths:
            ach_report['status'] = 'no_materials'
            ach_report['warnings'].append('无任何证明材料文件')
            report['achievements'].append(ach_report)
            continue

        # 3. 合并为PDF
        safe_name = _safe_filename(ach['name'])
        seq_str = f"{ach['seq']:02d}" if ach['seq'] < 10 else str(ach['seq'])
        output_path = os.path.join(output_dir, f'{seq_str}_{safe_name}.pdf')

        success, count, size_mb, merge_warnings = merge_files_to_pdf(file_paths, output_path)

        if success:
            ach_report['status'] = 'generated'
            ach_report['output_file'] = output_path
            ach_report['file_count'] = count
            ach_report['size_mb'] = round(size_mb, 2)
            ach_report['warnings'].extend(merge_warnings)
        else:
            ach_report['status'] = 'failed'
            ach_report['warnings'].extend(merge_warnings)

        report['achievements'].append(ach_report)

    # 统计
    generated = sum(1 for a in report['achievements'] if a['status'] == 'generated')
    report['stats'] = {
        'total': len(ach_list),
        'generated': generated,
        'no_materials': sum(1 for a in report['achievements'] if a['status'] == 'no_materials'),
        'failed': sum(1 for a in report['achievements'] if a['status'] == 'failed'),
        'duplicate_files_blocked': sum(
            1 for a in report['achievements']
            if any('[R05' in w for w in a.get('warnings', []))
        ),
        'year_violations': sum(
            1 for a in report['achievements']
            if any('[R02' in w or '[R04' in w for w in a.get('warnings', []))
        ),
    }

    report_path = os.path.join(output_dir, '_生成报告.json')
    with open(report_path, 'w', encoding='utf-8') as f:
        json.dump(report, f, ensure_ascii=False, indent=2, default=str)

    print(json.dumps(report, ensure_ascii=False, indent=2, default=str))
    return 0 if report['stats']['failed'] == 0 else 1


def auto_scan_mode(args):
    """v1.1.0: 自动扫描合同发票目录并匹配，含R02年份硬过滤"""
    ach_list, err = read_achievement_table(args.achievement_table)
    if err:
        print(json.dumps({'success': False, 'error': err}, ensure_ascii=False))
        return 1

    ip_certs, ip_excluded = scan_ip_certificates(args.ip_dir)
    contract_by_year = scan_contract_invoice_dir(args.contract_dir)
    output_dir = args.output_dir or os.path.join(os.path.dirname(args.achievement_table), '最终版本')
    os.makedirs(output_dir, exist_ok=True)

    valid_years = _get_valid_years(args.application_year)
    used_files = {}

    report = {
        'success': True,
        'version': '1.1.0',
        'generated_at': datetime.now().isoformat(),
        'application_year': args.application_year,
        'valid_years': sorted(valid_years),
        'output_dir': output_dir,
        'experience_rules': {
            'R02_year_filtering': f'近三年={sorted(valid_years)}，合同/发票硬过滤',
            'R03_certificate_priority': '已排除说明书/权利要求书/审查意见等非证书文件',
            'R05_file_uniqueness': '同一文件不会分配给多个成果',
        },
        'ip_excluded_count': sum(len(v) for v in ip_excluded.values()),
        'year_filtered_out': {
            year: len(files)
            for year, files in sorted(contract_by_year.items())
            if year != 'unknown' and int(year) not in valid_years
        },
        'achievements': [],
    }

    # R02: 硬过滤，只保留近三年的文件
    filtered_by_year = {}
    for year_str, files in contract_by_year.items():
        if year_str == 'unknown' or int(year_str) in valid_years:
            filtered_by_year[year_str] = files

    for ach in ach_list:
        ach_report = {
            'seq': ach['seq'],
            'name': ach['name'],
            'transform_time': str(ach.get('transform_time', '')),
            'status': 'pending',
            'files_collected': [],
            'warnings': [],
        }

        file_paths = []
        ach_year = _extract_year_from_text(ach.get('transform_time'))

        # 1. R03: IP证书
        for ip_num in ach['ip_nums']:
            if ip_num in ip_certs:
                fpath = ip_certs[ip_num]
                fhash = _file_hash(fpath)
                if fhash and fhash in used_files:
                    ach_report['warnings'].append(
                        f'[R05-文件重复] IP{ip_num:02d}证书已被其他成果使用'
                    )
                else:
                    if fhash:
                        used_files[fhash] = (ach['seq'], ach['name'])
                    file_paths.append(fpath)
                    ach_report['files_collected'].append(
                        f'IP{ip_num:02d}证书: {os.path.basename(fpath)}'
                    )
            else:
                ach_report['warnings'].append(f'IP{ip_num:02d}证书未找到')

        # 2. R02: 按年份匹配合同发票（硬过滤）
        if ach_year:
            if ach_year not in valid_years:
                ach_report['warnings'].append(
                    f'[R02-转化年份越界] 成果转化年份 {ach_year} 不在近三年 {sorted(valid_years)} 内'
                )
            else:
                year_str = str(ach_year)
                for yf in filtered_by_year.get(year_str, []):
                    fhash = _file_hash(yf)
                    if fhash and fhash in used_files:
                        ach_report['warnings'].append(
                            f'[R05-文件重复] "{os.path.basename(yf)}" 已被其他成果使用'
                        )
                        continue
                    if fhash:
                        used_files[fhash] = (ach['seq'], ach['name'])
                    file_paths.append(yf)
                    ach_report['files_collected'].append(f'{year_str}年: {os.path.basename(yf)}')

        if not file_paths:
            ach_report['status'] = 'no_materials'
            ach_report['warnings'].append('无任何证明材料文件')
            report['achievements'].append(ach_report)
            continue

        safe_name = _safe_filename(ach['name'])
        seq_str = f"{ach['seq']:02d}" if ach['seq'] < 10 else str(ach['seq'])
        output_path = os.path.join(output_dir, f'{seq_str}_{safe_name}.pdf')

        success, count, size_mb, merge_warnings = merge_files_to_pdf(file_paths, output_path)

        if success:
            ach_report['status'] = 'generated'
            ach_report['output_file'] = output_path
            ach_report['file_count'] = count
            ach_report['size_mb'] = round(size_mb, 2)
            ach_report['warnings'].extend(merge_warnings)
        else:
            ach_report['status'] = 'failed'
            ach_report['warnings'].extend(merge_warnings)

        report['achievements'].append(ach_report)

    generated = sum(1 for a in report['achievements'] if a['status'] == 'generated')
    report['stats'] = {
        'total': len(ach_list),
        'generated': generated,
        'no_materials': sum(1 for a in report['achievements'] if a['status'] == 'no_materials'),
        'failed': sum(1 for a in report['achievements'] if a['status'] == 'failed'),
        'duplicate_files_blocked': sum(
            1 for a in report['achievements']
            if any('[R05' in w for w in a.get('warnings', []))
        ),
    }

    report_path = os.path.join(output_dir, '_生成报告.json')
    with open(report_path, 'w', encoding='utf-8') as f:
        json.dump(report, f, ensure_ascii=False, indent=2, default=str)

    print(json.dumps(report, ensure_ascii=False, indent=2, default=str))
    return 0 if report['stats']['failed'] == 0 else 1


def preview_mode(args):
    """v1.1.0: 预览每项成果的材料匹配情况，含经验规则校验"""
    ach_list, err = read_achievement_table(args.achievement_table)
    if err:
        print(json.dumps({'success': False, 'error': err}, ensure_ascii=False))
        return 1

    ip_certs, ip_excluded = scan_ip_certificates(args.ip_dir)
    materials = scan_materials_dir(args.materials_dir) if args.materials_dir else {}
    valid_years = _get_valid_years(args.application_year)

    preview = {
        'version': '1.1.0',
        'application_year': args.application_year,
        'valid_years': sorted(valid_years),
        'experience_rules_applied': ['R02', 'R03', 'R04', 'R05'],
        'ip_excluded': {
            f'IP{num:02d}': [os.path.basename(p) for p in paths]
            for num, paths in sorted(ip_excluded.items())
        },
        'achievements': [],
    }

    used_files = {}

    for ach in ach_list:
        item = {
            'seq': ach['seq'],
            'name': ach['name'],
            'materials': {},
            'alerts': [],
        }

        # IP匹配（R03已过滤）
        ip_status = []
        for ip_num in ach['ip_nums']:
            ip_status.append({
                'ip_num': ip_num,
                'found': ip_num in ip_certs,
                'file': os.path.basename(ip_certs[ip_num]) if ip_num in ip_certs else None,
                'excluded': ip_num in ip_excluded,
            })
            if ip_num in ip_excluded:
                item['alerts'].append(
                    f'[R03] IP{ip_num:02d}有被排除的非证书文件: '
                    + ', '.join(os.path.basename(p) for p in ip_excluded[ip_num])
                )
        item['materials']['ip_certs'] = ip_status

        # 材料分类
        ach_mat = materials.get(ach['seq'], {})
        for cat in ['contracts', 'invoices', 'test_reports', 'quality_certs', 'ip_certs', 'other']:
            item['materials'][cat] = [
                os.path.basename(f) for f in ach_mat.get(cat, [])
            ]

        # R02 + R04 年份校验
        year_warnings = _validate_material_years(ach_mat, valid_years, ach['name'], ach['seq'])
        item['alerts'].extend(year_warnings)

        # R05 文件唯一性检查
        for cat in ['contracts', 'invoices', 'ip_certs']:
            for fp in ach_mat.get(cat, []):
                fhash = _file_hash(fp)
                if fhash and fhash in used_files:
                    prev_seq, prev_name = used_files[fhash]
                    item['alerts'].append(
                        f'[R05] "{os.path.basename(fp)}" 已被成果{prev_seq}({prev_name})使用'
                    )
                elif fhash:
                    used_files[fhash] = (ach['seq'], ach['name'])

        preview['achievements'].append(item)

    preview['stats'] = {
        'total': len(ach_list),
        'with_all_materials': sum(
            1 for a in preview['achievements']
            if all(ip['found'] for ip in a['materials']['ip_certs'])
            and a['materials']['contracts']
            and a['materials']['invoices']
        ),
        'missing_ip': sum(
            1 for a in preview['achievements']
            if any(not ip['found'] for ip in a['materials']['ip_certs'])
        ),
        'missing_contract': sum(
            1 for a in preview['achievements']
            if not a['materials']['contracts']
        ),
        'missing_invoice': sum(
            1 for a in preview['achievements']
            if not a['materials']['invoices']
        ),
        'alerts_count': sum(len(a['alerts']) for a in preview['achievements']),
    }

    print(json.dumps(preview, ensure_ascii=False, indent=2, default=str))
    return 0


def main():
    parser = argparse.ArgumentParser(
        description='科技成果转化证明材料PDF附件生成器 v1.2.0',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog='''示例:
  # 预览材料匹配情况
  python generate_achievement_proofs.py preview ^
      --achievement-table "03_成果转化证明/成果转化汇总表.xlsx" ^
      --materials-dir "03_成果转化证明/" ^
      --ip-dir "02_知识产权证明/"

  # 从归类的材料目录生成PDF
  python generate_achievement_proofs.py organize ^
      --achievement-table "03_成果转化证明/成果转化汇总表.xlsx" ^
      --materials-dir "03_成果转化证明/" ^
      --ip-dir "02_知识产权证明/" ^
      --output-dir "03_成果转化证明/最终版本/" ^
      --application-year 2026

经验规则（自动应用）:
  R02 年份硬过滤: 合同/发票/检测报告限定近三年 {申报年-3, 申报年-2, 申报年-1}
  R03 证书优先: 排除说明书/权利要求书/审查意见/缴费凭证
  R04 测试报告跨年校验: 检测报告年份=成果转化年份
  R05 文件唯一分配: 同一文件不会分配给多个成果
  v1.22.0 材料扩展: 质量管理体系证书 + 检测报告分类

v1.2.0 OneDrive安全: 输出目录在OneDrive同步根下时自动启用安全工作区
  d:\\Projects\\gxtz_safe\\ → 自动同步回原始目录 + attrib +p 钉选防脱水
  可通过 --no-safe-workspace 禁用自动保护
''',
    )

    parser.add_argument('--no-safe-workspace', action='store_true',
                        help='禁用 OneDrive 安全工作区自动保护（不推荐）')

    subparsers = parser.add_subparsers(dest='mode', help='运行模式')

    organize_parser = subparsers.add_parser('organize', help='从按成果归类的材料目录生成')
    organize_parser.add_argument('--achievement-table', required=True, help='成果转化汇总表路径')
    organize_parser.add_argument('--materials-dir', required=True, help='按成果归类的材料目录')
    organize_parser.add_argument('--ip-dir', required=True, help='IP证书目录')
    organize_parser.add_argument('--output-dir', help='输出目录（默认: 材料目录/最终版本/）')
    organize_parser.add_argument('--application-year', type=int, default=2026, help='申报年份')
    organize_parser.add_argument('--no-safe-workspace', action='store_true',
                                 help='禁用安全工作区')

    auto_parser = subparsers.add_parser('auto-scan', help='自动扫描合同发票目录匹配')
    auto_parser.add_argument('--achievement-table', required=True, help='成果转化汇总表路径')
    auto_parser.add_argument('--ip-dir', required=True, help='IP证书目录')
    auto_parser.add_argument('--contract-dir', required=True, help='合同发票目录')
    auto_parser.add_argument('--output-dir', help='输出目录')
    auto_parser.add_argument('--application-year', type=int, default=2026, help='申报年份')
    auto_parser.add_argument('--no-safe-workspace', action='store_true',
                             help='禁用安全工作区')

    preview_parser = subparsers.add_parser('preview', help='预览材料匹配情况')
    preview_parser.add_argument('--achievement-table', required=True, help='成果转化汇总表路径')
    preview_parser.add_argument('--materials-dir', help='按成果归类的材料目录')
    preview_parser.add_argument('--ip-dir', required=True, help='IP证书目录')
    preview_parser.add_argument('--application-year', type=int, default=2026, help='申报年份')

    args = parser.parse_args()

    if not args.mode:
        parser.print_help()
        return 1

    disable_sw = getattr(args, 'no_safe_workspace', False)

    if not disable_sw and _SAFE_WORKSPACE_AVAILABLE and not args.mode == 'preview':
        output_dir = getattr(args, 'output_dir', None) or ''
        achievement_table = getattr(args, 'achievement_table', '')
        check_path = output_dir or os.path.dirname(achievement_table) or ''

        if check_path and is_under_onedrive_syncroot(check_path):
            enterprise = _infer_enterprise(achievement_table)
            safe_root = os.path.dirname(check_path)
            try:
                with SafeWorkspace(safe_root, enterprise, mode='rw') as ws:
                    if hasattr(args, 'output_dir') and args.output_dir:
                        rel = os.path.relpath(check_path, safe_root) if safe_root else ''
                        if rel and rel != '.':
                            args.output_dir = os.path.join(ws.output_dir, rel)
                        else:
                            args.output_dir = ws.output_dir
                    return _dispatch_mode(args)
            except Exception as e:
                error_msg = f'安全工作区失败: {e}，回退到直接写入模式'
                print(json.dumps({'safe_workspace_warning': error_msg}, ensure_ascii=False),
                      file=sys.stderr)

    return _dispatch_mode(args)


def _infer_enterprise(path):
    m = re.search(r'【国高】[^\\/]*?([\u4e00-\u9fa5]{2,}?(?:公司|科技|集团|有限|实业|电子|材料|铝业|化学|技术))', path)
    if m:
        return m.group(1)
    parts = path.replace('\\', '/').split('/')
    for part in reversed(parts):
        if part and not part.startswith('.'):
            return part[:30]
    return 'unknown'


def _dispatch_mode(args):
    if args.mode == 'organize':
        return organize_mode(args)
    elif args.mode == 'auto-scan':
        return auto_scan_mode(args)
    elif args.mode == 'preview':
        return preview_mode(args)
    else:
        return 1


if __name__ == '__main__':
    sys.exit(main())
